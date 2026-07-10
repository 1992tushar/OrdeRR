"""
Vasy → OrdeRR import framework (Phase 2).

Vasy ERP is the source of truth for money; these importers mirror its Excel
exports into read-only OrdeRR tables (CustomerReceipt, OutstandingSnapshot) and
record an ImportLog per run. Idempotent: re-importing the same/overlapping file
upserts rather than duplicates.

Join to OrdeRR customers is phone-first (last-10 digits, where the export has a
phone) then normalized name — matching the reality confirmed against the real
2026-07-10 exports. Unmatched rows are recorded with customer_id = NULL
(unattributed bucket), never dropped.

Reuses: customer_service._to_amount / normalize_phone, and
customer_service.import_customers_from_xlsx for the customer/outstanding upsert
(so there is ONE customer-upsert code path).
"""
import io
import re
from datetime import date, datetime

import openpyxl
from sqlalchemy import func
from sqlalchemy.orm import Session

from orderr_core.dates import get_current_business_date

from orderr_core.models.customer import Customer
from orderr_core.models.customer_receipt import CustomerReceipt
from orderr_core.models.outstanding_snapshot import OutstandingSnapshot
from orderr_core.models.import_log import ImportLog
from orderr_core.models.vasy_invoice import VasyInvoice, VasyInvoiceItem
from orderr_core.models.vasy_purchase import VasyPurchase, VasyPurchaseItem
from orderr_core.models.vasy_expense import VasyExpense
from orderr_core.models.vasy_payment import VasyPayment
from orderr_core.models.vasy_supplier_bill import VasySupplierBill
from orderr_core.services.customer_service import (
    _to_amount, normalize_phone, import_customers_from_xlsx,
)
from orderr_core.services.template_parser import ERP_ITEMS


# ── shared helpers ──────────────────────────────────────────────────────────

def normalize_name(s) -> str:
    """Party-name join key: uppercase, strip every non-alphanumeric char.
    Robust to spacing / punctuation / case drift between exports."""
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


def _phone_last10(s) -> str:
    """Last 10 digits of a phone string ('91-9850410033' → '9850410033')."""
    digits = re.sub(r"\D", "", str(s or ""))
    return digits[-10:] if len(digits) >= 10 else ""


def _parse_date(v):
    """Parse a Vasy date cell: 'DD/MM/YYYY' string or a real date/datetime."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _find_header(ws, required_labels, max_scan=15):
    """Locate the header row and map wanted columns by label (resilient to
    column re-ordering / title rows). `required_labels` maps a logical key to a
    list of accepted lowercased header texts. Returns (header_row_index_0based,
    {key: col_idx}). Raises ValueError if a required column is missing."""
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        labels = [str(c).strip().lower() if c is not None else "" for c in row]
        colmap = {}
        for key, accepted in required_labels.items():
            for i, lbl in enumerate(labels):
                if lbl in accepted:
                    colmap[key] = i
                    break
        if ("receipt_no" in colmap or ("closing" in colmap and "party" in colmap)
                or ("voucher" in colmap and "party" in colmap)
                or ("bill" in colmap and "party" in colmap)
                or ("bill" in colmap and "vendor" in colmap)
                or "expense_no" in colmap or "payment_no" in colmap):
            # header row found (heuristic: a distinctive column is present)
            return r_idx, colmap
    raise ValueError("Could not locate the header row in the export.")


def _erp_code_to_name():
    """{Vasy erp_code → erp display name} from the SKU catalog."""
    out = {}
    for _name, item in ERP_ITEMS.items():
        code = item.get("erp_code")
        if code:
            out[str(code).strip()] = item.get("erp_name")
    return out


def _build_customer_lookup(db: Session):
    """Return (by_phone, by_name) dicts → customer_id.

    by_name is keyed by normalized name. It's built from BOTH the customer
    master AND the latest OutstandingSnapshot's clean Vasy party names. The
    latter matters because the customer master sometimes stores glued
    Name+Company names ("CHAINESE KATTA  CHAINESE KATTA") that don't match the
    single Vasy name on invoices/receipts — but the outstanding import already
    linked each clean party name to the right customer (phone-anchored), so
    that mapping recovers the join. Customer-master names take precedence.
    """
    by_phone, by_name = {}, {}
    for c in db.query(Customer.id, Customer.phone_number, Customer.restaurant_name).all():
        if c.phone_number:
            p = _phone_last10(c.phone_number)
            if p:
                by_phone.setdefault(p, c.id)
        if c.restaurant_name:
            by_name.setdefault(normalize_name(c.restaurant_name), c.id)

    # augment with the latest outstanding snapshot's clean party→customer links
    latest = db.query(func.max(OutstandingSnapshot.snapshot_date)).scalar()
    if latest is not None:
        for party_key, cid in (db.query(OutstandingSnapshot.party_key, OutstandingSnapshot.customer_id)
                               .filter(OutstandingSnapshot.snapshot_date == latest,
                                       OutstandingSnapshot.customer_id != None).all()):  # noqa: E711
            if party_key:
                by_name.setdefault(party_key, cid)
    return by_phone, by_name


# ── P2-2 receipts import ────────────────────────────────────────────────────

def import_receipts(db: Session, file_bytes: bytes, source_file: str = None) -> dict:
    """Import a Vasy receipt export into CustomerReceipt.

    Upsert on Receipt No.; join party → customer by normalized name (receipts
    carry no phone). Unmatched → customer_id NULL (unattributed). Idempotent.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"Could not read the receipt Excel file: {e}")
    ws = wb.active

    labels = {
        "receipt_no": ("receipt no.", "receipt no", "receipt number", "receipt"),
        "party": ("party name", "party", "name", "customer name"),
        "mode": ("mode", "payment mode"),
        "date": ("date", "receipt date"),
        "amount": ("amount", "amount (inr)", "received amount"),
        "status": ("status",),
        "created_by": ("created by",),
    }
    header_idx, col = _find_header(ws, labels)
    if "receipt_no" not in col or "party" not in col:
        raise ValueError("Receipt export must have 'Receipt No.' and 'Party Name' columns.")

    _, by_name = _build_customer_lookup(db)

    # preload existing receipts for idempotent upsert
    existing = {r.receipt_no: r for r in db.query(CustomerReceipt).all()}

    created = updated = unmatched = 0
    seen = set()

    def cell(row, key):
        i = col.get(key)
        return row[i] if (i is not None and i < len(row)) else None

    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        if not row:
            continue
        rno = cell(row, "receipt_no")
        rno = str(rno).strip() if rno is not None else ""
        if _is_total_row(rno) or rno in seen:  # skip footer Total / blank / dup
            continue
        seen.add(rno)

        party = str(cell(row, "party") or "").strip()
        mode = str(cell(row, "mode") or "").strip().lower() or None
        amount = _to_amount(cell(row, "amount"))
        rdate = _parse_date(cell(row, "date"))
        status = str(cell(row, "status") or "").strip().lower() or None
        created_by = str(cell(row, "created_by") or "").strip() or None

        cust_id = by_name.get(normalize_name(party))
        if cust_id is None:
            unmatched += 1

        rec = existing.get(rno)
        if rec is None:
            db.add(CustomerReceipt(
                receipt_no=rno, party_name=party, customer_id=cust_id,
                mode=mode, amount=amount, receipt_date=rdate,
                status=status, created_by=created_by,
            ))
            created += 1
        else:
            rec.party_name = party
            rec.customer_id = cust_id
            rec.mode = mode
            rec.amount = amount
            rec.receipt_date = rdate
            rec.status = status
            rec.created_by = created_by
            updated += 1

    log = ImportLog(entity="receipts", source_file=source_file,
                    rows_total=created + updated, created=created,
                    updated=updated, unmatched=unmatched)
    db.add(log)
    db.commit()
    wb.close()

    return {
        "entity": "receipts",
        "rows": created + updated,
        "created": created,
        "updated": updated,
        "unmatched": unmatched,
        "matched": created + updated - unmatched,
    }


# ── P2-3 outstanding import (customer refresh + daily snapshot) ─────────────

def import_outstanding(db: Session, file_bytes: bytes, snapshot_date: date = None,
                       source_file: str = None) -> dict:
    """Import a Vasy customer-outstanding export.

    Two layers:
      1. Customer master + `customer.outstanding` refresh (+ phone backfill,
         create-if-missing) — delegated to the existing, tested
         import_customers_from_xlsx so there is ONE customer-upsert path.
      2. A daily OutstandingSnapshot per party (opening/debit/credit/closing),
         upserted on (party_key, snapshot_date) so a same-day re-import updates
         rather than duplicates — this is what gives balances a history.

    Customer match is phone-first (Contact No. last-10) then normalized name.
    Unmatched parties still get a snapshot with customer_id NULL.
    """
    snapshot_date = snapshot_date or get_current_business_date()

    # ── layer 1: customer master + outstanding refresh (commits internally) ──
    cust_summary = import_customers_from_xlsx(db, file_bytes)

    # ── layer 2: snapshots ──
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"Could not read the outstanding Excel file: {e}")
    ws = wb.active

    labels = {
        "party": ("party name", "party", "name", "customer name"),
        "contact": ("contact no.", "contact no", "contact", "phone", "mobile no.", "mobile"),
        "opening": ("opening balance", "opening"),
        "debit": ("debit",),
        "credit": ("credit",),
        "closing": ("closing", "closing balance", "outstanding", "balance"),
    }
    header_idx, col = _find_header(ws, labels)
    if "party" not in col or "closing" not in col:
        raise ValueError("Outstanding export must have 'Party Name' and 'Closing' columns.")

    by_phone, by_name = _build_customer_lookup(db)  # after layer 1 → includes new customers

    existing = {s.party_key: s for s in db.query(OutstandingSnapshot)
                .filter(OutstandingSnapshot.snapshot_date == snapshot_date).all()}

    created = updated = unmatched = 0
    seen = set()

    def cell(row, key):
        i = col.get(key)
        return row[i] if (i is not None and i < len(row)) else None

    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        if not row:
            continue
        party = str(cell(row, "party") or "").strip()
        if _is_total_row(party):        # skip footer Total / blank
            continue
        key = normalize_name(party)
        if not key or key in seen:
            continue
        seen.add(key)

        contact = cell(row, "contact")
        contact = str(contact).strip() if contact not in (None, "") else None
        opening = _to_amount(cell(row, "opening"))
        debit = _to_amount(cell(row, "debit"))
        credit = _to_amount(cell(row, "credit"))
        closing = _to_amount(cell(row, "closing"))

        # phone-first, then name
        cust_id = None
        p10 = _phone_last10(contact)
        if p10:
            cust_id = by_phone.get(p10)
        if cust_id is None:
            cust_id = by_name.get(key)
        if cust_id is None:
            unmatched += 1

        snap = existing.get(key)
        if snap is None:
            db.add(OutstandingSnapshot(
                customer_id=cust_id, party_name=party, party_key=key,
                contact_no=contact, opening_balance=opening, debit=debit,
                credit=credit, closing=closing, snapshot_date=snapshot_date,
            ))
            created += 1
        else:
            snap.customer_id = cust_id
            snap.party_name = party
            snap.contact_no = contact
            snap.opening_balance = opening
            snap.debit = debit
            snap.credit = credit
            snap.closing = closing
            updated += 1

    db.add(ImportLog(entity="outstanding", source_file=source_file,
                     rows_total=created + updated, created=created,
                     updated=updated, unmatched=unmatched,
                     notes=f"snapshot_date={snapshot_date.isoformat()}; "
                           f"customers +{cust_summary.get('created',0)}/"
                           f"~{cust_summary.get('updated',0)}"))
    db.commit()
    wb.close()

    return {
        "entity": "outstanding",
        "snapshot_date": snapshot_date.isoformat(),
        "rows": created + updated,
        "created": created,
        "updated": updated,
        "unmatched": unmatched,
        "matched": created + updated - unmatched,
        "customers": cust_summary,
    }


# ── P2-15 Vasy sales-invoice import (authoritative revenue) ────────────────

def import_sales_invoices(db: Session, file_bytes: bytes, source_file: str = None) -> dict:
    """Import a Vasy sales-invoice export (line-item level) into
    VasyInvoice + VasyInvoiceItem.

    Rows are grouped by Voucher No; the invoice header total = Σ line net.
    Upsert on voucher_no and REPLACE its lines (clean idempotency — the export
    always carries the full invoice). Join party → customer phone-first
    (Mobile No. where present) then normalized name; unmatched → NULL.
    """
    try:
        # NOT read_only: the Vasy sales export declares a bad sheet dimension,
        # which makes openpyxl's read-only iterator truncate rows to 1 column.
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Could not read the sales-invoice Excel file: {e}")
    ws = wb.active

    labels = {
        "voucher": ("voucher no", "voucher no.", "voucher number", "invoice no", "invoice no."),
        "date": ("date", "invoice date"),
        "party": ("party name", "party", "customer name", "name"),
        "mobile": ("mobile no.", "mobile no", "mobile", "contact no.", "phone"),
        "category": ("category name", "category"),
        "item_code": ("item code", "item", "code"),
        "qty": ("qty", "quantity"),
        "net": ("net amount", "amount", "net"),
        "branch": ("branch",),
        "address": ("address",),
    }
    header_idx, col = _find_header(ws, labels)
    if "voucher" not in col or "party" not in col:
        raise ValueError("Sales-invoice export must have 'Voucher No' and 'Party Name' columns.")

    by_phone, by_name = _build_customer_lookup(db)
    erp = _erp_code_to_name()

    def cell(row, key):
        i = col.get(key)
        return row[i] if (i is not None and i < len(row)) else None

    # group lines by voucher_no
    invoices = {}
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        if not row:
            continue
        vno = cell(row, "voucher")
        vno = str(vno).strip() if vno is not None else ""
        if _is_total_row(vno):          # skip footer Total row / blank voucher
            continue
        inv = invoices.get(vno)
        if inv is None:
            party = str(cell(row, "party") or "").strip()
            mobile = cell(row, "mobile")
            inv = invoices[vno] = {
                "date": _parse_date(cell(row, "date")),
                "party": party,
                "mobile": str(mobile).strip() if mobile not in (None, "", "null", "-") else None,
                "branch": str(cell(row, "branch") or "").strip() or None,
                "address": str(cell(row, "address") or "").strip() or None,
                "lines": [],
            }
        code = str(cell(row, "item_code") or "").strip()
        qty = _to_amount(cell(row, "qty"))
        net = _to_amount(cell(row, "net"))
        inv["lines"].append({
            "item_code": code or None,
            "erp_name": erp.get(code) or (str(cell(row, "category") or "").strip() or None),
            "category": str(cell(row, "category") or "").strip() or None,
            "qty": qty, "net": net,
        })
    wb.close()

    # self-heal: remove any footer-artifact invoice left by the earlier import
    # bug (a "Total" row imported as a fake invoice with the grand total).
    for bad in db.query(VasyInvoice).filter(
            func.upper(VasyInvoice.voucher_no).in_(("TOTAL", ""))).all():
        db.delete(bad)

    existing = {v.voucher_no: v for v in db.query(VasyInvoice).all()
                if not _is_total_row(v.voucher_no)}
    created = updated = unmatched = auto_created = 0
    total_amount = 0.0
    new_by_key = {}   # party_key -> id for customers auto-created within this run

    for vno, inv in invoices.items():
        total = sum(float(l["net"]) for l in inv["lines"])
        total_amount += total
        key = normalize_name(inv["party"])
        cust_id = None
        p10 = _phone_last10(inv["mobile"])
        if p10:
            cust_id = by_phone.get(p10)
        if cust_id is None:
            cust_id = by_name.get(key)
        if cust_id is None and total > 0 and key:
            # A billed party with no customer record yet — e.g. a zero-outstanding
            # customer that never shows up in the outstanding export. Create one
            # from the clean Vasy name so its sales attribute. Gated on total>0 so
            # internal ₹0 accounts (PLANT WASTAGE, WORKERS DAILY FOOD) are NOT
            # turned into customers. Matching itself stays exact — never fuzzy.
            cust_id = new_by_key.get(key)
            if cust_id is None:
                newc = Customer(
                    restaurant_name=(inv["party"] or "").strip(),
                    phone_number=None,
                    onboarding_status="active",
                    is_active=True,
                    is_daily_order_customer=False,
                )
                db.add(newc)
                db.flush()   # assign PK without committing the outer transaction
                cust_id = newc.id
                new_by_key[key] = cust_id
                by_name.setdefault(key, cust_id)
                auto_created += 1
        if cust_id is None:
            unmatched += 1

        rec = existing.get(vno)
        if rec is None:
            rec = VasyInvoice(voucher_no=vno)
            db.add(rec)
            created += 1
        else:
            rec.items.clear()   # replace lines (cascade delete-orphan)
            updated += 1
        rec.invoice_date = inv["date"]
        rec.party_name = inv["party"]
        rec.party_key = key
        rec.customer_id = cust_id
        rec.total = round(total, 2)
        rec.item_count = len(inv["lines"])
        rec.branch = inv["branch"]
        rec.address = inv["address"]
        for l in inv["lines"]:
            rec.items.append(VasyInvoiceItem(
                item_code=l["item_code"], erp_name=l["erp_name"], category=l["category"],
                qty=l["qty"], net_amount=l["net"],
            ))

    db.add(ImportLog(entity="sales_invoices", source_file=source_file,
                     rows_total=created + updated, created=created,
                     updated=updated, unmatched=unmatched,
                     notes=f"lines={sum(len(i['lines']) for i in invoices.values())}; "
                           f"total={round(total_amount, 2)}; auto_created={auto_created}"))
    db.commit()

    return {
        "entity": "sales_invoices",
        "rows": created + updated,
        "invoices": created + updated,
        "created": created,
        "updated": updated,
        "unmatched": unmatched,
        "matched": created + updated - unmatched,
        "customers_created": auto_created,
        "total_fmt": _fmt_inr_local(total_amount),
    }


def _fmt_inr_local(amount):
    from orderr_core.services.analytics_service import fmt_inr
    return fmt_inr(amount)


def _is_total_row(v):
    """True if a key cell is blank or the export's footer 'Total' marker."""
    s = str(v or "").strip().lower()
    return s in ("", "total")


# ── P3-10 Purchases import (COGS; line-item, grouped by Bill No) ───────────

def import_purchases(db: Session, file_bytes: bytes, source_file: str = None) -> dict:
    """Import a Vasy purchase export into VasyPurchase + VasyPurchaseItem.
    Grouped by Bill No; total = Σ line amount. Upsert on bill_no, replace lines.
    Supplier party stored raw (no supplier master). Skips footer Total row."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Could not read the purchase Excel file: {e}")
    ws = wb.active
    labels = {
        "bill": ("bill no", "bill no.", "bill number"),
        "date": ("bill date", "date"),
        "party": ("party name", "party", "supplier", "name"),
        "hsn": ("hsn",),
        "product": ("product name", "product", "item name"),
        "item_code": ("item code", "code"),
        "rate": ("rate",),
        "qty": ("qty", "quantity"),
        "amount": ("total amount", "amount", "net amount"),
    }
    header_idx, col = _find_header(ws, labels)
    if "bill" not in col or "amount" not in col:
        raise ValueError("Purchase export must have 'Bill No' and 'Total Amount' columns.")

    def cell(row, key):
        i = col.get(key)
        return row[i] if (i is not None and i < len(row)) else None

    bills = {}
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        if not row:
            continue
        bno = cell(row, "bill")
        bno = str(bno).strip() if bno is not None else ""
        if _is_total_row(bno):
            continue
        b = bills.get(bno)
        if b is None:
            b = bills[bno] = {"date": _parse_date(cell(row, "date")),
                              "party": str(cell(row, "party") or "").strip(), "lines": []}
        b["lines"].append({
            "product": str(cell(row, "product") or "").strip() or None,
            "item_code": str(cell(row, "item_code") or "").strip() or None,
            "hsn": str(cell(row, "hsn") or "").strip() or None,
            "rate": _to_amount(cell(row, "rate")), "qty": _to_amount(cell(row, "qty")),
            "amount": _to_amount(cell(row, "amount")),
        })
    wb.close()

    existing = {p.bill_no: p for p in db.query(VasyPurchase).all()}
    created = updated = 0
    total_amount = 0.0
    for bno, b in bills.items():
        total = sum(float(l["amount"]) for l in b["lines"])
        total_amount += total
        rec = existing.get(bno)
        if rec is None:
            rec = VasyPurchase(bill_no=bno); db.add(rec); created += 1
        else:
            rec.items.clear(); updated += 1
        rec.bill_date = b["date"]
        rec.party_name = b["party"]
        rec.party_key = normalize_name(b["party"])
        rec.total = round(total, 2)
        rec.item_count = len(b["lines"])
        for l in b["lines"]:
            rec.items.append(VasyPurchaseItem(product_name=l["product"], item_code=l["item_code"],
                                              hsn=l["hsn"], rate=l["rate"], qty=l["qty"], amount=l["amount"]))
    db.add(ImportLog(entity="purchases", source_file=source_file, rows_total=created + updated,
                     created=created, updated=updated, unmatched=0,
                     notes=f"lines={sum(len(b['lines']) for b in bills.values())}; total={round(total_amount,2)}"))
    db.commit()
    return {"entity": "purchases", "rows": created + updated, "bills": created + updated,
            "created": created, "updated": updated, "unmatched": 0,
            "total_fmt": _fmt_inr_local(total_amount)}


# ── P3-10 Expenses import (opex; header level) ─────────────────────────────

def import_expenses(db: Session, file_bytes: bytes, source_file: str = None) -> dict:
    """Import a Vasy expense export into VasyExpense (upsert on expense_no)."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Could not read the expense Excel file: {e}")
    ws = wb.active
    labels = {
        "expense_no": ("expense no.", "expense no", "expense number"),
        "date": ("expense date", "date"),
        "party": ("party name", "party", "name"),
        "total": ("total",),
        "paid": ("paid",),
        "unpaid": ("unpaid",),
    }
    header_idx, col = _find_header(ws, labels)
    if "expense_no" not in col:
        raise ValueError("Expense export must have an 'Expense No.' column.")

    def cell(row, key):
        i = col.get(key)
        return row[i] if (i is not None and i < len(row)) else None

    existing = {e.expense_no: e for e in db.query(VasyExpense).all()}
    created = updated = 0
    total_amount = 0.0
    seen = set()
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        if not row:
            continue
        eno = cell(row, "expense_no")
        eno = str(eno).strip() if eno is not None else ""
        if _is_total_row(eno) or eno in seen:
            continue
        seen.add(eno)
        total = _to_amount(cell(row, "total"))
        total_amount += float(total)
        rec = existing.get(eno)
        if rec is None:
            rec = VasyExpense(expense_no=eno); db.add(rec); created += 1
        else:
            updated += 1
        rec.expense_date = _parse_date(cell(row, "date"))
        rec.party_name = str(cell(row, "party") or "").strip()
        rec.party_key = normalize_name(rec.party_name)
        rec.total = total
        rec.paid = _to_amount(cell(row, "paid"))
        rec.unpaid = _to_amount(cell(row, "unpaid"))
    db.add(ImportLog(entity="expenses", source_file=source_file, rows_total=created + updated,
                     created=created, updated=updated, unmatched=0, notes=f"total={round(total_amount,2)}"))
    db.commit()
    return {"entity": "expenses", "rows": created + updated, "created": created,
            "updated": updated, "unmatched": 0, "total_fmt": _fmt_inr_local(total_amount)}


# ── P3-10 Payments import (money out; header level) ────────────────────────

def import_payments(db: Session, file_bytes: bytes, source_file: str = None) -> dict:
    """Import a Vasy payment export (money out) into VasyPayment (upsert on
    payment_no). Distinct from CustomerReceipt despite the shared PAY prefix."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Could not read the payment Excel file: {e}")
    ws = wb.active
    labels = {
        "payment_no": ("payment no", "payment no.", "payment number"),
        "party": ("party name", "party", "name"),
        "mode": ("payment mode", "mode"),
        "date": ("date", "payment date"),
        "amount": ("amount",),
        "status": ("status",),
    }
    header_idx, col = _find_header(ws, labels)
    if "payment_no" not in col:
        raise ValueError("Payment export must have a 'Payment No' column.")

    def cell(row, key):
        i = col.get(key)
        return row[i] if (i is not None and i < len(row)) else None

    existing = {p.payment_no: p for p in db.query(VasyPayment).all()}
    created = updated = 0
    total_amount = 0.0
    seen = set()
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        if not row:
            continue
        pno = cell(row, "payment_no")
        pno = str(pno).strip() if pno is not None else ""
        if _is_total_row(pno) or pno in seen:
            continue
        seen.add(pno)
        amt = _to_amount(cell(row, "amount"))
        total_amount += float(amt)
        rec = existing.get(pno)
        if rec is None:
            rec = VasyPayment(payment_no=pno); db.add(rec); created += 1
        else:
            updated += 1
        rec.party_name = str(cell(row, "party") or "").strip()
        rec.party_key = normalize_name(rec.party_name)
        rec.mode = (str(cell(row, "mode") or "").strip().lower() or None)
        rec.payment_date = _parse_date(cell(row, "date"))
        rec.amount = amt
        rec.status = (str(cell(row, "status") or "").strip().lower() or None)
    db.add(ImportLog(entity="payments", source_file=source_file, rows_total=created + updated,
                     created=created, updated=updated, unmatched=0, notes=f"total={round(total_amount,2)}"))
    db.commit()
    return {"entity": "payments", "rows": created + updated, "created": created,
            "updated": updated, "unmatched": 0, "total_fmt": _fmt_inr_local(total_amount)}


# ── Supplier bills import (accounts payable; bill-level, upsert on Bill No) ─

def import_supplier_bills(db: Session, file_bytes: bytes, source_file: str = None) -> dict:
    """Import a Vasy Supplier Bill List into VasySupplierBill (upsert on
    bill_no). Bill-level AP with paid/due/due-date/status → enables true AP
    aging. Vendor stored raw. Skips footer Total row."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Could not read the supplier-outstanding Excel file: {e}")
    ws = wb.active
    labels = {
        "status": ("status",),
        "bill": ("bill no", "bill no.", "bill number"),
        "bill_date": ("bill date",),
        "vendor": ("vendor", "party name", "supplier", "party"),
        "amount": ("amount",),
        "paid": ("paid amount", "paid"),
        "due": ("due amount", "due", "outstanding"),
        "tax": ("tax amount", "tax"),
        "due_date": ("due date",),
    }
    header_idx, col = _find_header(ws, labels)
    if "bill" not in col or "due" not in col:
        raise ValueError("Supplier outstanding export must have 'Bill No' and 'Due Amount' columns.")

    def cell(row, key):
        i = col.get(key)
        return row[i] if (i is not None and i < len(row)) else None

    existing = {b.bill_no: b for b in db.query(VasySupplierBill).all()}
    created = updated = 0
    total_due = 0.0
    seen = set()
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        if not row:
            continue
        bno = cell(row, "bill")
        bno = str(bno).strip() if bno is not None else ""
        if _is_total_row(bno) or bno in seen:
            continue
        seen.add(bno)
        due = _to_amount(cell(row, "due"))
        total_due += float(due)
        rec = existing.get(bno)
        if rec is None:
            rec = VasySupplierBill(bill_no=bno); db.add(rec); created += 1
        else:
            updated += 1
        rec.bill_date = _parse_date(cell(row, "bill_date"))
        rec.due_date = _parse_date(cell(row, "due_date"))
        rec.vendor = str(cell(row, "vendor") or "").strip()
        rec.vendor_key = normalize_name(rec.vendor)
        rec.amount = _to_amount(cell(row, "amount"))
        rec.paid = _to_amount(cell(row, "paid"))
        rec.due = due
        rec.tax = _to_amount(cell(row, "tax"))
        rec.status = (str(cell(row, "status") or "").strip().lower() or None)
    db.add(ImportLog(entity="supplier_bills", source_file=source_file, rows_total=created + updated,
                     created=created, updated=updated, unmatched=0, notes=f"due_total={round(total_due,2)}"))
    db.commit()
    return {"entity": "supplier_bills", "rows": created + updated, "created": created,
            "updated": updated, "unmatched": 0, "total_fmt": _fmt_inr_local(total_due)}
