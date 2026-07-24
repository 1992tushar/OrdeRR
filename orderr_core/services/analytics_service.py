"""
Analytics service — read-only aggregations for the Analytics dashboard (Phase 1).

Phase 1 builds entirely on data already in OrdeRR (orders, invoices, actuals);
it has NO external dependency. Per the initiative's architecture principle,
Vasy ERP remains the money source-of-truth — so "sales" figures here are
OrdeRR's own (operational) *invoices*, and will be reconciled against / replaced
by mirrored Vasy revenue in Phase 2. Labels in the UI say so.

Conventions reused: orderr_core.dates.* for the business-date basis,
utils.fmt_qty for quantities. Money is rendered ₹ with en-IN (Indian) digit
grouping to match the app standard (client screens use toLocaleString('en-IN')).
All figures exclude cancelled orders and void invoices.
"""
import bisect
import calendar
from datetime import date, datetime, timedelta
from statistics import median

from sqlalchemy import func
from sqlalchemy.orm import Session

from orderr_core import config as _config
from orderr_core.models.order import Order
from orderr_core.models.invoice import Invoice, InvoiceItem
from orderr_core.models.customer import Customer
from orderr_core.models.salesperson import Salesperson
from orderr_core.models.actuals import OrderItemActual
from orderr_core.models.customer_receipt import CustomerReceipt
from orderr_core.models.outstanding_snapshot import OutstandingSnapshot
from orderr_core.models.bad_debt import BadDebt
from orderr_core.models.vasy_invoice import VasyInvoice
from orderr_core.models.vasy_purchase import VasyPurchase
from orderr_core.models.vasy_expense import VasyExpense
from orderr_core.models.vasy_payment import VasyPayment
from orderr_core.models.vasy_supplier_bill import VasySupplierBill
from orderr_core.services.template_parser import erp_display_name, ERP_ITEMS
from orderr_core.utils import safe_list, fmt_qty


# ── formatting helpers (Indian grouping; stdlib-only, locale-independent) ──────

def _group_indian(int_str: str) -> str:
    """Group an integer string Indian-style: 1234567 -> '12,34,567'."""
    if len(int_str) <= 3:
        return int_str
    head, tail = int_str[:-3], int_str[-3:]
    # insert a comma every 2 digits from the right of the head
    parts = []
    while len(head) > 2:
        parts.insert(0, head[-2:])
        head = head[:-2]
    parts.insert(0, head)
    return ",".join(parts) + "," + tail


def fmt_inr(amount) -> str:
    """Format a rupee amount as '₹1,23,456' (no paise for whole numbers,
    2 decimals otherwise), with Indian digit grouping."""
    try:
        n = float(amount or 0)
    except (TypeError, ValueError):
        return "₹0"
    neg = n < 0
    n = round(abs(n), 2)   # round FIRST so float noise (…9999) doesn't truncate a rupee
    if n == int(n):
        body = _group_indian(str(int(n)))
    else:
        whole = int(n)
        frac = round(n - whole, 2)
        body = _group_indian(str(whole)) + f"{frac:.2f}"[1:]  # keep ".xx"
    return ("-₹" if neg else "₹") + body


def fmt_kg(qty) -> str:
    """Format a kg quantity with Indian grouping, dropping trailing .0."""
    try:
        n = float(qty or 0)
    except (TypeError, ValueError):
        return "0"
    if n == int(n):
        return _group_indian(str(int(n)))
    return _group_indian(str(int(n))) + f"{round(n - int(n), 1):.1f}"[1:]


# ── period helpers ─────────────────────────────────────────────────────────

def _period_bounds(today: date, day: str = "today"):
    """Return (key, label, sublabel, start_date, end_date) for the pulse
    periods, all inclusive of `today`.

      today/yesterday — a single business day (toggled via `day`)
      week            — rolling last 7 days (today-6 .. today)
      month           — calendar month-to-date (1st of month .. today)

    The first card flips between today and yesterday; its key stays "today"
    so downstream consumers (e.g. the manager digest) keep working.
    """
    if day == "yesterday":
        d = today - timedelta(days=1)
        first = ("today", "Yesterday", d.strftime("%d %b"), d, d)
    else:
        first = ("today", "Today", today.strftime("%d %b"), today, today)
    return [
        first,
        ("week", "Last 7 Days", f"{(today - timedelta(days=6)).strftime('%d %b')} – {today.strftime('%d %b')}",
         today - timedelta(days=6), today),
        ("month", "Month to Date", today.strftime("%b %Y"),
         today.replace(day=1), today),
    ]


def _sales_for_range(db: Session, start: date, end: date):
    """(sales_rupees, sales_kg) from non-void invoices whose business_date is
    within [start, end]. Invoice.business_date is a Date column."""
    rupees = (
        db.query(func.coalesce(func.sum(Invoice.total), 0))
        .filter(Invoice.business_date >= start,
                Invoice.business_date <= end,
                Invoice.status != "void")
        .scalar()
    ) or 0

    kg = (
        db.query(func.coalesce(func.sum(InvoiceItem.quantity), 0))
        .join(Invoice, InvoiceItem.invoice_id == Invoice.id)
        .filter(Invoice.business_date >= start,
                Invoice.business_date <= end,
                Invoice.status != "void",
                func.lower(InvoiceItem.unit) == "kg")
        .scalar()
    ) or 0

    return float(rupees), float(kg)


def _orders_for_range(db: Session, start: date, end: date):
    """(order_count, active_customer_count) from non-cancelled orders whose
    business_date (stored as 'YYYY-MM-DD' string) is within [start, end].
    ISO date strings sort lexicographically, so string comparison is safe."""
    s, e = start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")
    base = db.query(Order).filter(
        Order.business_date >= s,
        Order.business_date <= e,
        Order.is_cancelled == False,  # noqa: E712
    )
    order_count = base.count()
    active_customers = (
        base.with_entities(Order.customer_phone)
        .distinct()
        .count()
    )
    return order_count, active_customers


def _vasy_sales_for_range(db: Session, start: date, end: date):
    """(revenue, invoice_count, billed_customers) from Vasy sales invoices
    (the authoritative revenue source) with invoice_date in [start, end]."""
    rev, cnt, custs = (
        db.query(
            func.coalesce(func.sum(VasyInvoice.total), 0),
            func.count(VasyInvoice.id),
            func.count(func.distinct(VasyInvoice.customer_id)),
        )
        .filter(VasyInvoice.invoice_date >= start, VasyInvoice.invoice_date <= end)
        .one()
    )
    return float(rev or 0), int(cnt or 0), int(custs or 0)


def _vasy_revenue_by_customer(db: Session, start=None, end=None):
    """{customer_id: revenue} from Vasy invoices (matched customers only),
    optionally bounded by invoice_date in [start, end]."""
    q = db.query(VasyInvoice.customer_id, func.coalesce(func.sum(VasyInvoice.total), 0)) \
        .filter(VasyInvoice.customer_id != None)   # noqa: E711
    if start is not None:
        q = q.filter(VasyInvoice.invoice_date >= start)
    if end is not None:
        q = q.filter(VasyInvoice.invoice_date <= end)
    return {cid: float(t) for cid, t in q.group_by(VasyInvoice.customer_id).all()}


def _vasy_invoice_dates_by_customer(db: Session, upto: date):
    """{customer_id: [dates asc]} of Vasy invoice dates up to `upto` — the
    billing-activity signal (recency / frequency / cadence)."""
    rows = (
        db.query(VasyInvoice.customer_id, VasyInvoice.invoice_date)
        .filter(VasyInvoice.customer_id != None,          # noqa: E711
                VasyInvoice.invoice_date != None,          # noqa: E711
                VasyInvoice.invoice_date <= upto)
        .all()
    )
    out = {}
    for cid, d in rows:
        out.setdefault(cid, []).append(d)
    for cid in out:
        out[cid].sort()
    return out


def business_pulse(db: Session, today: date, day: str = "today") -> dict:
    """P1-1 — Business Pulse KPI strip.

    Sales come from **Vasy** sales invoices (the source of truth for revenue);
    OrdeRR is only the operational order-consolidation layer. Per period
    (Today|Yesterday / Last 7 Days / Month-to-Date): revenue ₹, invoice count,
    billed customers. `day` toggles the first card between today and yesterday.
    """
    periods = []
    for key, label, sublabel, start, end in _period_bounds(today, day):
        rupees, invoices, customers = _vasy_sales_for_range(db, start, end)
        periods.append({
            "key": key,
            "label": label,
            "sublabel": sublabel,
            "sales_rupees": rupees,
            "sales_rupees_fmt": fmt_inr(rupees),
            "invoices": invoices,
            "active_customers": customers,
        })
    return {"periods": periods}


def sales_breakdown(db: Session, today: date, period: str = "today",
                    day: str = "today") -> dict:
    """Drill-down behind a Sales KPI on the Business Pulse card: the Vasy
    invoices that make up a period's revenue, grouped by customer ('sold to
    whom'), each with its invoice count. Rows sort by revenue, largest first.

    `period` is a pulse key ("today" | "week" | "month"); `day` mirrors the card
    toggle so the first ("today") card resolves to today vs yesterday exactly as
    business_pulse() does.
    """
    bounds = {b[0]: b for b in _period_bounds(today, day)}
    key, label, sublabel, start, end = bounds.get(period, bounds["today"])

    rows_q = (
        db.query(VasyInvoice.party_name, VasyInvoice.customer_id,
                 func.count(VasyInvoice.id),
                 func.coalesce(func.sum(VasyInvoice.total), 0))
        .filter(VasyInvoice.invoice_date >= start,
                VasyInvoice.invoice_date <= end)
        .group_by(VasyInvoice.party_name, VasyInvoice.customer_id)
        .all()
    )

    ordered = sorted(
        ({"party_name": (party or "—"), "customer_id": cust_id,
          "count": int(cnt or 0), "total": float(amt or 0)}
         for party, cust_id, cnt, amt in rows_q),
        key=lambda r: r["total"], reverse=True,
    )
    rows = [{
        "party_name":   r["party_name"],
        "customer_id":  r["customer_id"],
        "unattributed": r["customer_id"] is None,
        "count":        r["count"],
        "total":        round(r["total"], 2), "total_fmt": fmt_inr(r["total"]),
    } for r in ordered]

    grand = sum(r["total"] for r in ordered)
    return {
        "period":       key,
        "label":        label,
        "window_label": f"{label} · {sublabel}",
        "rows":         rows,
        "party_count":  len(rows),
        "total_fmt":    fmt_inr(grand),
    }


# ── P2 money helpers: latest / previous outstanding snapshot ───────────────

def _latest_two_snapshot_dates(db: Session):
    """The two most recent snapshot_dates present (latest, previous|None)."""
    dates = [d[0] for d in db.query(OutstandingSnapshot.snapshot_date)
             .distinct().order_by(OutstandingSnapshot.snapshot_date.desc()).limit(2).all()]
    latest = dates[0] if dates else None
    prev = dates[1] if len(dates) > 1 else None
    return latest, prev


def _closing_by_customer(db: Session, snapshot_date) -> dict:
    """customer_id → SUM of closing across that customer's snapshot rows.

    Summed, not last-row-wins: one customer can legitimately own several Vasy
    ledgers (two outlets merged via alias, running + old account). A plain
    {customer_id: closing} dict comprehension silently dropped all but one row
    — the 13 Jul 2026 ₹6.85L undercount in Customer 360."""
    if snapshot_date is None:
        return {}
    rows = (db.query(OutstandingSnapshot.customer_id,
                     func.coalesce(func.sum(OutstandingSnapshot.closing), 0))
            .filter(OutstandingSnapshot.snapshot_date == snapshot_date,
                    OutstandingSnapshot.customer_id != None)  # noqa: E711
            .group_by(OutstandingSnapshot.customer_id).all())
    return {cid: float(total) for cid, total in rows}


def _bad_debt_ids(db: Session) -> set:
    """customer_ids whose balance the owner has written off as unrecoverable
    (hotel closed, owner fled…). AR views exclude these customers' balances;
    the write-off itself lives only in OrdeRR — the Vasy ledger is untouched."""
    return {cid for (cid,) in db.query(BadDebt.customer_id).all()}


def _outstanding_total(db: Session, snapshot_date, bad_ids: set = None):
    """Gross AR = sum of DEBTOR balances only (closing > 0), excluding customers
    written off as bad debt. Customers in credit (negative closing) are a
    separate advance/liability, not netted into AR — so this matches the
    Receivables page's 'Total outstanding (AR)'. Netting credits in here
    understated AR and made the two screens disagree."""
    if snapshot_date is None:
        return 0.0
    if bad_ids is None:
        bad_ids = _bad_debt_ids(db)
    q = db.query(func.coalesce(func.sum(OutstandingSnapshot.closing), 0)) \
        .filter(OutstandingSnapshot.snapshot_date == snapshot_date,
                OutstandingSnapshot.closing > 0)
    if bad_ids:
        q = q.filter((OutstandingSnapshot.customer_id == None) |               # noqa: E711
                     (OutstandingSnapshot.customer_id.notin_(bad_ids)))
    return float(q.scalar() or 0)


def _collections_for_range(db: Session, start: date, end: date):
    """(total, cash, bank) collections from receipts with receipt_date in range."""
    q = db.query(CustomerReceipt.mode, func.coalesce(func.sum(CustomerReceipt.amount), 0)) \
        .filter(CustomerReceipt.receipt_date >= start, CustomerReceipt.receipt_date <= end) \
        .group_by(CustomerReceipt.mode).all()
    total = cash = bank = 0.0
    for mode, amt in q:
        amt = float(amt or 0)
        total += amt
        if (mode or "").lower() == "cash":
            cash += amt
        elif (mode or "").lower() == "bank":
            bank += amt
    return total, cash, bank


def money_pulse(db: Session, today: date, day: str = "today") -> dict:
    """P2-6 — money KPIs: collections per period (from receipts), billed vs
    collected (this month), and total outstanding + trend vs the previous
    snapshot. Returns has_data=False when no receipts/snapshots imported yet.
    `day` toggles the first card between today and yesterday.
    """
    has_receipts = db.query(CustomerReceipt.id).first() is not None
    latest_snap, prev_snap = _latest_two_snapshot_dates(db)

    periods = []
    for key, label, sublabel, start, end in _period_bounds(today, day):
        total, cash, bank = _collections_for_range(db, start, end)
        billed, _, _ = _vasy_sales_for_range(db, start, end)   # Vasy invoices = billed truth
        periods.append({
            "key": key, "label": label, "sublabel": sublabel,
            "collections": total, "collections_fmt": fmt_inr(total),
            "cash_fmt": fmt_inr(cash), "bank_fmt": fmt_inr(bank),
            "billed": billed, "billed_fmt": fmt_inr(billed),
            "net": billed - total, "net_fmt": fmt_inr(billed - total),
        })

    total_ar = _outstanding_total(db, latest_snap)
    prev_ar = _outstanding_total(db, prev_snap) if prev_snap else None
    ar_delta = (total_ar - prev_ar) if prev_ar is not None else None

    return {
        "has_data": has_receipts or latest_snap is not None,
        "periods": periods,
        "total_outstanding": total_ar,
        "total_outstanding_fmt": fmt_inr(total_ar),
        "outstanding_as_of": latest_snap.strftime("%d %b %Y") if latest_snap else None,
        "outstanding_delta_fmt": (fmt_inr(ar_delta) if ar_delta is not None else None),
        "outstanding_direction": (None if ar_delta is None else
                                  ("up" if ar_delta > 0 else "down" if ar_delta < 0 else "flat")),
        "prev_snap": prev_snap.strftime("%d %b %Y") if prev_snap else None,
    }


def collections_breakdown(db: Session, today: date, period: str = "today",
                          day: str = "today") -> dict:
    """Drill-down behind a Collected KPI on the money pulse card: the receipts
    that make up the collected total for one period, grouped by party (i.e.
    'collected against whom'), each with its cash/bank split and receipt count.

    `period` is a pulse key ("today" | "week" | "month"); `day` mirrors the card
    toggle so the first ("today") card resolves to today vs yesterday exactly as
    money_pulse() does. Rows sort by amount collected, largest first.
    """
    bounds = {b[0]: b for b in _period_bounds(today, day)}
    key, label, sublabel, start, end = bounds.get(period, bounds["today"])

    rows_q = (
        db.query(CustomerReceipt.party_name, CustomerReceipt.customer_id,
                 CustomerReceipt.mode, func.count(CustomerReceipt.id),
                 func.coalesce(func.sum(CustomerReceipt.amount), 0))
        .filter(CustomerReceipt.receipt_date >= start,
                CustomerReceipt.receipt_date <= end)
        .group_by(CustomerReceipt.party_name, CustomerReceipt.customer_id,
                  CustomerReceipt.mode)
        .all()
    )

    agg: dict = {}
    for party, cust_id, mode, cnt, amt in rows_q:
        amt = float(amt or 0)
        name = party or "—"
        p = agg.setdefault(name, {"party_name": name, "customer_id": cust_id,
                                  "count": 0, "cash": 0.0, "bank": 0.0, "total": 0.0})
        p["count"] += int(cnt or 0)
        p["total"] += amt
        m = (mode or "").lower()
        if m == "cash":
            p["cash"] += amt
        elif m == "bank":
            p["bank"] += amt
        if p["customer_id"] is None and cust_id is not None:
            p["customer_id"] = cust_id

    ordered = sorted(agg.values(), key=lambda r: r["total"], reverse=True)
    rows = [{
        "party_name":   r["party_name"],
        "customer_id":  r["customer_id"],
        "unattributed": r["customer_id"] is None,
        "count":        r["count"],
        "cash":         round(r["cash"], 2), "cash_fmt": fmt_inr(r["cash"]),
        "bank":         round(r["bank"], 2), "bank_fmt": fmt_inr(r["bank"]),
        "total":        round(r["total"], 2), "total_fmt": fmt_inr(r["total"]),
    } for r in ordered]

    grand = sum(r["total"] for r in ordered)
    return {
        "period":       key,
        "label":        label,
        "window_label": f"{label} · {sublabel}",
        "rows":         rows,
        "party_count":  len(rows),
        "total_fmt":    fmt_inr(grand),
    }


# ── P2-10 Collection velocity + P2-4 Unattributed receipts ─────────────────

def collections(db: Session, today: date, weeks: int = 12) -> dict:
    """P2-10 — weekly collection velocity (₹/week, cash vs bank) over the last
    `weeks` weeks, plus P2-4 unattributed receipts (customer_id NULL).

    Reconciliation: attributed + unattributed = grand total collected.
    """
    has_receipts = db.query(CustomerReceipt.id).first() is not None
    if not has_receipts:
        return {"has_data": False}

    window_start = today - timedelta(days=7 * weeks - 1)

    # weekly buckets (oldest → newest, each ending on a 7-day boundary at today)
    buckets = []
    for i in range(weeks - 1, -1, -1):
        wk_end = today - timedelta(days=7 * i)
        wk_start = wk_end - timedelta(days=6)
        buckets.append({"start": wk_start, "end": wk_end,
                        "label": wk_start.strftime("%d %b"),
                        "total": 0.0, "cash": 0.0, "bank": 0.0})

    rows = (
        db.query(CustomerReceipt.receipt_date, CustomerReceipt.mode,
                 func.coalesce(func.sum(CustomerReceipt.amount), 0))
        .filter(CustomerReceipt.receipt_date >= window_start,
                CustomerReceipt.receipt_date <= today)
        .group_by(CustomerReceipt.receipt_date, CustomerReceipt.mode).all()
    )
    for rdate, mode, amt in rows:
        amt = float(amt or 0)
        # find bucket
        idx = (rdate - window_start).days // 7
        if 0 <= idx < len(buckets):
            b = buckets[idx]
            b["total"] += amt
            if (mode or "").lower() == "cash":
                b["cash"] += amt
            elif (mode or "").lower() == "bank":
                b["bank"] += amt

    win_total = sum(b["total"] for b in buckets)
    win_cash = sum(b["cash"] for b in buckets)
    win_bank = sum(b["bank"] for b in buckets)

    series = [{"label": b["label"],
               "total": round(b["total"], 2), "total_fmt": fmt_inr(b["total"]),
               "cash": round(b["cash"], 2), "bank": round(b["bank"], 2)}
              for b in buckets]

    # P2-4 unattributed receipts (all time, customer_id NULL)
    un_rows = (
        db.query(CustomerReceipt.party_name,
                 func.count(CustomerReceipt.id),
                 func.coalesce(func.sum(CustomerReceipt.amount), 0))
        .filter(CustomerReceipt.customer_id == None)   # noqa: E711
        .group_by(CustomerReceipt.party_name)
        .order_by(func.sum(CustomerReceipt.amount).desc()).all()
    )
    unattributed = [{"party_name": p, "count": n, "total": round(float(t), 2),
                     "total_fmt": fmt_inr(t)} for p, n, t in un_rows]
    un_total = float(db.query(func.coalesce(func.sum(CustomerReceipt.amount), 0))
                     .filter(CustomerReceipt.customer_id == None).scalar() or 0)  # noqa: E711
    grand_total = float(db.query(func.coalesce(func.sum(CustomerReceipt.amount), 0)).scalar() or 0)

    return {
        "has_data": True,
        "weeks": weeks,
        "window_from": window_start.strftime("%d %b %Y"),
        "series": series,
        "win_total_fmt": fmt_inr(win_total),
        "win_cash_fmt": fmt_inr(win_cash),
        "win_bank_fmt": fmt_inr(win_bank),
        "win_cash_pct": round(win_cash / win_total * 100) if win_total else 0,
        "win_bank_pct": round(win_bank / win_total * 100) if win_total else 0,
        "avg_per_week_fmt": fmt_inr(win_total / weeks) if weeks else fmt_inr(0),
        "unattributed": unattributed,
        "unattributed_total_fmt": fmt_inr(un_total),
        "unattributed_count": len(unattributed),
        "attributed_total_fmt": fmt_inr(grand_total - un_total),
        "grand_total_fmt": fmt_inr(grand_total),
    }


# ── P3-1/2/4/5 Credit intelligence (risk score, classification, at-risk) ───

def _lin(x, lo, hi):
    """Map x in [lo,hi] → 0..100 (clamped). Below lo→0, above hi→100."""
    if x is None:
        return 0.0
    if x <= lo:
        return 0.0
    if x >= hi:
        return 100.0
    return (x - lo) / (hi - lo) * 100.0


def credit_intelligence(db: Session, today: date) -> dict:
    """P3-4 credit-risk score (0–100, higher = riskier) fused from:
      • payment lateness  — days since last payment (P2 receipts)
      • exposure burden   — outstanding ÷ avg monthly collections (P3-2 ratio)
      • order slowdown    — gap since last order ÷ own cadence (P1-4 signal)
      • balance trend     — rising outstanding vs previous snapshot (P2-9)
    → P3-5 classification (Grow / Watch / Chase / Hold-credit),
    → P3-1 at-risk flag with human-readable reasons,
    → P3-3 credit-limit breach.

    Explainable: every row carries its sub-scores + reasons. Scored on
    customers with any exposure / order / payment history.
    """
    latest_snap, prev_snap = _latest_two_snapshot_dates(db)
    snap_now = _closing_by_customer(db, latest_snap)    # summed per customer
    snap_prev = _closing_by_customer(db, prev_snap)
    rstats = _receipt_stats_by_customer(db)
    dates_by_cust = _vasy_invoice_dates_by_customer(db, today)   # billing cadence
    sp_name = {s.id: s.name for s in db.query(Salesperson).all()}

    # receipt window span in months (denominator for avg monthly collections)
    rd = db.query(func.min(CustomerReceipt.receipt_date),
                  func.max(CustomerReceipt.receipt_date)).one()
    span_months = 1
    if rd[0] and rd[1]:
        span_months = max(1, (rd[1].year - rd[0].year) * 12 + (rd[1].month - rd[0].month) + 1)

    if not latest_snap and not rstats:
        return {"has_data": False}

    rows = []
    areas, salespeople = set(), set()
    class_counts = {"Grow": 0, "Watch": 0, "Chase": 0, "Hold-credit": 0}
    at_risk_count = breach_count = 0

    for c in db.query(Customer).all():
        outstanding = snap_now.get(c.id, 0.0)
        rec = rstats.get(c.id)              # (last_date, count, total) or None
        dates = dates_by_cust.get(c.id)
        has_orders = bool(dates)
        has_activity = outstanding != 0 or rec is not None or has_orders
        if not has_activity:
            continue

        reasons = []

        # ── payment lateness ──
        if rec and rec[0]:
            days_since_pay = (today - rec[0]).days
            payment_lateness = _lin(days_since_pay, 7, 60)
            if days_since_pay >= 30:
                reasons.append(f"last paid {days_since_pay}d ago")
        elif outstanding > 0:
            days_since_pay = None
            payment_lateness = 100.0
            reasons.append("no payment in the imported receipts")
        else:
            days_since_pay = None
            payment_lateness = 0.0

        # ── exposure burden (P3-2 ratio) ──
        collected = rec[2] if rec else 0.0
        avg_monthly = collected / span_months if collected else 0.0
        if outstanding <= 0:
            exposure_burden = 0.0
            exposure_months = 0.0
        elif avg_monthly > 0:
            exposure_months = outstanding / avg_monthly
            exposure_burden = _lin(exposure_months, 0.5, 3.0)
            if exposure_months >= 2:
                reasons.append(f"exposure ≈ {round(exposure_months,1)} months of collections")
        else:
            exposure_months = None
            exposure_burden = 100.0
            reasons.append("owes money with no collections on record")

        # ── order slowdown (cadence) ──
        order_slowdown = None
        if has_orders and len(dates) >= 3:
            gaps = [(dates[i] - dates[i - 1]).days for i in range(1, len(dates))]
            gaps = [g for g in gaps if g > 0] or gaps
            cadence = median(gaps) if gaps else None
            if cadence:
                days_since_order = (today - dates[-1]).days
                ratio = days_since_order / cadence
                order_slowdown = _lin(ratio, 1.0, 3.0)
                if ratio >= 2:
                    reasons.append(f"orders slowed to {round(ratio,1)}× usual cadence")

        # ── balance trend ──
        rising = False
        if c.id in snap_now and c.id in snap_prev:
            if snap_now[c.id] > snap_prev[c.id]:
                rising = True
                reasons.append(f"balance rising (+{fmt_inr(snap_now[c.id]-snap_prev[c.id])})")

        # ── composite score ──
        comps = [(payment_lateness, 0.45), (exposure_burden, 0.35)]
        if order_slowdown is not None:
            comps.append((order_slowdown, 0.20))
        wsum = sum(w for _, w in comps)
        score = sum(v * w for v, w in comps) / wsum
        if rising:
            score = min(100.0, score + 5)
        score = round(score)

        # ── breach ──
        limit = float(c.credit_limit) if c.credit_limit is not None else None
        breach = limit is not None and outstanding > limit
        if breach:
            breach_count += 1
            reasons.insert(0, f"over limit by {fmt_inr(outstanding - limit)}")

        # ── classification ──
        if breach or score >= 70:
            cls = "Hold-credit"
        elif score >= 50:
            cls = "Chase"
        elif score >= 30:
            cls = "Watch"
        else:
            cls = "Grow"
        class_counts[cls] += 1

        at_risk = breach or score >= 55
        if at_risk:
            at_risk_count += 1

        sp = (sp_name.get(c.salesperson_id) if c.salesperson_id else "") or ""
        area = c.area or ""
        if area:
            areas.add(area)
        if sp:
            salespeople.add(sp)

        rows.append({
            "customer_id": c.id,
            "name": c.restaurant_name or (c.phone_number or f"#{c.id}"),
            "phone": c.phone_number or "",
            "area": area, "salesperson": sp,
            "outstanding": round(outstanding, 2), "outstanding_fmt": fmt_inr(outstanding),
            "credit_limit": float(limit) if limit is not None else "",
            "credit_limit_fmt": fmt_inr(limit) if limit is not None else "—",
            "breach": breach,
            "score": score,
            "classification": cls,
            "at_risk": at_risk,
            "days_since_payment": days_since_pay if days_since_pay is not None else "",
            "exposure_months": round(exposure_months, 1) if exposure_months is not None else "",
            "sub": {"payment": round(payment_lateness), "exposure": round(exposure_burden),
                    "orders": round(order_slowdown) if order_slowdown is not None else None},
            "reasons": reasons,
        })

    rows.sort(key=lambda r: (r["breach"], r["score"], r["outstanding"]), reverse=True)

    return {
        "has_data": True,
        "as_of": latest_snap.strftime("%d %b %Y") if latest_snap else None,
        "has_trend": prev_snap is not None,
        "span_months": span_months,
        "rows": rows,
        "counts": class_counts,
        "at_risk": at_risk_count,
        "breach": breach_count,
        "scored": len(rows),
        "areas": sorted(areas), "salespeople": sorted(salespeople),
    }


# ── Imports data-coverage (what's loaded, for the Imports page) ────────────

def import_coverage(db: Session, today: date) -> list:
    """Per-entity coverage for the Imports page: how many rows and over what
    date range (ledgers), or how many daily snapshots (outstanding). Lets the
    user see at a glance what's loaded and what's stale/missing."""
    ledgers = [
        ("Receipts", CustomerReceipt, CustomerReceipt.receipt_date, "receipts"),
        ("Sales invoices", VasyInvoice, VasyInvoice.invoice_date, "sales-invoices"),
        ("Purchases", VasyPurchase, VasyPurchase.bill_date, "purchases"),
        ("Expenses", VasyExpense, VasyExpense.expense_date, "expenses"),
        ("Payments", VasyPayment, VasyPayment.payment_date, "payments"),
        ("Supplier bills (AP)", VasySupplierBill, VasySupplierBill.bill_date, "supplier-outstanding"),
    ]
    rows = []

    def _push_ledger(label, model, dcol, key):
        n = db.query(func.count(model.id)).scalar() or 0
        lo, hi = db.query(func.min(dcol), func.max(dcol)).one()
        stale_days = (today - hi).days if hi else None
        rows.append({
            "label": label, "key": key, "kind": "ledger", "rows": n,
            "from": lo.strftime("%d %b %Y") if lo else None,
            "to": hi.strftime("%d %b %Y") if hi else None,
            "stale_days": stale_days,
        })

    # order: Receipts, Outstanding, Sales invoices, Purchases, Expenses, Payments
    _push_ledger(*ledgers[0])
    # Outstanding — snapshot entity
    snap_dates = [d[0] for d in db.query(OutstandingSnapshot.snapshot_date).distinct().all()]
    latest = max(snap_dates) if snap_dates else None
    n_latest = (db.query(func.count(OutstandingSnapshot.id))
                .filter(OutstandingSnapshot.snapshot_date == latest).scalar()) if latest else 0
    rows.append({
        "label": "Outstanding", "key": "outstanding", "kind": "snapshot",
        "snapshots": len(snap_dates), "parties": n_latest,
        "to": latest.strftime("%d %b %Y") if latest else None,
        "stale_days": (today - latest).days if latest else None,
    })
    for l in ledgers[1:]:
        _push_ledger(*l)
    return rows


# ── P3-11/12/13/14 Plant financials (P&L, cash-flow, margin, AP) ───────────

def _monthly_bucket(db, model, date_col, amount_col, keys):
    """Sum `amount_col` by business-month for `model`, restricted to `keys`
    (list of 'YYYY-MM'). Returns {month_key: total}."""
    key_set = set(keys)
    start = date(int(keys[0][:4]), int(keys[0][5:]), 1)
    end = date(int(keys[-1][:4]), int(keys[-1][5:]), 28) + timedelta(days=7)
    out = {k: 0.0 for k in keys}
    for d, amt in db.query(date_col, amount_col).filter(date_col != None,           # noqa: E711
                                                        date_col >= start,
                                                        date_col <= end).all():
        try:
            mk = _month_key(d)
        except Exception:
            continue
        if mk in key_set:
            out[mk] += float(amt or 0)
    return out


def _date_range(db, date_col):
    lo, hi = db.query(func.min(date_col), func.max(date_col)).one()
    return (lo.strftime("%d %b %Y") if lo else None, hi.strftime("%d %b %Y") if hi else None)


def plant_financials(db: Session, today: date, months: int = 12) -> dict:
    """P3-11 P&L (revenue − COGS − expenses), P3-13 gross margin, P3-12 cash
    flow (receipts − payments, running balance), P3-14 payables.

    Revenue = Vasy sales invoices; COGS = Vasy purchases; expenses = Vasy
    expenses (accrual); cash-in = receipts, cash-out = payments (NOT expenses —
    they're realized via payments; adding both double-counts). Monthly series
    for the last `months`. Entities are imported over different date ranges, so
    the UI shows per-entity coverage and treats period P&L as indicative until
    ranges align.
    """
    has_any = any(db.query(m.id).first() is not None
                  for m in (VasyInvoice, VasyPurchase, VasyExpense, VasyPayment, CustomerReceipt))
    if not has_any:
        return {"has_data": False}

    keys = _last_n_months(today, months)
    rev = _monthly_bucket(db, VasyInvoice, VasyInvoice.invoice_date, VasyInvoice.total, keys)
    cogs = _monthly_bucket(db, VasyPurchase, VasyPurchase.bill_date, VasyPurchase.total, keys)
    exp = _monthly_bucket(db, VasyExpense, VasyExpense.expense_date, VasyExpense.total, keys)
    cin = _monthly_bucket(db, CustomerReceipt, CustomerReceipt.receipt_date, CustomerReceipt.amount, keys)
    cout = _monthly_bucket(db, VasyPayment, VasyPayment.payment_date, VasyPayment.amount, keys)

    pnl, cash = [], []
    running = 0.0
    for k in keys:
        r, c, e = round(rev[k], 2), round(cogs[k], 2), round(exp[k], 2)
        gp = r - c
        net = gp - e
        pnl.append({
            "key": k, "label": _month_label(k),
            "revenue": r, "revenue_fmt": fmt_inr(r),
            "cogs": c, "cogs_fmt": fmt_inr(c),
            "gross_profit": round(gp, 2), "gross_profit_fmt": fmt_inr(gp),
            "expenses": e, "expenses_fmt": fmt_inr(e),
            "net": round(net, 2), "net_fmt": fmt_inr(net),
            "margin_pct": round(gp / r * 100, 1) if r else None,
        })
        ci, co = round(cin[k], 2), round(cout[k], 2)
        running += ci - co
        cash.append({
            "key": k, "label": _month_label(k),
            "cash_in": ci, "cash_in_fmt": fmt_inr(ci),
            "cash_out": co, "cash_out_fmt": fmt_inr(co),
            "net_cash": round(ci - co, 2), "net_cash_fmt": fmt_inr(ci - co),
            "running_fmt": fmt_inr(running),
        })

    # totals (over the window)
    t_rev, t_cogs, t_exp = sum(rev.values()), sum(cogs.values()), sum(exp.values())
    t_gp = t_rev - t_cogs
    t_in, t_out = sum(cin.values()), sum(cout.values())

    # ── P3-14 payables (true AP from supplier bills, if imported) ──
    has_ap = db.query(VasySupplierBill.id).first() is not None
    ap = {"has_bills": has_ap}
    if has_ap:
        bills = db.query(VasySupplierBill).filter(VasySupplierBill.due > 0).all()
        gross_due = sum(float(b.due) for b in bills)

        # Big-supplier payments are often entered as standalone vouchers that
        # never get allocated to bills in Vasy, so bill `due` overstates real
        # AP (audit 2026-07-14: Yewale showed ₹73L due with ₹67L already paid
        # by voucher). Per vendor: unallocated credit = payment vouchers −
        # bill-allocated paid − expense-register paid (vouchers to expense
        # heads mirror the expense ledger, not bills). Net the credit against
        # that vendor's open bills oldest-first.
        vendor_keys = {b.vendor_key for b in bills}
        vouchers = dict(db.query(VasyPayment.party_key,
                                 func.coalesce(func.sum(VasyPayment.amount), 0))
                        .filter(VasyPayment.party_key.in_(vendor_keys))
                        .group_by(VasyPayment.party_key).all())
        bill_paid = dict(db.query(VasySupplierBill.vendor_key,
                                  func.coalesce(func.sum(VasySupplierBill.paid), 0))
                         .group_by(VasySupplierBill.vendor_key).all())
        exp_paid = dict(db.query(VasyExpense.party_key,
                                 func.coalesce(func.sum(VasyExpense.paid), 0))
                        .filter(VasyExpense.party_key.in_(vendor_keys))
                        .group_by(VasyExpense.party_key).all())
        credit = {}
        for vk in vendor_keys:
            c = (float(vouchers.get(vk, 0)) - float(bill_paid.get(vk, 0))
                 - float(exp_paid.get(vk, 0)))
            if c > 0.005:
                credit[vk] = c

        remaining = dict(credit)
        eff_due = {}
        for b in sorted(bills, key=lambda b: (b.due_date or b.bill_date or today, b.bill_no)):
            due = float(b.due)
            take = min(due, remaining.get(b.vendor_key, 0.0))
            if take > 0:
                due -= take
                remaining[b.vendor_key] -= take
            eff_due[b.id] = due

        open_bills = [b for b in bills if eff_due[b.id] > 0.005]
        ap_total = sum(eff_due[b.id] for b in open_bills)
        applied = sum(credit.values()) - sum(remaining.values())

        # aging by due date (days overdue = today − due_date), on netted dues
        buckets = {"current": 0.0, "0-30": 0.0, "31-60": 0.0, "61-90": 0.0, "90+": 0.0, "no-date": 0.0}
        for b in open_bills:
            due = eff_due[b.id]
            if b.due_date is None:
                buckets["no-date"] += due
            else:
                od = (today - b.due_date).days
                if od <= 0:
                    buckets["current"] += due
                elif od <= 30:
                    buckets["0-30"] += due
                elif od <= 60:
                    buckets["31-60"] += due
                elif od <= 90:
                    buckets["61-90"] += due
                else:
                    buckets["90+"] += due
        creditors = {}
        for b in open_bills:
            creditors[b.vendor] = creditors.get(b.vendor, 0.0) + eff_due[b.id]
        top_creditors = sorted(creditors.items(), key=lambda kv: kv[1], reverse=True)[:10]

        vendor_name = {}
        for b in bills:
            vendor_name.setdefault(b.vendor_key, b.vendor)
        netted = sorted(((vendor_name[vk], credit[vk] - remaining.get(vk, 0.0))
                         for vk in credit if credit[vk] - remaining.get(vk, 0.0) > 0.005),
                        key=lambda kv: kv[1], reverse=True)
        ap.update({
            "ap_total_fmt": fmt_inr(ap_total),
            "gross_due_fmt": fmt_inr(gross_due),
            "unallocated_fmt": fmt_inr(applied),
            "has_netting": bool(netted),
            "netting": [{"vendor": v, "credit_fmt": fmt_inr(a)} for v, a in netted[:5]],
            "open_bills": len(open_bills),
            "aging": [
                {"bucket": "Not yet due", "amount_fmt": fmt_inr(buckets["current"])},
                {"bucket": "1–30 days overdue", "amount_fmt": fmt_inr(buckets["0-30"])},
                {"bucket": "31–60 days overdue", "amount_fmt": fmt_inr(buckets["31-60"])},
                {"bucket": "61–90 days overdue", "amount_fmt": fmt_inr(buckets["61-90"])},
                {"bucket": "90+ days overdue", "amount_fmt": fmt_inr(buckets["90+"])},
            ],
            "top_creditors": [{"vendor": v, "due_fmt": fmt_inr(d)} for v, d in top_creditors],
        })

    # supporting: unpaid expenses + top payees (money out)
    unpaid_total = float(db.query(func.coalesce(func.sum(VasyExpense.unpaid), 0)).scalar() or 0)
    payee_rows = db.query(VasyPayment.party_name, func.count(VasyPayment.id),
                          func.coalesce(func.sum(VasyPayment.amount), 0)) \
        .group_by(VasyPayment.party_name).order_by(func.sum(VasyPayment.amount).desc()).limit(15).all()
    top_payees = [{"party": p, "count": n, "total_fmt": fmt_inr(t)} for p, n, t in payee_rows]

    return {
        "has_data": True,
        "pnl": pnl, "cash": cash,
        "totals": {
            "revenue_fmt": fmt_inr(t_rev), "cogs_fmt": fmt_inr(t_cogs),
            "gross_profit_fmt": fmt_inr(t_gp),
            "margin_pct": round(t_gp / t_rev * 100, 1) if t_rev else None,
            "margin_threshold": _config.MARGIN_ALERT_PCT,
            "margin_alert": bool(t_rev and (t_gp / t_rev * 100) < _config.MARGIN_ALERT_PCT),
            "expenses_fmt": fmt_inr(t_exp), "net_fmt": fmt_inr(t_gp - t_exp),
            "cash_in_fmt": fmt_inr(t_in), "cash_out_fmt": fmt_inr(t_out),
            "net_cash_fmt": fmt_inr(t_in - t_out),
        },
        "coverage": {
            "revenue": _date_range(db, VasyInvoice.invoice_date),
            "cogs": _date_range(db, VasyPurchase.bill_date),
            "expenses": _date_range(db, VasyExpense.expense_date),
            "receipts": _date_range(db, CustomerReceipt.receipt_date),
            "payments": _date_range(db, VasyPayment.payment_date),
        },
        "payables": {
            "ap": ap,
            "unpaid_total_fmt": fmt_inr(unpaid_total),
            "top_payees": top_payees,
        },
    }


def business_overview(db: Session, start: date, end: date, row_cap: int = 1000) -> dict:
    """Business-at-a-glance for a [start, end] date window: four money ledgers —
    Sales, Purchases, Expenses and Received — each with its period total, count
    and the underlying rows (for the drill-down on the Business tab).

    All four come straight from the Vasy mirror (source of truth for money):
      Sales     = Vasy sales invoices  (invoice_date, total)
      Purchases = Vasy purchase bills  (bill_date, total)
      Expenses  = Vasy expenses        (expense_date, total)
      Received  = customer receipts    (receipt_date, amount)

    Each ledger's `rows` are newest-first and capped at `row_cap` (the total /
    count always reflect the full window; `truncated` flags a capped list).
    """
    def _fmt_d(d):
        return d.strftime("%d %b %Y") if d else "—"

    def _grouped(model, party_col, date_col, amt_col, card_total,
                 unpaid_col=None, cap=25):
        """Group a ledger by party within the window — the 'by head/payee' view.
        Returns (rows, group_count); rows are top `cap` by total, each with the
        share of the card total and (optionally) an unpaid sum."""
        cols = [party_col, func.count(model.id), func.coalesce(func.sum(amt_col), 0)]
        if unpaid_col is not None:
            cols.append(func.coalesce(func.sum(unpaid_col), 0))
        rows = (db.query(*cols)
                .filter(date_col != None,                    # noqa: E711
                        date_col >= start, date_col <= end)
                .group_by(party_col)
                .order_by(func.sum(amt_col).desc()).all())
        grand = card_total or 1.0
        out = []
        for r in rows[:cap]:
            item = {"name": r[0] or "—", "count": int(r[1]),
                    "total": float(r[2]), "total_fmt": fmt_inr(r[2]),
                    "pct": round(float(r[2]) / grand * 100, 1)}
            if unpaid_col is not None:
                item["unpaid_fmt"] = fmt_inr(r[3])
            out.append(item)
        return out, len(rows)

    # ── Sales (Vasy invoices) ────────────────────────────────────────────────
    sales_q = (db.query(VasyInvoice)
               .filter(VasyInvoice.invoice_date != None,               # noqa: E711
                       VasyInvoice.invoice_date >= start,
                       VasyInvoice.invoice_date <= end)
               .order_by(VasyInvoice.invoice_date.desc(), VasyInvoice.id.desc()))
    sales_total = float(db.query(func.coalesce(func.sum(VasyInvoice.total), 0))
                        .filter(VasyInvoice.invoice_date != None,      # noqa: E711
                                VasyInvoice.invoice_date >= start,
                                VasyInvoice.invoice_date <= end).scalar() or 0)
    sales_count = sales_q.count()
    sales_rows = [{
        "date": _fmt_d(r.invoice_date), "ref": r.voucher_no,
        "party": r.party_name, "meta": f"{r.item_count} item(s)",
        "amount": float(r.total or 0), "amount_fmt": fmt_inr(r.total),
    } for r in sales_q.limit(row_cap).all()]

    # ── Purchases (Vasy purchase bills) ──────────────────────────────────────
    purch_q = (db.query(VasyPurchase)
               .filter(VasyPurchase.bill_date != None,                 # noqa: E711
                       VasyPurchase.bill_date >= start,
                       VasyPurchase.bill_date <= end)
               .order_by(VasyPurchase.bill_date.desc(), VasyPurchase.id.desc()))
    purch_total = float(db.query(func.coalesce(func.sum(VasyPurchase.total), 0))
                        .filter(VasyPurchase.bill_date != None,        # noqa: E711
                                VasyPurchase.bill_date >= start,
                                VasyPurchase.bill_date <= end).scalar() or 0)
    purch_count = purch_q.count()
    purch_rows = [{
        "date": _fmt_d(r.bill_date), "ref": r.bill_no,
        "party": r.party_name, "meta": f"{r.item_count} item(s)",
        "amount": float(r.total or 0), "amount_fmt": fmt_inr(r.total),
    } for r in purch_q.limit(row_cap).all()]

    # ── Expenses (Vasy expenses) ─────────────────────────────────────────────
    exp_q = (db.query(VasyExpense)
             .filter(VasyExpense.expense_date != None,                 # noqa: E711
                     VasyExpense.expense_date >= start,
                     VasyExpense.expense_date <= end)
             .order_by(VasyExpense.expense_date.desc(), VasyExpense.id.desc()))
    exp_total = float(db.query(func.coalesce(func.sum(VasyExpense.total), 0))
                      .filter(VasyExpense.expense_date != None,        # noqa: E711
                              VasyExpense.expense_date >= start,
                              VasyExpense.expense_date <= end).scalar() or 0)
    exp_count = exp_q.count()
    exp_rows = [{
        "date": _fmt_d(r.expense_date), "ref": r.expense_no,
        "party": r.party_name,
        "meta": ("unpaid " + fmt_inr(r.unpaid)) if float(r.unpaid or 0) > 0 else "paid",
        "amount": float(r.total or 0), "amount_fmt": fmt_inr(r.total),
    } for r in exp_q.limit(row_cap).all()]

    # ── Received (customer receipts) ─────────────────────────────────────────
    rcpt_q = (db.query(CustomerReceipt)
              .filter(CustomerReceipt.receipt_date != None,            # noqa: E711
                      CustomerReceipt.receipt_date >= start,
                      CustomerReceipt.receipt_date <= end)
              .order_by(CustomerReceipt.receipt_date.desc(), CustomerReceipt.id.desc()))
    rcpt_total = float(db.query(func.coalesce(func.sum(CustomerReceipt.amount), 0))
                       .filter(CustomerReceipt.receipt_date != None,   # noqa: E711
                               CustomerReceipt.receipt_date >= start,
                               CustomerReceipt.receipt_date <= end).scalar() or 0)
    rcpt_count = rcpt_q.count()
    rcpt_rows = [{
        "date": _fmt_d(r.receipt_date), "ref": r.receipt_no,
        "party": r.party_name, "meta": (r.mode or "").upper(),
        "amount": float(r.amount or 0), "amount_fmt": fmt_inr(r.amount),
    } for r in rcpt_q.limit(row_cap).all()]

    # ── Grouped-by-party breakdowns (the 'by head/payee' view) ───────────────
    sales_brk, sales_grp = _grouped(VasyInvoice, VasyInvoice.party_name,
                                    VasyInvoice.invoice_date, VasyInvoice.total, sales_total)
    purch_brk, purch_grp = _grouped(VasyPurchase, VasyPurchase.party_name,
                                    VasyPurchase.bill_date, VasyPurchase.total, purch_total)
    exp_brk, exp_grp = _grouped(VasyExpense, VasyExpense.party_name,
                                VasyExpense.expense_date, VasyExpense.total, exp_total,
                                unpaid_col=VasyExpense.unpaid)
    rcpt_brk, rcpt_grp = _grouped(CustomerReceipt, CustomerReceipt.party_name,
                                  CustomerReceipt.receipt_date, CustomerReceipt.amount, rcpt_total)

    cards = [
        {"key": "sales",    "label": "Sales",     "icon": "🧾", "tone": "green",
         "hint": "Vasy sales invoices", "total": sales_total, "total_fmt": fmt_inr(sales_total),
         "count": sales_count, "rows": sales_rows, "truncated": sales_count > len(sales_rows),
         "ref_head": "Voucher", "meta_head": "Items",
         "group_head": "Customer", "breakdown": sales_brk, "group_count": sales_grp,
         "has_unpaid": False},
        {"key": "purchases", "label": "Purchases", "icon": "📦", "tone": "amber",
         "hint": "Vasy purchase bills", "total": purch_total, "total_fmt": fmt_inr(purch_total),
         "count": purch_count, "rows": purch_rows, "truncated": purch_count > len(purch_rows),
         "ref_head": "Bill no", "meta_head": "Items",
         "group_head": "Supplier", "breakdown": purch_brk, "group_count": purch_grp,
         "has_unpaid": False},
        {"key": "expenses", "label": "Expenses",  "icon": "💡", "tone": "red",
         "hint": "Vasy expenses (opex)", "total": exp_total, "total_fmt": fmt_inr(exp_total),
         "count": exp_count, "rows": exp_rows, "truncated": exp_count > len(exp_rows),
         "ref_head": "Expense no", "meta_head": "Status",
         "group_head": "Expense head / payee", "breakdown": exp_brk, "group_count": exp_grp,
         "has_unpaid": True},
        {"key": "received", "label": "Received",  "icon": "💰", "tone": "blue",
         "hint": "Money in (receipts)", "total": rcpt_total, "total_fmt": fmt_inr(rcpt_total),
         "count": rcpt_count, "rows": rcpt_rows, "truncated": rcpt_count > len(rcpt_rows),
         "ref_head": "Receipt no", "meta_head": "Mode",
         "group_head": "Customer", "breakdown": rcpt_brk, "group_count": rcpt_grp,
         "has_unpaid": False},
    ]

    # Manual monthly overheads (salaries etc.) that Vasy's register never books.
    # Each row sits on the 1st of its month; include the rows whose month-start
    # falls in the window. These are accrual P&L lines, so they reduce Net.
    from orderr_core.models.monthly_overhead import MonthlyOverhead
    oh_rows_q = (db.query(MonthlyOverhead)
                 .filter(MonthlyOverhead.period >= start,
                         MonthlyOverhead.period <= end)
                 .order_by(MonthlyOverhead.period.desc(), MonthlyOverhead.id.desc())
                 .all())
    overhead_total = float(sum(float(o.amount) for o in oh_rows_q))
    overheads = [{
        "id": o.id,
        "period_fmt": o.period.strftime("%b %Y"),
        "head": o.head,
        "amount_fmt": fmt_inr(float(o.amount)),
        "note": o.note or "",
    } for o in oh_rows_q]

    vasy_out = purch_total + exp_total          # Vasy-booked purchases + expenses
    money_out = vasy_out + overhead_total       # + manual overheads (salaries)
    net = sales_total - money_out
    return {
        "has_data": bool(sales_count or purch_count or exp_count or rcpt_count),
        "start": start, "end": end,
        "start_fmt": _fmt_d(start), "end_fmt": _fmt_d(end),
        "start_iso": start.isoformat(), "end_iso": end.isoformat(),
        "cards": cards,
        # Accrual margin read for the range: Sales − (Purchases + Expenses +
        # manual overheads). Receipts (cash-in) show separately in the Received
        # card; loans / owner cash-in are cash-only and live in the Cash Book.
        "sales_fmt": fmt_inr(sales_total),
        "vasy_out_fmt": fmt_inr(vasy_out),
        "overhead_total": round(overhead_total, 2),
        "overhead_total_fmt": fmt_inr(overhead_total),
        "overheads": overheads,
        "money_out_fmt": fmt_inr(money_out),
        "net_fmt": fmt_inr(net),
        "net_positive": net >= 0,
    }


def add_overhead(db: Session, data: dict) -> str | None:
    """Add/replace a manual monthly overhead (salaries etc.). `period` accepts a
    YYYY-MM (or any YYYY-MM-DD) and is normalised to the 1st. One figure per
    (month, head) — re-saving the same pair overwrites. Returns an error string,
    or None on success (house convention)."""
    from orderr_core.models.monthly_overhead import MonthlyOverhead

    raw = (data.get("period") or "").strip()
    period = None
    for fmt in ("%Y-%m", "%Y-%m-%d"):
        try:
            period = datetime.strptime(raw, fmt).date().replace(day=1)
            break
        except ValueError:
            continue
    if period is None:
        return "Pick the month this figure is for."
    head = (data.get("head") or "Salaries").strip() or "Salaries"
    try:
        amount = round(float(str(data.get("amount") or "").replace(",", "").strip()), 2)
    except ValueError:
        return "Enter the amount."
    if amount <= 0:
        return "Amount must be more than zero."
    note = (data.get("note") or "").strip() or None

    existing = (db.query(MonthlyOverhead)
                .filter(MonthlyOverhead.period == period,
                        func.lower(MonthlyOverhead.head) == head.lower())
                .first())
    if existing:
        existing.amount = amount
        existing.note = note
        existing.head = head
    else:
        db.add(MonthlyOverhead(period=period, head=head, amount=amount, note=note))
    db.commit()
    return None


def delete_overhead(db: Session, overhead_id: int) -> str | None:
    """Remove one manual overhead row. Returns an error string, or None."""
    from orderr_core.models.monthly_overhead import MonthlyOverhead

    rec = db.get(MonthlyOverhead, overhead_id)
    if rec is None:
        return "Overhead not found."
    db.delete(rec)
    db.commit()
    return None


# ledger key → (model, party column, date column, amount column, ref column,
# ref label, extra unpaid column). Shared by the day-wise drill-down below.
_LEDGER_MAP = {
    "sales":     (VasyInvoice,     "party_name", "invoice_date", "total",  "voucher_no",  "Voucher",     None),
    "purchases": (VasyPurchase,    "party_name", "bill_date",    "total",  "bill_no",     "Bill no",     None),
    "expenses":  (VasyExpense,     "party_name", "expense_date", "total",  "expense_no",  "Expense no",  "unpaid"),
    "received":  (CustomerReceipt, "party_name", "receipt_date", "amount", "receipt_no",  "Receipt no",  None),
}


def ledger_daywise(db: Session, ledger: str, party: str,
                   start: date = None, end: date = None) -> dict:
    """Day-wise drill-down behind one row of a Business/Expenses breakdown: for a
    single party (customer / supplier / expense head), the per-day entry count and
    amount, newest day first. Window is optional — the Expenses tab groups
    all-time, the Business tab passes its [start, end]."""
    cfg = _LEDGER_MAP.get(ledger)
    if cfg is None:
        return {"ok": False, "rows": []}
    model, party_attr, date_attr, amt_attr, ref_attr, ref_head, unpaid_attr = cfg
    party_col = getattr(model, party_attr)
    date_col = getattr(model, date_attr)
    amt_col = getattr(model, amt_attr)

    def _fmt_d(d):
        return d.strftime("%d %b %Y") if d else "—"

    # party_name can be NULL (unmatched) — treat the "—" label as a NULL match.
    party_filter = (party_col == None) if (party is None or party == "—") else (party_col == party)  # noqa: E711
    filters = [date_col != None, party_filter]                          # noqa: E711
    if start is not None:
        filters.append(date_col >= start)
    if end is not None:
        filters.append(date_col <= end)

    cols = [date_col, func.count(model.id), func.coalesce(func.sum(amt_col), 0)]
    has_unpaid = unpaid_attr is not None
    if has_unpaid:
        cols.append(func.coalesce(func.sum(getattr(model, unpaid_attr)), 0))

    grp = (db.query(*cols).filter(*filters)
           .group_by(date_col).order_by(date_col.desc()).all())

    rows, total, count = [], 0.0, 0
    for r in grp:
        d, n, t = r[0], int(r[1]), float(r[2])
        total += t
        count += n
        item = {"date": _fmt_d(d), "count": n, "total_fmt": fmt_inr(t)}
        if has_unpaid:
            item["unpaid_fmt"] = fmt_inr(r[3])
        rows.append(item)

    label = party if (party and party != "—") else "(unmatched)"
    return {
        "ok": True,
        "ledger": ledger,
        "party": label,
        "ref_head": ref_head,
        "has_unpaid": has_unpaid,
        "rows": rows,
        "day_count": len(rows),
        "count": count,
        "total_fmt": fmt_inr(total),
        "window_label": (f"{_fmt_d(start)} → {_fmt_d(end)}"
                         if (start is not None or end is not None) else "All time"),
    }


def plant_expenses(db: Session, today: date, months: int = 12) -> dict:
    """Standalone opex (expense-ledger) view: monthly trend, paid/unpaid split,
    and a breakdown by expense head (party_name). Sourced only from Vasy
    expenses — this is the accrual opex ledger, NOT cash-out (payments realize
    expenses; the Cash book / Financials own the cash side). Vasy exports no
    category column, so `party_name` (the expense head / payee) is the grouping
    dimension.
    """
    if db.query(VasyExpense.id).first() is None:
        return {"has_data": False}

    keys = _last_n_months(today, months)
    tot = _monthly_bucket(db, VasyExpense, VasyExpense.expense_date, VasyExpense.total, keys)
    paid = _monthly_bucket(db, VasyExpense, VasyExpense.expense_date, VasyExpense.paid, keys)
    unpaid = _monthly_bucket(db, VasyExpense, VasyExpense.expense_date, VasyExpense.unpaid, keys)
    cnt_rows = db.query(VasyExpense.expense_date, VasyExpense.id).filter(VasyExpense.expense_date != None).all()  # noqa: E711
    counts = {k: 0 for k in keys}
    for d, _ in cnt_rows:
        try:
            mk = _month_key(d)
        except Exception:
            continue
        if mk in counts:
            counts[mk] += 1

    monthly = []
    for k in keys:
        t, p, u, n = round(tot[k], 2), round(paid[k], 2), round(unpaid[k], 2), counts[k]
        monthly.append({
            "key": k, "label": _month_label(k), "count": n,
            "total": t, "total_fmt": fmt_inr(t),
            "paid_fmt": fmt_inr(p),
            "unpaid": u, "unpaid_fmt": fmt_inr(u),
        })

    # window totals (over the last `months`)
    t_total, t_paid, t_unpaid = sum(tot.values()), sum(paid.values()), sum(unpaid.values())
    t_count = sum(counts.values())
    this_key = _month_key(today)
    this_month = tot.get(this_key, 0.0)
    active_months = sum(1 for k in keys if tot[k] > 0) or 1
    monthly_avg = t_total / active_months

    # ── breakdown by expense head (all-time, so a head that was quiet this
    #    window still shows) ──
    head_rows = db.query(
        VasyExpense.party_name,
        func.count(VasyExpense.id),
        func.coalesce(func.sum(VasyExpense.total), 0),
        func.coalesce(func.sum(VasyExpense.unpaid), 0),
    ).group_by(VasyExpense.party_name).order_by(func.sum(VasyExpense.total).desc()).all()
    grand = sum(float(t) for _, _, t, _ in head_rows) or 1.0
    heads = [{
        "head": name, "count": n,
        "total": float(t), "total_fmt": fmt_inr(t),
        "pct": round(float(t) / grand * 100, 1),
        "unpaid_fmt": fmt_inr(u),
    } for name, n, t, u in head_rows[:25]]

    # cash vs bank split, where the Expense Register's Payment Data covered it
    cash_known = db.query(
        func.coalesce(func.sum(VasyExpense.cash_paid), 0),
        func.coalesce(func.sum(VasyExpense.noncash_paid), 0),
    ).filter(VasyExpense.cash_paid != None).one()  # noqa: E711
    cash_amt, bank_amt = float(cash_known[0]), float(cash_known[1])
    has_mode = (cash_amt + bank_amt) > 0.005

    return {
        "has_data": True,
        "monthly": monthly,
        "heads": heads,
        "head_count": len(head_rows),
        "totals": {
            "total_fmt": fmt_inr(t_total),
            "paid_fmt": fmt_inr(t_paid),
            "unpaid_fmt": fmt_inr(t_unpaid),
            "count": t_count,
            "this_month_fmt": fmt_inr(this_month),
            "this_month_label": _month_label(this_key),
            "monthly_avg_fmt": fmt_inr(monthly_avg),
        },
        "mode": {
            "has_mode": has_mode,
            "cash_fmt": fmt_inr(cash_amt),
            "bank_fmt": fmt_inr(bank_amt),
        },
        "coverage": _date_range(db, VasyExpense.expense_date),
    }


# ── P3-6 Manager daily digest ──────────────────────────────────────────────

def manager_digest(db: Session, today: date) -> dict:
    """P3-6 — compose the manager's daily digest: today's pulse (sales +
    collections), total outstanding, at-risk count + top names, top collection
    chases, and churn risk. Returns a WhatsApp-friendly text + the structured
    parts (for a preview page). Pure/testable — sending is done by the caller.
    """
    from orderr_core.config import PLANT_NAME

    pulse = business_pulse(db, today)
    money = money_pulse(db, today)
    ci = credit_intelligence(db, today)
    churn = churn_risk(db, today)

    today_sales = next((p for p in pulse["periods"] if p["key"] == "today"), None)
    today_coll = next((p for p in money["periods"] if p["key"] == "today"), None) if money.get("has_data") else None

    at_risk_rows = sorted([r for r in ci.get("rows", []) if r["at_risk"]],
                          key=lambda r: r["score"], reverse=True) if ci.get("has_data") else []
    chases = sorted([r for r in ci.get("rows", []) if r["outstanding"] > 0],
                    key=lambda r: r["outstanding"] * r["score"] / 100, reverse=True)[:5] \
        if ci.get("has_data") else []

    lines = [f"📊 {PLANT_NAME} — Daily Digest", today.strftime("%d %b %Y"), ""]
    lines.append(f"Sales today: {today_sales['sales_rupees_fmt'] if today_sales else '₹0'}"
                 f" · {today_sales['invoices'] if today_sales else 0} invoices")
    if today_coll is not None:
        lines.append(f"Collected today: {today_coll['collections_fmt']}")
    if money.get("has_data"):
        lines.append(f"Outstanding (AR): {money['total_outstanding_fmt']}")
    lines.append("")

    # Margin guard
    try:
        fin = plant_financials(db, today)
        if fin.get("has_data") and fin["totals"].get("margin_alert"):
            lines.append(f"📉 Gross margin {fin['totals']['margin_pct']}% — below "
                         f"{fin['totals']['margin_threshold']:.0f}% floor. Review rates vs cost.")
            lines.append("")
    except Exception:
        pass

    # Wastage today (internal PLANT WASTAGE / WORKERS DAILY FOOD)
    try:
        w = wastage(db, today, days=30)
        if w.get("today_qty", 0) > 0:
            lines.append(f"🗑️ Wastage today: {fmt_qty(w['today_qty'])} kg "
                         f"({w['pct_of_volume']}% of volume, 30d)")
            lines.append("")
    except Exception:
        pass

    # Staff on leave / late today
    try:
        from orderr_core.models.leave import Leave
        from orderr_core.models.late_mark import LateMark
        from orderr_core.models.employee import Employee
        tstr = today.strftime("%Y-%m-%d")
        on_leave = [n for (n,) in db.query(Employee.name)
                    .join(Leave, Leave.employee_id == Employee.id)
                    .filter(Leave.date == tstr).all()]
        late = [n for (n,) in db.query(Employee.name)
                .join(LateMark, LateMark.employee_id == Employee.id)
                .filter(LateMark.date == tstr).all()]
        if on_leave:
            lines.append(f"🏖️ On leave today: {', '.join(on_leave)}")
        if late:
            lines.append(f"⏰ Late today: {', '.join(late)}")
        if on_leave or late:
            lines.append("")
    except Exception:
        pass

    if ci.get("has_data"):
        lines.append(f"⚠️ At-risk customers: {ci['at_risk']} ({ci['breach']} over limit)")
        for r in at_risk_rows[:3]:
            reason = r["reasons"][0] if r["reasons"] else f"risk {r['score']}"
            lines.append(f"  • {r['name']} — {r['outstanding_fmt']} ({reason})")
        lines.append("")
        if chases:
            lines.append("📞 Top collection calls:")
            for r in chases[:5]:
                lines.append(f"  • {r['name']} — {r['outstanding_fmt']}"
                             + (f" ☎ {r['phone']}" if r["phone"] else ""))
            lines.append("")

    if churn.get("rows"):
        lines.append(f"🔻 Silent-churn risk: {len(churn['rows'])} customers "
                     f"({churn['high']} high)")

    # Registers & Reminders — attention section (counts + worst three)
    try:
        from orderr_core.services import reminders_service
        att = reminders_service.attention_digest_lines(db, today)
        if att:
            if lines[-1] != "":
                lines.append("")
            lines.extend(att)
    except Exception:
        pass

    text = "\n".join(lines).strip()
    return {
        "text": text,
        "date_display": today.strftime("%d %b %Y"),
        "at_risk": ci.get("at_risk", 0),
        "breach": ci.get("breach", 0),
        "chase_count": len(chases),
        "churn_count": len(churn.get("rows", [])),
        "has_money": money.get("has_data", False),
    }


# ── Wastage tracking (PLANT WASTAGE / WORKERS DAILY FOOD internal accounts) ──

# Normalized party keys of the internal "non-customer" accounts (see the
# ₹0-internal split in the sales importer). party_key = UPPER + strip non-alnum.
WASTAGE_PARTY_KEYS = ["PLANTWASTAGE", "WORKERSDAILYFOOD"]


def wastage(db: Session, today: date, days: int = 30) -> dict:
    """Chicken booked to the internal PLANT WASTAGE / WORKERS DAILY FOOD accounts
    — a direct margin leak. Returns totals (kg), per-account + per-SKU splits, a
    daily trend, today's figure, and wastage as a % of total volume handled."""
    from orderr_core.models.vasy_sales_item import VasySalesItem

    start = today - timedelta(days=days - 1)
    rows = (db.query(VasySalesItem)
            .filter(VasySalesItem.party_key.in_(WASTAGE_PARTY_KEYS),
                    VasySalesItem.invoice_date != None,           # noqa: E711
                    VasySalesItem.invoice_date >= start,
                    VasySalesItem.invoice_date <= today).all())

    by_party, by_product, by_day = {}, {}, {}
    total_qty = 0.0
    for it in rows:
        qty = float(it.qty or 0)
        total_qty += qty
        by_party[it.party_name or "?"] = by_party.get(it.party_name or "?", 0.0) + qty
        by_product[it.product_name or "?"] = by_product.get(it.product_name or "?", 0.0) + qty
        if it.invoice_date:
            by_day[it.invoice_date] = by_day.get(it.invoice_date, 0.0) + qty

    # Total volume handled in the window (all sales lines) → wastage as a %.
    total_billed_qty = float(db.query(func.coalesce(func.sum(VasySalesItem.qty), 0))
                             .filter(VasySalesItem.invoice_date >= start,
                                     VasySalesItem.invoice_date <= today).scalar() or 0)
    pct = round(total_qty / total_billed_qty * 100, 2) if total_billed_qty else 0.0

    series = [{"date": (start + timedelta(days=i)).strftime("%d %b"),
               "qty": round(by_day.get(start + timedelta(days=i), 0.0), 1)}
              for i in range(days)]

    return {
        "has_data": bool(rows),
        "days": days,
        "total_qty": round(total_qty, 1),
        "today_qty": round(by_day.get(today, 0.0), 1),
        "pct_of_volume": pct,
        "by_party": [{"party": p, "qty": round(q, 1)}
                     for p, q in sorted(by_party.items(), key=lambda kv: kv[1], reverse=True)],
        "by_product": [{"product": p, "qty": round(q, 1)}
                       for p, q in sorted(by_product.items(), key=lambda kv: kv[1], reverse=True)],
        "series": series,
    }


# ── P3-7 Collection chase list ("call today") ──────────────────────────────

def chase_list(db: Session, today: date, top_n: int = 20) -> dict:
    """P3-7 — auto-ranked "call today" list: outstanding × risk = ₹-at-risk.

    Reuses the credit-intelligence signals; ranks debtors (outstanding > 0) by
    expected ₹-at-risk (outstanding × score/100) so the biggest, riskiest
    exposures surface first. Top `top_n` flagged as today's calls.
    """
    ci = credit_intelligence(db, today)
    if not ci.get("has_data"):
        return {"has_data": False}

    rows = []
    for r in ci["rows"]:
        if r["outstanding"] <= 0:
            continue
        priority = round(r["outstanding"] * r["score"] / 100, 2)
        rows.append({
            "customer_id": r["customer_id"],
            "name": r["name"], "phone": r["phone"],
            "area": r["area"], "salesperson": r["salesperson"],
            "outstanding": r["outstanding"], "outstanding_fmt": r["outstanding_fmt"],
            "score": r["score"], "classification": r["classification"],
            "days_since_payment": r["days_since_payment"],
            "breach": r["breach"],
            "priority": priority, "priority_fmt": fmt_inr(priority),
            "reason": " · ".join(r["reasons"][:2]) if r["reasons"] else "",
        })
    rows.sort(key=lambda r: r["priority"], reverse=True)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
        r["call_today"] = i < top_n

    today_value = sum(r["priority"] for r in rows[:top_n])
    total_value = sum(r["priority"] for r in rows)
    return {
        "has_data": True,
        "as_of": ci.get("as_of"),
        "rows": rows,
        "top_n": top_n,
        "call_today_count": min(top_n, len(rows)),
        "today_value_fmt": fmt_inr(today_value),
        "total_at_risk_fmt": fmt_inr(total_value),
        "areas": ci["areas"], "salespeople": ci["salespeople"],
    }


# ── P2-15/16 Vasy revenue + OrdeRR↔Vasy reconciliation ─────────────────────

def reconciliation(db: Session, target_date: date = None, today: date = None) -> dict:
    """P2-16 — flag OrdeRR deliveries with no matching Vasy invoice (billing
    leakage), for a given date, matching by customer.

    Also surfaces the day's Vasy-invoiced revenue (P2-15 authoritative revenue)
    and the reverse list (Vasy-invoiced customers with no OrdeRR order that day
    — e.g. phone/manual orders). Match is customer_id + exact date; the order's
    business-date basis vs Vasy invoice date can differ by a day, so this is a
    same-day approximation (labelled).
    """
    if today is None:
        from orderr_core.dates import get_current_business_date
        today = get_current_business_date()
    dates = [d[0] for d in db.query(VasyInvoice.invoice_date).distinct()
             .filter(VasyInvoice.invoice_date != None)                 # noqa: E711
             .order_by(VasyInvoice.invoice_date.desc()).all()]
    if not dates:
        return {"has_data": False}
    # default to the latest invoice date that isn't in the future (skip stray
    # post-dated invoices), else the most recent available
    target = target_date or next((d for d in dates if d <= today), dates[0])
    ds = target.strftime("%Y-%m-%d")

    # customers + phone→id map
    customers = {c.id: c for c in db.query(Customer).all()}
    phone_to_id = {}
    for c in customers.values():
        if c.phone_number:
            phone_to_id[c.phone_number] = c.id

    # Vasy invoices that day, grouped by customer
    vasy = db.query(VasyInvoice).filter(VasyInvoice.invoice_date == target).all()
    vasy_by_cust = {}
    vasy_total = 0.0
    for v in vasy:
        vasy_total += float(v.total or 0)
        if v.customer_id is not None:
            vasy_by_cust.setdefault(v.customer_id, []).append(v)

    # OrdeRR orders that day (non-cancelled), grouped by customer
    orders = db.query(Order).filter(
        Order.business_date == ds, Order.is_cancelled == False).all()  # noqa: E712
    orders_by_cust = {}
    for o in orders:
        cid = phone_to_id.get(o.customer_phone)
        orders_by_cust.setdefault(cid, []).append(o)

    # unbilled: OrdeRR customer ordered that day but no Vasy invoice
    unbilled = []
    for cid, olist in orders_by_cust.items():
        if cid is None:
            continue  # order from a phone with no customer record
        if cid not in vasy_by_cust:
            c = customers.get(cid)
            items = 0
            for o in olist:
                items += len(safe_list(o.parsed_items))
            unbilled.append({
                "customer_id": cid,
                "name": (c.restaurant_name if c else None) or o.customer_phone,
                "area": (c.area if c else "") or "",
                "orders": len(olist),
                "items": items,
            })
    unbilled.sort(key=lambda r: r["name"].lower())

    # reverse: Vasy invoiced that day but no OrdeRR order
    no_order = []
    for cid, vlist in vasy_by_cust.items():
        if cid not in orders_by_cust:
            c = customers.get(cid)
            tot = sum(float(v.total or 0) for v in vlist)
            no_order.append({
                "customer_id": cid,
                "name": (c.restaurant_name if c else None) or vlist[0].party_name,
                "vouchers": ", ".join(v.voucher_no for v in vlist),
                "total_fmt": fmt_inr(tot),
            })
    no_order.sort(key=lambda r: r["name"].lower())

    matched = sum(1 for cid in orders_by_cust if cid in vasy_by_cust)

    return {
        "has_data": True,
        "date": ds,
        "date_display": target.strftime("%d %b %Y"),
        "available_dates": [d.strftime("%Y-%m-%d") for d in dates],
        "vasy_invoice_count": len(vasy),
        "vasy_total_fmt": fmt_inr(vasy_total),
        "ordered_customers": len([c for c in orders_by_cust if c is not None]),
        "matched": matched,
        "unbilled": unbilled,
        "no_order": no_order,
    }


# ── P2-8/9/11/12 Receivables (AR exposure, debtors, aging proxy) ───────────

def _receipt_stats_by_customer(db: Session):
    """{customer_id: (last_date, count, total)} over matched receipts."""
    rows = (
        db.query(CustomerReceipt.customer_id,
                 func.max(CustomerReceipt.receipt_date),
                 func.count(CustomerReceipt.id),
                 func.coalesce(func.sum(CustomerReceipt.amount), 0))
        .filter(CustomerReceipt.customer_id != None)   # noqa: E711
        .group_by(CustomerReceipt.customer_id).all()
    )
    return {cid: (last, cnt, float(tot)) for cid, last, cnt, tot in rows}


def receivables(db: Session, today: date) -> dict:
    """P2-8/9/11/12 — receivables from the latest outstanding snapshot fused
    with receipt history.

    Per debtor (closing > 0): outstanding, balance direction vs previous
    snapshot (P2-9), days-since-last-payment (P2-11), last-payment date, avg
    receipt. Plus total AR + top-10 concentration (P2-8) and aging bucketed by
    days-since-last-payment (P2-12). This is a RUNNING-ACCOUNT business — a single
    balance per customer that lump-sum payments reduce, with no invoice-level
    settlement — so days-since-last-payment is the correct aging, not a stopgap.
    """
    latest_snap, prev_snap = _latest_two_snapshot_dates(db)
    if latest_snap is None:
        return {"has_data": False}

    snaps = db.query(OutstandingSnapshot).filter(
        OutstandingSnapshot.snapshot_date == latest_snap).all()
    prev_closing = {}
    if prev_snap:
        prev_closing = {s.party_key: float(s.closing) for s in db.query(OutstandingSnapshot)
                        .filter(OutstandingSnapshot.snapshot_date == prev_snap).all()}
    rstats = _receipt_stats_by_customer(db)
    sp_name = {s.id: s.name for s in db.query(Salesperson).all()}
    cust = {c.id: c for c in db.query(Customer).all()}
    bad_ids = _bad_debt_ids(db)

    rows = []
    total_ar = 0.0
    credit_total = 0.0
    bad_debt_total = 0.0
    bad_debt_customers = set()
    areas, salespeople = set(), set()
    aging = {"0-30": 0.0, "31-60": 0.0, "61-90": 0.0, "90+": 0.0, "no-payment": 0.0}

    for s in snaps:
        closing = float(s.closing)
        if s.customer_id in bad_ids:
            # written off as bad debt — out of AR, tallied separately
            if closing > 0:
                bad_debt_total += closing
                bad_debt_customers.add(s.customer_id)
            continue
        if closing < 0:
            credit_total += closing
        if closing <= 0:
            continue
        total_ar += closing

        c = cust.get(s.customer_id) if s.customer_id else None
        area = (c.area if c else "") or ""
        sp = (sp_name.get(c.salesperson_id) if c and c.salesperson_id else "") or ""
        if area:
            areas.add(area)
        if sp:
            salespeople.add(sp)

        last_date = count = avg = None
        days_since = None
        if s.customer_id in rstats:
            last_date, count, tot = rstats[s.customer_id]
            avg = tot / count if count else 0
            if last_date:
                # clamp at 0: a receipt dated after `today` (future-dated import)
                # must not read as negative days-since-payment.
                days_since = max(0, (today - last_date).days)

        # direction vs previous snapshot
        if prev_snap and s.party_key in prev_closing:
            pc = prev_closing[s.party_key]
            direction = "up" if closing > pc else "down" if closing < pc else "flat"
        else:
            direction = None

        # aging proxy bucket (by payment recency)
        if days_since is None:
            aging["no-payment"] += closing
        elif days_since <= 30:
            aging["0-30"] += closing
        elif days_since <= 60:
            aging["31-60"] += closing
        elif days_since <= 90:
            aging["61-90"] += closing
        else:
            aging["90+"] += closing

        rows.append({
            "customer_id": s.customer_id,
            "name": (c.restaurant_name if c else None) or s.party_name,
            "area": area, "salesperson": sp,
            "outstanding": round(closing, 2), "outstanding_fmt": fmt_inr(closing),
            "direction": direction,
            "last_payment": last_date.strftime("%Y-%m-%d") if last_date else "",
            "last_payment_display": last_date.strftime("%d %b %Y") if last_date else "never",
            "days_since_payment": days_since if days_since is not None else "",
            "avg_receipt_fmt": fmt_inr(avg) if avg is not None else "—",
        })

    rows.sort(key=lambda r: r["outstanding"], reverse=True)
    top10 = sum(r["outstanding"] for r in rows[:10])
    top5 = sum(r["outstanding"] for r in rows[:5])

    # Debtors with no payment on record (biggest first) — a collection blind
    # spot: unaged in the recency buckets, so surfaced separately.
    no_payment_rows = [r for r in rows if r["days_since_payment"] == ""]
    no_payment_total = round(sum(r["outstanding"] for r in no_payment_rows), 2)

    return {
        "has_data": True,
        "as_of": latest_snap.strftime("%d %b %Y"),
        "has_trend": prev_snap is not None,
        "prev_as_of": prev_snap.strftime("%d %b %Y") if prev_snap else None,
        "rows": rows,
        "debtor_count": len(rows),
        "total_ar": round(total_ar, 2), "total_ar_fmt": fmt_inr(total_ar),
        "credit_total_fmt": fmt_inr(round(credit_total, 2)),
        "bad_debt_total": round(bad_debt_total, 2),
        "bad_debt_total_fmt": fmt_inr(round(bad_debt_total, 2)),
        "bad_debt_count": len(bad_debt_customers),
        "no_payment_count": len(no_payment_rows),
        "no_payment_total_fmt": fmt_inr(no_payment_total),
        "no_payment_pct": round(no_payment_total / total_ar * 100, 1) if total_ar else 0,
        "no_payment_top": no_payment_rows[:10],
        "top10_fmt": fmt_inr(top10), "top5_fmt": fmt_inr(top5),
        "top10_pct": round(top10 / total_ar * 100, 1) if total_ar else 0,
        "top5_pct": round(top5 / total_ar * 100, 1) if total_ar else 0,
        "aging": [
            {"bucket": "0–30 days", "amount": round(aging["0-30"], 2), "amount_fmt": fmt_inr(aging["0-30"])},
            {"bucket": "31–60 days", "amount": round(aging["31-60"], 2), "amount_fmt": fmt_inr(aging["31-60"])},
            {"bucket": "61–90 days", "amount": round(aging["61-90"], 2), "amount_fmt": fmt_inr(aging["61-90"])},
            {"bucket": "90+ days", "amount": round(aging["90+"], 2), "amount_fmt": fmt_inr(aging["90+"])},
            {"bucket": "No payment on record", "amount": round(aging["no-payment"], 2), "amount_fmt": fmt_inr(aging["no-payment"])},
        ],
        "areas": sorted(areas), "salespeople": sorted(salespeople),
    }


# ── P1-2 Customer 360 (sales side) ─────────────────────────────────────────

# Windows offered in the C360 period selector. days=None → all time.
C360_WINDOWS = {"7": 7, "30": 30, "90": 90, "all": None}


def customer_360(db: Session, today: date, days=30) -> dict:
    """P1-2 — one sales-side row per customer.

    Per customer: revenue (₹, selected window, from OrdeRR invoices), order
    count in the window, days-since-last-order (recency, all-time), and a
    product-mix summary (top SKUs by quantity in the window, ERP display names).

    `days` bounds the revenue / order-count / mix window; None = all time.
    Recency is always all-time (a customer's true last order). Grouped queries
    keep this to a handful of round-trips regardless of customer count.
    """
    window_start = None if days is None else today - timedelta(days=days - 1)
    ws = window_start.strftime("%Y-%m-%d") if window_start else None
    today_str = today.strftime("%Y-%m-%d")

    # salesperson id → name
    sp_name = {s.id: s.name for s in db.query(Salesperson).all()}

    # ── P2-7 payment enrichment: outstanding + receipt behaviour per customer ──
    latest_snap, prev_snap = _latest_two_snapshot_dates(db)
    snap_now = _closing_by_customer(db, latest_snap)    # summed per customer
    snap_prev = _closing_by_customer(db, prev_snap)
    rstats = _receipt_stats_by_customer(db)
    has_money = bool(latest_snap) or bool(rstats)

    # revenue + invoice count by customer within window (Vasy = revenue truth)
    revenue = _vasy_revenue_by_customer(db, window_start, today)
    vc_q = db.query(VasyInvoice.customer_id, func.count(VasyInvoice.id)).filter(
        VasyInvoice.customer_id != None, VasyInvoice.invoice_date <= today)  # noqa: E711
    if window_start:
        vc_q = vc_q.filter(VasyInvoice.invoice_date >= window_start)
    inv_win = {cid: n for cid, n in vc_q.group_by(VasyInvoice.customer_id).all()}

    # last invoice date (all time) per customer
    last_q = db.query(
        VasyInvoice.customer_id, func.max(VasyInvoice.invoice_date)
    ).filter(VasyInvoice.customer_id != None).group_by(VasyInvoice.customer_id).all()  # noqa: E711
    last_inv = {cid: d for cid, d in last_q if d}

    # product mix within window — parsed_items is JSON, so aggregate in Python
    mix_q = db.query(Order.customer_phone, Order.parsed_items).filter(
        Order.is_cancelled == False,  # noqa: E712
        Order.business_date <= today_str,
    )
    if ws:
        mix_q = mix_q.filter(Order.business_date >= ws)
    mix = {}  # phone → {erp_name: qty}
    for ph, parsed in mix_q.all():
        bucket = mix.setdefault(ph, {})
        for item in safe_list(parsed):
            if not isinstance(item, dict):
                continue
            name = erp_display_name(item.get("product", "") or "Unknown")
            try:
                qty = float(item.get("quantity") or 0)
            except (TypeError, ValueError):
                qty = 0
            bucket[name] = bucket.get(name, 0) + qty

    rows = []
    areas, salespeople = set(), set()
    for c in db.query(Customer).all():
        ph = c.phone_number
        rev = revenue.get(c.id, 0.0)
        n_orders = inv_win.get(c.id, 0)          # invoice count (Vasy)
        last = last_inv.get(c.id)                # date or None
        recency_days = (today - last).days if last else None

        # top-3 products by quantity → summary (from OrdeRR orders — operational)
        prod_qty = mix.get(ph, {})
        top = sorted(prod_qty.items(), key=lambda kv: kv[1], reverse=True)[:3]
        mix_summary = ", ".join(
            f"{n} ({fmt_qty(q)})" for n, q in top
        ) if top else "—"

        sp = sp_name.get(c.salesperson_id) if c.salesperson_id else None
        area = c.area or None
        if area:
            areas.add(area)
        if sp:
            salespeople.add(sp)

        # payment enrichment
        outstanding = snap_now.get(c.id)
        bal_dir = None
        if c.id in snap_now and c.id in snap_prev:
            bal_dir = ("up" if snap_now[c.id] > snap_prev[c.id]
                       else "down" if snap_now[c.id] < snap_prev[c.id] else "flat")
        last_pay = last_pay_days = avg_receipt = None
        if c.id in rstats:
            lp, cnt, tot = rstats[c.id]
            avg_receipt = tot / cnt if cnt else 0
            if lp:
                last_pay = lp
                last_pay_days = (today - lp).days

        rows.append({
            "customer_id": c.id,
            "phone": ph or "",
            "name": c.restaurant_name or (ph or f"#{c.id}"),
            "area": area or "",
            "salesperson": sp or "",
            "is_active": bool(c.is_active),
            "revenue": rev,
            "revenue_fmt": fmt_inr(rev),
            "orders": n_orders,
            "last_order": last.strftime("%Y-%m-%d") if last else "",
            "recency_days": recency_days if recency_days is not None else "",
            "mix_summary": mix_summary,
            # P2-7 payment columns
            "outstanding": round(outstanding, 2) if outstanding is not None else "",
            "outstanding_fmt": fmt_inr(outstanding) if outstanding is not None else "—",
            "balance_dir": bal_dir,
            "last_payment_display": last_pay.strftime("%d %b %Y") if last_pay else "—",
            "days_since_payment": last_pay_days if last_pay_days is not None else "",
            "avg_receipt_fmt": fmt_inr(avg_receipt) if avg_receipt is not None else "—",
        })

    # default sort: revenue desc, then most-recent
    rows.sort(key=lambda r: (r["revenue"], -(r["recency_days"] if isinstance(r["recency_days"], int) else 10**9)), reverse=True)

    return {
        "rows": rows,
        "areas": sorted(areas),
        "salespeople": sorted(salespeople),
        "days": days,
        "window_label": "All time" if days is None else f"Last {days} days",
        "has_money": has_money,   # whether Vasy payment columns have data
    }


def _norm_name(s) -> str:
    """Name-join key: uppercase, strip every non-alphanumeric char. Mirrors
    vasy_import.normalize_name so this report reasons about the same keys the
    importer used."""
    import re
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


def customer_split_report(db: Session, today: date) -> dict:
    """Data-health: find customers split across two OrdeRR records — one holding
    the Vasy *invoices* (revenue) and another holding the *outstanding/receipts*
    (AR) — because the sales-invoice export's party name/phone didn't match the
    customer master (the PATILVADA / "PATILVADA HOTEL" case).

    A record is 'invoice-only' if it has invoices but no outstanding and no
    receipts; 'AR-only' if it has outstanding or receipts but no invoices. We
    pair an invoice-only record to an AR-only record when one's normalized name
    contains the other's (e.g. PATILVADA ⊂ PATILVADAHOTEL) — the exact signature
    of an import name-mismatch. Read-only; suggests merges, changes nothing.
    """
    sp_name = {s.id: s.name for s in db.query(Salesperson).all()}

    revenue = _vasy_revenue_by_customer(db, None, today)          # {cid: ₹ all-time}
    inv_cnt = {cid: n for cid, n in db.query(
        VasyInvoice.customer_id, func.count(VasyInvoice.id)).filter(
        VasyInvoice.customer_id != None).group_by(VasyInvoice.customer_id).all()}  # noqa: E711
    last_inv = {cid: d for cid, d in db.query(
        VasyInvoice.customer_id, func.max(VasyInvoice.invoice_date)).filter(
        VasyInvoice.customer_id != None).group_by(VasyInvoice.customer_id).all() if d}  # noqa: E711

    latest_snap, _prev = _latest_two_snapshot_dates(db)
    snap_now = _closing_by_customer(db, latest_snap)    # summed per customer
    rstats = _receipt_stats_by_customer(db)

    invoice_only, ar_only = [], []
    for c in db.query(Customer).all():
        rev = revenue.get(c.id, 0.0)
        ninv = inv_cnt.get(c.id, 0)
        out = snap_now.get(c.id)
        rec = rstats.get(c.id)
        has_sales = ninv > 0
        has_ar = (out is not None and round(out, 2) != 0) or rec is not None
        info = {
            "id": c.id,
            "name": c.restaurant_name or (c.phone_number or f"#{c.id}"),
            "norm": _norm_name(c.restaurant_name),
            "phone": c.phone_number or "",
            "area": c.area or "",
            "salesperson": (sp_name.get(c.salesperson_id) if c.salesperson_id else "") or "",
            "revenue": round(rev, 2), "revenue_fmt": fmt_inr(rev),
            "invoices": ninv,
            "last_invoice": last_inv[c.id].strftime("%d %b %Y") if c.id in last_inv else "never",
            "outstanding": round(out, 2) if out is not None else "",
            "outstanding_fmt": fmt_inr(out) if out is not None else "—",
            "last_payment": rec[0].strftime("%d %b %Y") if rec and rec[0] else "never",
        }
        if has_sales and not has_ar:
            invoice_only.append(info)
        elif has_ar and not has_sales:
            ar_only.append(info)

    # pair invoice-only ↔ AR-only when one normalized name is a PREFIX of the
    # other (PATILVADA→PATILVADAHOTEL, BABA→BABACHINESE, 712→712HOTEL). Prefix is
    # tighter than substring, so mid-name coincidences don't create bad merges;
    # the shorter key must be ≥3 chars to avoid 1–2 char false hits.
    pairs, used_inv = [], set()
    for ar in ar_only:
        an = ar["norm"]
        if len(an) < 3:
            continue
        match = None
        for iv in invoice_only:
            ivn = iv["norm"]
            if iv["id"] in used_inv or len(ivn) < 3:
                continue
            if an.startswith(ivn) or ivn.startswith(an):
                match = iv
                break
        if match:
            used_inv.add(match["id"])
            pairs.append({
                "invoice_side": match,   # holds revenue, usually no phone (auto-created)
                "ar_side": ar,           # holds outstanding + phone + salesperson
                "combined_out_fmt": ar["outstanding_fmt"],
                "combined_rev_fmt": match["revenue_fmt"],
            })

    pairs.sort(key=lambda p: (p["ar_side"]["outstanding"] or 0), reverse=True)
    orphan_invoice = [iv for iv in invoice_only if iv["id"] not in used_inv]
    orphan_invoice.sort(key=lambda r: r["revenue"], reverse=True)

    return {
        "has_data": bool(inv_cnt or snap_now),
        "as_of": latest_snap.strftime("%d %b %Y") if latest_snap else None,
        "pairs": pairs,
        "orphan_invoice": orphan_invoice,   # invoice-only, no AR twin found
        "counts": {
            "pairs": len(pairs),
            "invoice_only": len(invoice_only),
            "ar_only": len(ar_only),
            "orphan_invoice": len(orphan_invoice),
        },
    }


def merge_customer_split(db: Session, invoice_customer_id: int, ar_customer_id: int) -> dict:
    """Permanently merge a split: move all money rows from the invoice-only
    record into the AR record, record name aliases so future Vasy imports
    auto-link (never re-split), then delete the now-empty duplicate.

    `invoice_customer_id` is the auto-created invoice-holding record; it is
    absorbed into `ar_customer_id` (the phone/outstanding-bearing real customer).
    No customer_id column here has a unique constraint, so the re-point can't
    collide. Approval-based — the Data-health screen calls this per pair.
    """
    from orderr_core.models.customer_alias import CustomerAlias
    from orderr_core.models.vasy_sales_item import VasySalesItem

    if invoice_customer_id == ar_customer_id:
        raise ValueError("Cannot merge a customer into itself.")
    src = db.get(Customer, invoice_customer_id)   # absorbed
    dst = db.get(Customer, ar_customer_id)         # canonical
    if src is None or dst is None:
        raise ValueError("Both customers must exist.")

    # 1) record aliases (the permanent fix): the duplicate's normalized name and
    #    every party_key its invoices carried → the canonical customer.
    keys = set()
    if src.restaurant_name:
        keys.add(_norm_name(src.restaurant_name))
    for (pk,) in (db.query(VasyInvoice.party_key)
                  .filter(VasyInvoice.customer_id == invoice_customer_id).distinct().all()):
        if pk:
            keys.add(pk)
    for k in keys:
        row = db.query(CustomerAlias).filter_by(alias_key=k).first()
        if row:
            row.customer_id = ar_customer_id
        else:
            db.add(CustomerAlias(alias_key=k, customer_id=ar_customer_id,
                                 source="data-health-merge"))

    # 2) re-point money rows src → dst (the 4 tables that carry customer_id)
    moved = {
        "invoices": db.query(VasyInvoice).filter_by(customer_id=invoice_customer_id)
                      .update({"customer_id": ar_customer_id}, synchronize_session=False),
        "sales_items": db.query(VasySalesItem).filter_by(customer_id=invoice_customer_id)
                         .update({"customer_id": ar_customer_id}, synchronize_session=False),
        "receipts": db.query(CustomerReceipt).filter_by(customer_id=invoice_customer_id)
                      .update({"customer_id": ar_customer_id}, synchronize_session=False),
        "outstanding": db.query(OutstandingSnapshot).filter_by(customer_id=invoice_customer_id)
                         .update({"customer_id": ar_customer_id}, synchronize_session=False),
    }

    # 3) delete the now-empty duplicate
    absorbed_name = src.restaurant_name
    db.delete(src)
    db.commit()

    return {
        "status": "ok",
        "absorbed_id": invoice_customer_id,
        "absorbed_name": absorbed_name,
        "into_id": ar_customer_id,
        "into_name": dst.restaurant_name,
        "aliases": sorted(keys),
        "moved": moved,
    }


# ── P1-3 Customer detail ───────────────────────────────────────────────────

def _month_key(d: date) -> str:
    return d.strftime("%Y-%m")


def _last_n_months(today: date, n: int = 12):
    """List of the last n month keys ('YYYY-MM'), oldest → newest, ending at
    today's month."""
    ym = today.year * 12 + (today.month - 1)
    keys = []
    for i in range(n - 1, -1, -1):
        m = ym - i
        keys.append(f"{m // 12:04d}-{(m % 12) + 1:02d}")
    return keys


def _month_label(key: str) -> str:
    y, m = key.split("-")
    return date(int(y), int(m), 1).strftime("%b %y")


def _reliability_score(cur_out, days_since_pay, utilization_pct, pay_months, tenure_months):
    """Payment-reliability score (0–100) + grade A–E, for the running-account
    model. Components: payment recency (50%), credit utilisation (30%), payment
    consistency = paid in most active months (20%). Single source of truth for
    both the per-customer profile and the portfolio map.
    Returns (score, grade, components_dict)."""
    if days_since_pay is None:
        c_recency = 0 if cur_out > 0 else 60
    else:
        c_recency = max(0, min(100, round(100 - (days_since_pay / 75.0) * 100)))
    c_util = 60 if utilization_pct is None else max(0, min(100, round(100 - max(0.0, utilization_pct - 25) / 1.25)))
    c_consistency = max(0, min(100, round(pay_months / tenure_months * 100))) if tenure_months else 60
    score = round(0.5 * c_recency + 0.3 * c_util + 0.2 * c_consistency)
    grade = ("A" if score >= 80 else "B" if score >= 65 else "C" if score >= 50
             else "D" if score >= 35 else "E")
    return score, grade, {"recency": c_recency, "utilization": c_util, "consistency": c_consistency}


def credit_gate(db: Session, customer, today: date):
    """Assess whether posting an order for `customer` should warn the manager
    (over credit limit, or owes money with no recent payment). Running-account
    model. Returns (should_warn: bool, message: str)."""
    latest = db.query(func.max(OutstandingSnapshot.snapshot_date)).scalar()
    out = None
    if latest is not None:
        # sum, not .first() — a customer can own several Vasy ledgers (aliased outlets)
        row = (db.query(func.sum(OutstandingSnapshot.closing))
               .filter(OutstandingSnapshot.snapshot_date == latest,
                       OutstandingSnapshot.customer_id == customer.id).scalar())
        if row is not None:
            out = float(row)
    if out is None:
        out = float(customer.outstanding or 0)
    limit = float(customer.credit_limit) if customer.credit_limit else None
    lp = (db.query(func.max(CustomerReceipt.receipt_date))
          .filter(CustomerReceipt.customer_id == customer.id).scalar())
    days = (today - lp).days if lp else None

    over = bool(limit and out > limit)
    overdue = bool(out > 0 and (days is None or days > 45))
    if not (over or overdue):
        return False, ""

    name = customer.restaurant_name or "This customer"
    parts = [f"{name} owes {fmt_inr(out)}"]
    if over:
        parts.append(f"— over their {fmt_inr(limit)} credit limit")
    if days is None and out > 0:
        parts.append("with no payment on record")
    elif days is not None and days > 45:
        parts.append(f"and last paid {days} days ago")
    return True, " ".join(parts) + ". Post the order anyway?"


def _lifecycle_stage(first_inv_days, last_inv_days, trajectory_pct):
    """Customer lifecycle from billing recency + 3-month spend trajectory.
    Shared by the per-customer profile and the lifecycle/cohorts screen."""
    if first_inv_days is None:
        return "No billing"
    if last_inv_days is not None and last_inv_days > 90:
        return "Lost"
    if last_inv_days is not None and last_inv_days > 45:
        return "Dormant"
    if first_inv_days <= 60:
        return "New"
    if trajectory_pct is not None and trajectory_pct >= 15:
        return "Growing"
    if trajectory_pct is not None and trajectory_pct <= -15:
        return "Declining"
    return "Mature"


def customer_detail(db: Session, customer_id: int, today: date, months: int = 12):
    """P1-3 — full sales-side detail for one customer.

    Returns None if the customer id is unknown. Otherwise: header/KPIs, a
    12-month revenue trend (from OrdeRR invoices, gaps filled with 0), an
    all-time product mix (ERP names), and order + invoice history.
    """
    customer = db.query(Customer).get(customer_id)
    if customer is None:
        return None

    phone = customer.phone_number
    sp = None
    if customer.salesperson_id:
        sp_row = db.query(Salesperson).get(customer.salesperson_id)
        sp = sp_row.name if sp_row else None

    # ── invoices (revenue) — from Vasy sales invoices (source of truth) ──
    invoices = []
    monthly = {}  # 'YYYY-MM' → revenue
    total_revenue = 0.0
    inv_rows = (
        db.query(VasyInvoice)
        .filter(VasyInvoice.customer_id == customer.id)
        .order_by(VasyInvoice.invoice_date.desc())
        .all()
    )
    for inv in inv_rows:
        amt = float(inv.total or 0)
        total_revenue += amt
        if inv.invoice_date is not None:
            monthly[_month_key(inv.invoice_date)] = monthly.get(_month_key(inv.invoice_date), 0.0) + amt
        invoices.append({
            "number": inv.voucher_no,
            "date": inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else "",
            "date_display": inv.invoice_date.strftime("%d %b %Y") if inv.invoice_date else "",
            "total": amt,
            "total_fmt": fmt_inr(amt),
            "status": "",   # VasyInvoice header export carries no per-invoice status
        })

    # ── revenue trend (last N months, gaps = 0) ──
    trend = [{"key": k, "label": _month_label(k), "revenue": round(monthly.get(k, 0.0), 2)}
             for k in _last_n_months(today, months)]

    # ── orders + product mix ──
    orders = []
    mix = {}  # (erp_name, unit) → qty
    last_order = None
    first_order = None
    if phone:
        ord_rows = (
            db.query(Order)
            .filter(Order.customer_phone == phone, Order.is_cancelled == False)  # noqa: E712
            .order_by(Order.business_date.desc())
            .all()
        )
        for o in ord_rows:
            items = []
            for it in safe_list(o.parsed_items):
                if not isinstance(it, dict):
                    continue
                name = erp_display_name(it.get("product", "") or "Unknown")
                unit = (it.get("unit", "kg") or "kg").lower()
                try:
                    qty = float(it.get("quantity") or 0)
                except (TypeError, ValueError):
                    qty = 0
                items.append({"name": name, "qty": fmt_qty(qty), "unit": unit})
                mix[(name, unit)] = mix.get((name, unit), 0) + qty
            orders.append({
                "date": o.business_date or "",
                "date_display": (date.fromisoformat(o.business_date).strftime("%d %b %Y")
                                 if o.business_date else ""),
                "status": o.status or "",
                "line_items": items,
            })
        dates = [o.business_date for o in ord_rows if o.business_date]
        if dates:
            last_order, first_order = max(dates), min(dates)

    mix_list = [{"name": n, "unit": u, "qty": fmt_qty(q), "qty_raw": q}
                for (n, u), q in sorted(mix.items(), key=lambda kv: kv[1], reverse=True)]

    # ── P2-14 payments & balance history ──
    receipts = []
    collected = 0.0
    rec_rows = (db.query(CustomerReceipt)
                .filter(CustomerReceipt.customer_id == customer.id)
                .order_by(CustomerReceipt.receipt_date.desc(),
                          CustomerReceipt.receipt_no.desc()).all())
    for r in rec_rows:
        amt = float(r.amount or 0)
        collected += amt
        receipts.append({
            "receipt_no": r.receipt_no,
            "date_display": r.receipt_date.strftime("%d %b %Y") if r.receipt_date else "",
            "mode": (r.mode or "").title(),
            "amount_fmt": fmt_inr(amt),
            "status": (r.status or "").title(),
        })
    snap_rows = (db.query(OutstandingSnapshot)
                 .filter(OutstandingSnapshot.customer_id == customer.id)
                 .order_by(OutstandingSnapshot.snapshot_date).all())
    balance_trend = [{"date": s.snapshot_date.strftime("%Y-%m-%d"),
                      "label": s.snapshot_date.strftime("%d %b"),
                      "closing": round(float(s.closing), 2)} for s in snap_rows]
    current_outstanding = float(snap_rows[-1].closing) if snap_rows else None
    last_payment = rec_rows[0].receipt_date if rec_rows else None
    n_rec = len(rec_rows)

    # ── P3-9 share-of-wallet / upsell ──
    bought = {n for (n, _u) in mix.keys()}
    peer = {}                       # erp_name → set(customer_phone) who order it
    all_orderers = set()
    for ph2, parsed2 in (db.query(Order.customer_phone, Order.parsed_items)
                         .filter(Order.is_cancelled == False).all()):  # noqa: E712
        if ph2:
            all_orderers.add(ph2)
        seen = set()
        for it in safe_list(parsed2):
            if isinstance(it, dict):
                seen.add(erp_display_name(it.get("product", "") or "Unknown"))
        for nm in seen:
            peer.setdefault(nm, set()).add(ph2)
    n_orderers = max(1, len(all_orderers))
    suggestions = sorted(((nm, len(s)) for nm, s in peer.items() if nm not in bought),
                         key=lambda x: x[1], reverse=True)[:5]
    catalog = [it["erp_name"] for it in ERP_ITEMS.values()]
    catalog_set = set(catalog)
    sow_pct = round(len(bought & catalog_set) / len(catalog_set) * 100) if catalog_set else 0
    upsell = {
        "share_of_wallet_pct": sow_pct,
        "bought_count": len(bought & catalog_set),
        "catalog_count": len(catalog_set),
        "suggestions": [{"name": nm, "peer_count": cnt,
                         "peer_pct": round(cnt / n_orderers * 100)} for nm, cnt in suggestions],
    }

    recency_days = (today - date.fromisoformat(last_order)).days if last_order else None
    n_invoices = len(invoices)
    avg_order_value = (total_revenue / n_invoices) if n_invoices else 0.0

    # ── Deep intelligence: credit reliability + customer behaviour ─────────
    from orderr_core.models.vasy_sales_item import VasySalesItem

    cur_out = current_outstanding if current_outstanding is not None else float(customer.outstanding or 0)
    credit_limit = float(customer.credit_limit) if customer.credit_limit else None
    utilization_pct = round(cur_out / credit_limit * 100, 1) if (credit_limit and credit_limit > 0) else None
    days_since_pay = (today - last_payment).days if last_payment else None

    inv_dates = sorted([i.invoice_date for i in inv_rows if i.invoice_date])
    tenure_days = (today - inv_dates[0]).days if inv_dates else 0
    tenure_months = max(1, round(tenure_days / 30)) if tenure_days else 0
    gaps = [(inv_dates[k] - inv_dates[k - 1]).days for k in range(1, len(inv_dates))]
    avg_gap_days = round(sum(gaps) / len(gaps), 1) if gaps else None
    invoices_per_month = round(n_invoices / tenure_months, 1) if tenure_months else None

    def _rev_between(d0, d1):
        return sum(float(i.total or 0) for i in inv_rows
                   if i.invoice_date and d0 <= i.invoice_date < d1)
    last3 = _rev_between(today - timedelta(days=90), today + timedelta(days=1))
    prev3 = _rev_between(today - timedelta(days=180), today - timedelta(days=90))
    trajectory_pct = (round((last3 - prev3) / prev3 * 100, 1) if prev3 > 0
                      else (100.0 if last3 > 0 else None))

    avg_daily_bill = last3 / 90.0
    dso_days = round(cur_out / avg_daily_bill) if (avg_daily_bill > 0 and cur_out > 0) else None

    last_inv_days = (today - inv_dates[-1]).days if inv_dates else None
    first_inv_days = (today - inv_dates[0]).days if inv_dates else None
    lifecycle = _lifecycle_stage(first_inv_days, last_inv_days, trajectory_pct)

    bal_dir = None
    if len(snap_rows) >= 2:
        first_c, last_c = float(snap_rows[0].closing), float(snap_rows[-1].closing)
        bal_dir = "up" if last_c > first_c else "down" if last_c < first_c else "flat"
    deteriorating = bool(cur_out > 0 and (days_since_pay is None or days_since_pay > 45)
                         and bal_dir in (None, "up"))

    # Payment-reliability score — running-account model (see _reliability_score).
    pay_months = len({(r.receipt_date.year, r.receipt_date.month) for r in rec_rows if r.receipt_date})
    score, grade, _comps = _reliability_score(cur_out, days_since_pay, utilization_pct, pay_months, tenure_months)

    sku_mix = []
    sk_rows = (db.query(
                 func.coalesce(VasySalesItem.product_name, VasySalesItem.item_code, "Unknown"),
                 func.coalesce(func.sum(VasySalesItem.qty), 0),
                 func.coalesce(func.sum(VasySalesItem.net_amount), 0))
               .filter(VasySalesItem.customer_id == customer.id)
               .group_by(func.coalesce(VasySalesItem.product_name, VasySalesItem.item_code, "Unknown"))
               .all())
    sk_total = sum(float(v) for _, _, v in sk_rows) or 0.0
    for nm, q, v in sorted(sk_rows, key=lambda x: float(x[2]), reverse=True):
        sku_mix.append({
            "product": str(nm), "qty_fmt": fmt_qty(float(q)),
            "value_fmt": fmt_inr(float(v)),
            "pct": round(float(v) / sk_total * 100, 1) if sk_total else 0.0,
        })

    intelligence = {
        "score": score, "grade": grade,
        "score_components": _comps,
        "dso_days": dso_days,
        "credit_limit_fmt": (fmt_inr(credit_limit) if credit_limit else None),
        "utilization_pct": utilization_pct,
        "current_outstanding_fmt": fmt_inr(cur_out),
        "days_since_payment": days_since_pay if days_since_pay is not None else None,
        "lifecycle": lifecycle,
        "trajectory_pct": trajectory_pct,
        "avg_gap_days": avg_gap_days,
        "invoices_per_month": invoices_per_month,
        "tenure_months": tenure_months,
        "balance_direction": bal_dir,
        "deteriorating": deteriorating,
        "avg_order_value_fmt": fmt_inr(avg_order_value),
    }

    return {
        "customer": {
            "id": customer.id,
            "name": customer.restaurant_name or (phone or f"#{customer.id}"),
            "phone": phone or "",
            "area": customer.area or "",
            "salesperson": sp or "",
            "is_active": bool(customer.is_active),
            "outstanding_fmt": fmt_inr(customer.outstanding or 0),
        },
        "kpis": {
            "total_revenue_fmt": fmt_inr(total_revenue),
            "orders": len(orders),
            "invoices": n_invoices,
            "avg_order_value_fmt": fmt_inr(avg_order_value),
            "recency_days": recency_days if recency_days is not None else "",
            "last_order_display": (date.fromisoformat(last_order).strftime("%d %b %Y")
                                   if last_order else "never"),
            "first_order_display": (date.fromisoformat(first_order).strftime("%d %b %Y")
                                    if first_order else "—"),
        },
        "intelligence": intelligence,
        "sku_mix": sku_mix,
        "trend": trend,
        "mix": mix_list,
        "upsell": upsell,
        "orders": orders,
        "invoices": invoices,
        "payments": {
            "has_data": bool(rec_rows) or bool(snap_rows),
            "current_outstanding_fmt": fmt_inr(current_outstanding) if current_outstanding is not None else "—",
            "collected_fmt": fmt_inr(collected),
            "receipt_count": n_rec,
            "avg_receipt_fmt": fmt_inr(collected / n_rec) if n_rec else "—",
            "last_payment_display": last_payment.strftime("%d %b %Y") if last_payment else "never",
            "days_since_payment": (today - last_payment).days if last_payment else "",
            "receipts": receipts,
            "balance_trend": balance_trend,
        },
    }


# ── shared: distinct order dates per customer phone ────────────────────────

def _order_dates_by_phone(db: Session, upto: date):
    """{phone: [date, …] sorted ascending} of DISTINCT non-cancelled order
    business-dates up to and including `upto`. Malformed date strings skipped."""
    rows = (
        db.query(Order.customer_phone, Order.business_date)
        .filter(Order.is_cancelled == False,  # noqa: E712
                Order.business_date != None,   # noqa: E711
                Order.business_date <= upto.strftime("%Y-%m-%d"))
        .all()
    )
    out = {}
    for ph, ds in rows:
        try:
            d = date.fromisoformat(ds)
        except (TypeError, ValueError):
            continue
        out.setdefault(ph, set()).add(d)
    return {ph: sorted(s) for ph, s in out.items()}


# ── Phase 4.2 — Value × Risk portfolio map ─────────────────────────────────

def portfolio(db: Session, today: date) -> dict:
    """Every transacting customer placed in a 2×2 by all-time revenue (value) and
    payment-reliability score (risk): protect / watch / maintain / tighten. The
    firm-wide view that drives daily credit decisions. Reuses _reliability_score
    (same scoring as the per-customer profile). Batch queries — no per-customer
    detail calls.
    """
    customers = {c.id: c for c in db.query(Customer).all()}
    sp_name = {s.id: s.name for s in db.query(Salesperson).all()}
    bad_recs = {b.customer_id: b for b in db.query(BadDebt).all()}
    rev = _vasy_revenue_by_customer(db, None, today)   # {cid: all-time revenue}

    inv_span = {}
    for cid, lo, _hi in (db.query(VasyInvoice.customer_id,
                                  func.min(VasyInvoice.invoice_date),
                                  func.max(VasyInvoice.invoice_date))
                         .filter(VasyInvoice.customer_id != None)          # noqa: E711
                         .group_by(VasyInvoice.customer_id).all()):
        inv_span[cid] = lo

    last_pay = {cid: d for cid, d in
                (db.query(CustomerReceipt.customer_id, func.max(CustomerReceipt.receipt_date))
                 .filter(CustomerReceipt.customer_id != None)             # noqa: E711
                 .group_by(CustomerReceipt.customer_id).all())}

    pay_months = {}
    for cid, rd in (db.query(CustomerReceipt.customer_id, CustomerReceipt.receipt_date)
                    .filter(CustomerReceipt.customer_id != None,          # noqa: E711
                            CustomerReceipt.receipt_date != None).all()):  # noqa: E711
        pay_months.setdefault(cid, set()).add((rd.year, rd.month))

    latest_snap = db.query(func.max(OutstandingSnapshot.snapshot_date)).scalar()
    out_by = {}
    if latest_snap is not None:
        for cid, closing in (db.query(OutstandingSnapshot.customer_id, OutstandingSnapshot.closing)
                             .filter(OutstandingSnapshot.snapshot_date == latest_snap,
                                     OutstandingSnapshot.customer_id != None).all()):  # noqa: E711
            out_by[cid] = float(closing)

    rows = []
    bad_rows = []
    for cid, c in customers.items():
        revenue = float(rev.get(cid, 0.0))
        cur_out = out_by.get(cid, float(c.outstanding or 0))
        if cur_out < 0:
            cur_out = 0.0                               # credit balance → no AR risk
        if cid in bad_recs:
            # written off as bad debt — kept out of the risk map, listed apart
            b = bad_recs[cid]
            bad_rows.append({
                "customer_id": cid,
                "name": c.restaurant_name or (c.phone_number or f"#{cid}"),
                "area": c.area or "",
                "salesperson": (sp_name.get(c.salesperson_id, "") if c.salesperson_id else ""),
                "amount": float(b.amount or 0), "amount_fmt": fmt_inr(b.amount),
                "current_outstanding_fmt": fmt_inr(cur_out),
                "reason": b.reason, "note": b.note or "",
                "written_off_on": b.written_off_on.strftime("%d %b %Y") if b.written_off_on else "",
            })
            continue
        if revenue <= 0 and cur_out <= 0:
            continue                                    # never transacted → skip
        lo = inv_span.get(cid)
        tenure_months = max(1, round((today - lo).days / 30)) if lo else 0
        lp = last_pay.get(cid)
        days_since = (today - lp).days if lp else None
        credit_limit = float(c.credit_limit) if c.credit_limit else None
        util = round(cur_out / credit_limit * 100, 1) if (credit_limit and credit_limit > 0) else None
        pm = len(pay_months.get(cid, set()))
        score, grade, _ = _reliability_score(cur_out, days_since, util, pm, tenure_months)
        rows.append({
            "customer_id": cid,
            "name": c.restaurant_name or (c.phone_number or f"#{cid}"),
            "area": c.area or "", "salesperson": (sp_name.get(c.salesperson_id, "") if c.salesperson_id else ""),
            "revenue": round(revenue, 2), "revenue_fmt": fmt_inr(revenue),
            "outstanding": round(cur_out, 2), "outstanding_fmt": fmt_inr(cur_out),
            "score": score, "grade": grade,
            "days_since_payment": days_since if days_since is not None else "",
        })

    revs = sorted(r["revenue"] for r in rows)
    median_rev = revs[len(revs) // 2] if revs else 0.0
    RISK_CUT = 55
    quad_def = {
        "watch":    {"label": "Watch closely",  "sub": "High value · High risk", "accent": "red"},
        "protect":  {"label": "Protect",        "sub": "High value · Low risk",  "accent": "good"},
        "tighten":  {"label": "Tighten credit", "sub": "Low value · High risk",  "accent": "amber"},
        "maintain": {"label": "Maintain",       "sub": "Low value · Low risk",   "accent": "blue"},
    }
    for r in rows:
        hv = r["revenue"] >= median_rev
        hr = r["score"] < RISK_CUT
        r["quadrant"] = ("protect" if (hv and not hr) else "watch" if (hv and hr)
                         else "maintain" if (not hv and not hr) else "tighten")

    quads = []
    for key, meta in quad_def.items():
        members = [r for r in rows if r["quadrant"] == key]
        out_sum = sum(m["outstanding"] for m in members)
        quads.append({
            "key": key, "label": meta["label"], "sub": meta["sub"], "accent": meta["accent"],
            "count": len(members),
            "revenue_fmt": fmt_inr(sum(m["revenue"] for m in members)),
            "outstanding": out_sum, "outstanding_fmt": fmt_inr(out_sum),
        })

    rows.sort(key=lambda r: r["outstanding"], reverse=True)
    bad_rows.sort(key=lambda r: r["amount"], reverse=True)
    bad_total = sum(r["amount"] for r in bad_rows)

    # Bad debt grouped by salesperson (who carried the customers that went bad)
    bd_by_sp = {}
    for r in bad_rows:
        sp = r["salesperson"] or "Unassigned"
        e = bd_by_sp.setdefault(sp, {"salesperson": sp, "amount": 0.0, "count": 0})
        e["amount"] += r["amount"]
        e["count"] += 1
    bad_by_sp = sorted(bd_by_sp.values(), key=lambda x: x["amount"], reverse=True)
    for e in bad_by_sp:
        e["amount"] = round(e["amount"], 2)
        e["amount_fmt"] = fmt_inr(e["amount"])
        e["pct"] = round(e["amount"] / bad_total * 100) if bad_total else 0

    return {
        "rows": rows,
        "quadrants": quads,
        "total": len(rows),
        "median_rev_fmt": fmt_inr(median_rev),
        "risk_cut": RISK_CUT,
        "areas": sorted({r["area"] for r in rows if r["area"]}),
        "salespeople": sorted({r["salesperson"] for r in rows if r["salesperson"]}),
        "as_of": latest_snap.strftime("%d %b %Y") if latest_snap else None,
        "bad_debt": {
            "rows": bad_rows,
            "count": len(bad_rows),
            "total_fmt": fmt_inr(round(bad_total, 2)),
            "by_salesperson": bad_by_sp,
        },
    }


def write_off_bad_debt(db: Session, customer_id: int, reason: str, note: str, today: date):
    """Mark a customer's balance as bad debt (unrecoverable). Records the
    balance at write-off (latest snapshot sum, falling back to the scalar
    customer.outstanding) for the audit trail. OrdeRR-side overlay only — the
    Vasy ledger is untouched. Returns an error string, or None on success."""
    customer = db.query(Customer).get(customer_id)
    if customer is None:
        return "Customer not found."
    if db.query(BadDebt).filter_by(customer_id=customer_id).first():
        return "This customer is already written off."

    latest = db.query(func.max(OutstandingSnapshot.snapshot_date)).scalar()
    amount = None
    if latest is not None:
        row = (db.query(func.sum(OutstandingSnapshot.closing))
               .filter(OutstandingSnapshot.snapshot_date == latest,
                       OutstandingSnapshot.customer_id == customer_id).scalar())
        if row is not None:
            amount = float(row)
    if amount is None:
        amount = float(customer.outstanding or 0)
    if amount <= 0:
        return "This customer has no outstanding balance to write off."

    db.add(BadDebt(customer_id=customer_id, amount=round(amount, 2),
                   reason=(reason or "Other").strip() or "Other",
                   note=(note or "").strip() or None,
                   written_off_on=today))
    db.commit()
    return None


def undo_bad_debt(db: Session, customer_id: int):
    """Remove a bad-debt write-off — the customer's balance rejoins AR on the
    next screen load. Returns an error string, or None on success."""
    rec = db.query(BadDebt).filter_by(customer_id=customer_id).first()
    if rec is None:
        return "No bad-debt write-off on record for this customer."
    db.delete(rec)
    db.commit()
    return None


# ── Phase 4.3 — Payment behaviour: DSO, concentration, early warnings ───────

def payment_behaviour(db: Session, today: date) -> dict:
    """Portfolio-level payment health: DSO, AR concentration, an AR-over-time
    series, and a ranked 'deteriorating payers' early-warning list. Reuses
    receivables() for the debtor rows (outstanding, balance direction,
    days-since-payment)."""
    rec = receivables(db, today)
    if not rec.get("has_data"):
        return {"has_data": False}
    rows = rec["rows"]                          # debtors, sorted desc by outstanding
    total_ar = rec["total_ar"] or 0.0

    # Portfolio DSO = total AR ÷ average daily billing (trailing 90 days).
    start90 = today - timedelta(days=90)
    bill90 = float(db.query(func.coalesce(func.sum(VasyInvoice.total), 0))
                   .filter(VasyInvoice.invoice_date >= start90,
                           VasyInvoice.invoice_date <= today,
                           VasyInvoice.total > 0).scalar() or 0)
    avg_daily = bill90 / 90.0
    dso = round(total_ar / avg_daily) if avg_daily > 0 else None

    # Concentration: how many customers make up 80% of AR.
    cum = 0.0
    n80 = 0
    for r in rows:
        cum += r["outstanding"]
        n80 += 1
        if total_ar and cum >= 0.8 * total_ar:
            break
    top20 = sum(r["outstanding"] for r in rows[:20])

    # AR over time (total positive balances per snapshot date, bad debt excluded
    # throughout so the series matches the netted Total AR KPI).
    bad_ids = _bad_debt_ids(db)
    ar_series = []
    for (d,) in (db.query(OutstandingSnapshot.snapshot_date).distinct()
                 .order_by(OutstandingSnapshot.snapshot_date).all()):
        tot = _outstanding_total(db, d, bad_ids)
        ar_series.append({"date": d.strftime("%d %b"), "amount": round(tot, 2), "amount_fmt": fmt_inr(tot)})

    # Deteriorating payers: rising balance and/or long silence, ranked by
    # exposure × staleness.
    watch = []
    for r in rows:
        dsp = r["days_since_payment"]
        d = 999 if dsp == "" else int(dsp)
        reasons = []
        if r.get("direction") == "up":
            reasons.append("balance rising")
        if dsp == "":
            reasons.append("no payment on record")
        elif d > 30:
            reasons.append(f"{d}d since payment")
        if not reasons:
            continue
        severity = r["outstanding"] * (1 + min(d, 120) / 60.0)
        watch.append({
            "customer_id": r["customer_id"], "name": r["name"],
            "area": r["area"], "salesperson": r["salesperson"],
            "outstanding_fmt": r["outstanding_fmt"], "outstanding": r["outstanding"],
            "days_since_payment": dsp, "last_payment_display": r["last_payment_display"],
            "direction": r.get("direction"), "reasons": ", ".join(reasons),
            "_sev": severity,
        })
    watch.sort(key=lambda x: x["_sev"], reverse=True)
    watch_exposure = sum(w["outstanding"] for w in watch)

    return {
        "has_data": True,
        "as_of": rec["as_of"],
        "total_ar_fmt": rec["total_ar_fmt"], "debtor_count": rec["debtor_count"],
        "dso": dso, "bill90_fmt": fmt_inr(bill90),
        "top5_pct": rec["top5_pct"], "top10_pct": rec["top10_pct"],
        "top20_pct": round(top20 / total_ar * 100, 1) if total_ar else 0,
        "n80": n80,
        "ar_series": ar_series,
        "watch": watch[:50],
        "watch_count": len(watch),
        "watch_exposure_fmt": fmt_inr(watch_exposure),
        "areas": rec["areas"], "salespeople": rec["salespeople"],
    }


# ── Phase 4.4 — Lifecycle stages & acquisition cohorts ─────────────────────

def lifecycle_cohorts(db: Session, today: date) -> dict:
    """Customer lifecycle distribution, biggest spend movers (3-mo trajectory),
    and acquisition cohorts (first-invoice month → how many still active)."""
    customers = {c.id: c for c in db.query(Customer).all()}
    rev = _vasy_revenue_by_customer(db, None, today)          # all-time revenue

    span = {}
    for cid, lo, hi in (db.query(VasyInvoice.customer_id,
                                 func.min(VasyInvoice.invoice_date),
                                 func.max(VasyInvoice.invoice_date))
                        .filter(VasyInvoice.customer_id != None)          # noqa: E711
                        .group_by(VasyInvoice.customer_id).all()):
        span[cid] = (lo, hi)

    last3 = _vasy_revenue_by_customer(db, today - timedelta(days=90), today)
    prev3 = {}
    for cid, amt in (db.query(VasyInvoice.customer_id, func.sum(VasyInvoice.total))
                     .filter(VasyInvoice.customer_id != None,             # noqa: E711
                             VasyInvoice.invoice_date >= today - timedelta(days=180),
                             VasyInvoice.invoice_date < today - timedelta(days=90))
                     .group_by(VasyInvoice.customer_id).all()):
        prev3[cid] = float(amt or 0)

    STAGES = ["New", "Growing", "Mature", "Declining", "Dormant", "Lost"]
    ACCENT = {"New": "blue", "Growing": "good", "Mature": "", "Declining": "amber",
              "Dormant": "amber", "Lost": "red"}
    stage_agg = {s: {"count": 0, "revenue": 0.0} for s in STAGES}
    movers = []
    cohorts = {}                     # 'YYYY-MM' -> {"acq": int, "active": int}

    for cid, c in customers.items():
        if cid not in span or not span[cid][0]:
            continue
        lo, hi = span[cid]
        first_inv_days = (today - lo).days
        last_inv_days = (today - hi).days
        l3, p3 = last3.get(cid, 0.0), prev3.get(cid, 0.0)
        traj = round((l3 - p3) / p3 * 100, 1) if p3 > 0 else (100.0 if l3 > 0 else None)
        stage = _lifecycle_stage(first_inv_days, last_inv_days, traj)
        if stage in stage_agg:
            stage_agg[stage]["count"] += 1
            stage_agg[stage]["revenue"] += float(rev.get(cid, 0.0))

        if traj is not None and abs(traj) >= 15 and (l3 > 0 or p3 > 0) and last_inv_days <= 45:
            movers.append({"customer_id": cid, "name": c.restaurant_name or f"#{cid}",
                           "trajectory_pct": traj, "last3_fmt": fmt_inr(l3), "prev3_fmt": fmt_inr(p3)})

        ck = lo.strftime("%Y-%m")
        co = cohorts.setdefault(ck, {"acq": 0, "active": 0})
        co["acq"] += 1
        if last_inv_days <= 45:                      # still buying
            co["active"] += 1

    base = sum(s["count"] for s in stage_agg.values()) or 1
    stages = [{"stage": s, "accent": ACCENT[s], "count": stage_agg[s]["count"],
               "pct": round(stage_agg[s]["count"] / base * 100),
               "revenue_fmt": fmt_inr(stage_agg[s]["revenue"])} for s in STAGES]

    movers.sort(key=lambda m: m["trajectory_pct"], reverse=True)
    growing = [m for m in movers if m["trajectory_pct"] > 0][:10]
    shrinking = sorted([m for m in movers if m["trajectory_pct"] < 0],
                       key=lambda m: m["trajectory_pct"])[:10]

    cohort_rows = []
    for ck in sorted(cohorts.keys()):
        co = cohorts[ck]
        y, m = ck.split("-")
        cohort_rows.append({
            "label": date(int(y), int(m), 1).strftime("%b %Y"),
            "acquired": co["acq"], "active": co["active"],
            "retention_pct": round(co["active"] / co["acq"] * 100) if co["acq"] else 0,
        })

    return {
        "has_data": bool(span),
        "total": base if base > 1 else sum(s["count"] for s in stage_agg.values()),
        "stages": stages,
        "growing": growing, "shrinking": shrinking,
        "cohorts": cohort_rows,
    }


# ── P1-4 Silent-churn detector ─────────────────────────────────────────────

def churn_risk(db: Session, today: date, min_orders: int = 3,
               ratio_threshold: float = 2.0, floor_days: int = 3) -> dict:
    """P1-4 — customers overdue relative to their OWN billing cadence.

    Cadence = median gap (days) between a customer's Vasy invoice dates. A
    customer is flagged when days-since-last-invoice exceeds
    `ratio_threshold` × cadence (and at least `floor_days`). Needs `min_orders`
    invoices for a stable cadence; fewer-history customers excluded. Uses Vasy
    invoices (billing = actual business), not OrdeRR orders.

    ratio = days_since_last / cadence. Severity: ≥3 high, ≥2 medium.
    Returns only flagged customers, most-overdue (by ratio) first.
    """
    dates_by_customer = _vasy_invoice_dates_by_customer(db, today)
    sp_name = {s.id: s.name for s in db.query(Salesperson).all()}
    customers = {c.id: c for c in db.query(Customer).all()}

    rows = []
    areas, salespeople = set(), set()
    for cid, dates in dates_by_customer.items():
        cust = customers.get(cid)
        if cust is None or len(dates) < min_orders:
            continue
        gaps = [(dates[i] - dates[i - 1]).days for i in range(1, len(dates))]
        gaps = [g for g in gaps if g > 0] or gaps  # guard all-zero
        cadence = median(gaps) if gaps else None
        if not cadence or cadence <= 0:
            continue
        last = dates[-1]
        days_since = (today - last).days
        ratio = days_since / cadence
        if days_since < floor_days or ratio < ratio_threshold:
            continue

        severity = "high" if ratio >= 3 else "medium"
        sp = sp_name.get(cust.salesperson_id) if cust.salesperson_id else ""
        area = cust.area or ""
        if area:
            areas.add(area)
        if sp:
            salespeople.add(sp)

        rows.append({
            "customer_id": cust.id,
            "name": cust.restaurant_name or (cust.phone_number or f"#{cid}"),
            "phone": cust.phone_number or "",
            "area": area,
            "salesperson": sp,
            "orders": len(dates),
            "cadence_days": round(cadence, 1),
            "days_since_last": days_since,
            "last_order": last.strftime("%Y-%m-%d"),
            "last_order_display": last.strftime("%d %b %Y"),
            "ratio": round(ratio, 2),
            "severity": severity,
        })

    rows.sort(key=lambda r: r["ratio"], reverse=True)
    return {
        "rows": rows,
        "areas": sorted(areas),
        "salespeople": sorted(salespeople),
        "high": sum(1 for r in rows if r["severity"] == "high"),
        "medium": sum(1 for r in rows if r["severity"] == "medium"),
        "params": {"min_orders": min_orders, "ratio_threshold": ratio_threshold, "floor_days": floor_days},
    }


# ── P1-5 Revenue trend + MoM ───────────────────────────────────────────────

def _pct_change(curr: float, prev: float):
    """MoM % change; None when there is no prior base (prev == 0)."""
    if prev == 0:
        return None
    return round((curr - prev) / prev * 100, 1)


def revenue_trends(db: Session, today: date, months: int = 12) -> dict:
    """P1-5 — overall revenue over time (with MoM %) plus a per-customer
    current-vs-previous-month comparison.

    Revenue is Vasy sales-invoice totals (the source of truth), bucketed by
    invoice-month. Overall totals include unmatched invoices; the per-customer
    table covers matched customers only.
    """
    keys = _last_n_months(today, months)
    key_set = set(keys)

    curr_key = keys[-1]
    prev_key = keys[-2] if len(keys) >= 2 else None

    # Is the current calendar month still in progress? (today before month-end)
    days_in_month = calendar.monthrange(today.year, today.month)[1]
    current_partial = today.day < days_in_month

    # overall monthly totals + per-(customer,month) totals in one pass.
    # prev_mtd = previous-month revenue up to the SAME day-of-month as today, so
    # the current (partial) month is compared like-for-like, not against a full month.
    monthly = {k: 0.0 for k in keys}
    per_cust = {}  # customer_id → {month_key: revenue}
    per_cust_prev_mtd = {}  # customer_id → prev-month revenue up to today's day
    scan_start = date(int(keys[0][:4]), int(keys[0][5:]), 1)
    prev_mtd = 0.0

    inv_rows = (
        db.query(VasyInvoice.customer_id, VasyInvoice.invoice_date, VasyInvoice.total)
        .filter(VasyInvoice.invoice_date >= scan_start,
                VasyInvoice.invoice_date <= today)
        .all()
    )
    for cid, bdate, total in inv_rows:
        if bdate is None:
            continue
        mk = _month_key(bdate)
        amt = float(total or 0)
        if mk in monthly:
            monthly[mk] += amt
        if prev_key and mk == prev_key and bdate.day <= today.day:
            prev_mtd += amt
            if cid is not None:
                per_cust_prev_mtd[cid] = per_cust_prev_mtd.get(cid, 0.0) + amt
        if cid is not None:
            per_cust.setdefault(cid, {})[mk] = per_cust.get(cid, {}).get(mk, 0.0) + amt

    # overall trend with MoM (full-month vs full-month for history)
    trend = []
    prev_rev = None
    for k in keys:
        rev = round(monthly[k], 2)
        mom = _pct_change(rev, prev_rev) if prev_rev is not None else None
        trend.append({
            "key": k, "label": _month_label(k),
            "revenue": rev, "revenue_fmt": fmt_inr(rev), "mom_pct": mom,
            "partial": False,
        })
        prev_rev = rev

    # The current month is partial: its full-vs-full MoM overstates a "drop".
    # Restate it as MTD-vs-same-period-last-month and flag it for the UI.
    prev_mtd = round(prev_mtd, 2)
    if current_partial and trend:
        trend[-1]["partial"] = True
        trend[-1]["mom_pct"] = _pct_change(trend[-1]["revenue"], prev_mtd)

    # Trim leading months that precede the first month with any revenue, so the
    # chart doesn't show a run of ₹0 bars for months simply not yet imported.
    first_nz = next((i for i, t in enumerate(trend) if t["revenue"] > 0), 0)
    trend = trend[first_nz:]

    # per-customer curr vs prev month
    sp_name = {s.id: s.name for s in db.query(Salesperson).all()}
    customers = {c.id: c for c in db.query(Customer).all()}
    cust_rows = []
    areas, salespeople = set(), set()
    for cid, by_month in per_cust.items():
        curr = round(by_month.get(curr_key, 0.0), 2)
        # Compare like-for-like: when the current month is partial, the previous
        # side is that month up to the same day (not the whole month), so a
        # customer isn't shown as collapsing just because the month is young.
        if current_partial:
            prev = round(per_cust_prev_mtd.get(cid, 0.0), 2)
        else:
            prev = round(by_month.get(prev_key, 0.0), 2) if prev_key else 0.0
        if curr == 0 and prev == 0:
            continue
        cust = customers.get(cid)
        name = (cust.restaurant_name if cust else None) or f"#{cid}"
        area = (cust.area if cust else "") or ""
        sp = (sp_name.get(cust.salesperson_id) if cust and cust.salesperson_id else "") or ""
        if area:
            areas.add(area)
        if sp:
            salespeople.add(sp)
        pct = _pct_change(curr, prev)
        if prev == 0 and curr > 0:
            direction = "new"
        elif curr == 0 and prev > 0:
            direction = "lost"
        elif curr > prev:
            direction = "up"
        elif curr < prev:
            direction = "down"
        else:
            direction = "flat"
        cust_rows.append({
            "customer_id": cid,
            "name": name, "area": area, "salesperson": sp,
            "curr": curr, "curr_fmt": fmt_inr(curr),
            "prev": prev, "prev_fmt": fmt_inr(prev),
            "delta": round(curr - prev, 2), "delta_fmt": fmt_inr(curr - prev),
            "pct": pct, "direction": direction,
        })

    cust_rows.sort(key=lambda r: r["delta"], reverse=True)

    curr_total = round(monthly[curr_key], 2)
    prev_total = round(monthly[prev_key], 2) if prev_key else 0.0

    # Headline MoM: like-for-like when the month is partial (MTD vs prev-month
    # same-period), else full-vs-full. Run-rate projects the partial month to a
    # full-month estimate at the current daily pace.
    if current_partial:
        headline_mom = _pct_change(curr_total, prev_mtd)
        projection = round(curr_total / today.day * days_in_month, 2) if today.day else 0.0
    else:
        headline_mom = _pct_change(curr_total, prev_total)
        projection = curr_total

    return {
        "trend": trend,
        "current_label": _month_label(curr_key),
        "prev_label": _month_label(prev_key) if prev_key else "",
        "current_revenue_fmt": fmt_inr(curr_total),
        "prev_revenue_fmt": fmt_inr(prev_total),
        "current_mom_pct": headline_mom,
        "current_partial": current_partial,
        "day_of_month": today.day,
        "prev_mtd_fmt": fmt_inr(prev_mtd),
        "projection_fmt": fmt_inr(projection),
        "customers": cust_rows,
        "areas": sorted(areas),
        "salespeople": sorted(salespeople),
    }


# ── P1-6 New vs lost customers ─────────────────────────────────────────────

def new_vs_lost(db: Session, today: date, months: int = 12) -> dict:
    """P1-6 — monthly customer acquisitions vs attrition.

    Acquisition[m] = customers whose FIRST Vasy invoice month == m.
    Attrition[m]   = customers whose LAST invoice month == m, excluding the
                     current month (can't be declared lost mid-month).
    Derived from Vasy sales-invoice history (billing = actual business).
    """
    rows = (
        db.query(VasyInvoice.customer_id,
                 func.min(VasyInvoice.invoice_date), func.max(VasyInvoice.invoice_date))
        .filter(VasyInvoice.customer_id != None,     # noqa: E711
                VasyInvoice.invoice_date != None)    # noqa: E711
        .group_by(VasyInvoice.customer_id)
        .all()
    )

    keys = _last_n_months(today, months)
    key_set = set(keys)
    curr_key = _month_key(today)
    acq = {k: 0 for k in keys}
    att = {k: 0 for k in keys}
    acq_ids = {k: [] for k in keys}   # month → [(cid, date)] first invoice
    att_ids = {k: [] for k in keys}   # month → [(cid, date)] last invoice

    for cid, first_d, last_d in rows:
        try:
            fm = _month_key(first_d)
            lm = _month_key(last_d)
        except (TypeError, ValueError, AttributeError):
            continue
        if fm in key_set:
            acq[fm] += 1
            acq_ids[fm].append((cid, first_d))
        if lm in key_set and lm != curr_key:
            att[lm] += 1
            att_ids[lm].append((cid, last_d))

    # resolve customer display info for the drill-down lists
    sp_name = {s.id: s.name for s in db.query(Salesperson).all()}
    customers = {c.id: c for c in db.query(Customer).all()}

    def _people(pairs):
        out = []
        for cid, d in pairs:
            cust = customers.get(cid)
            out.append({
                "id": cid,
                "name": (cust.restaurant_name if cust else None) or f"#{cid}",
                "area": (cust.area if cust else "") or "",
                "salesperson": (sp_name.get(cust.salesperson_id)
                                if cust and cust.salesperson_id else "") or "",
                "date": d.strftime("%d %b %y") if d else "",
            })
        out.sort(key=lambda r: r["name"].lower())
        return out

    series = [{"key": k, "label": _month_label(k), "new": acq[k], "lost": att[k],
               "new_customers": _people(acq_ids[k]), "lost_customers": _people(att_ids[k]),
               "net": acq[k] - att[k]} for k in keys]

    # Drop leading months before the first with any activity (data-history edge:
    # months simply not yet imported shouldn't read as "zero acquisitions").
    first_active = next((i for i, s in enumerate(series) if s["new"] or s["lost"]), 0)
    series = series[first_active:]

    return {
        "series": series,
        "total_new": sum(acq.values()),
        "total_lost": sum(att.values()),
        "net": sum(acq.values()) - sum(att.values()),
    }


# ── P1-7 Product mix (value + volume) ──────────────────────────────────────

def product_mix(db: Session, today: date, days=30) -> dict:
    """P1-7 — per-SKU billed value (₹) and volume (qty) over a window, from the
    Vasy Sales Item Register (billed truth, source of truth). % of total value.
    `days` bounds the window; None = all time.

    Volume is a single QTY per SKU: the register has no unit column and units
    differ across SKUs (kg for cuts, nos for whole birds), so a per-SKU qty is
    meaningful while a cross-SKU volume total is not — hence value drives ranking.
    """
    from orderr_core.models.vasy_sales_item import VasySalesItem

    window_start = None if days is None else today - timedelta(days=days - 1)

    q = (
        db.query(
            func.coalesce(VasySalesItem.product_name, VasySalesItem.item_code, "Unknown"),
            func.coalesce(func.sum(VasySalesItem.qty), 0),
            func.coalesce(func.sum(VasySalesItem.net_amount), 0),
            func.count(VasySalesItem.id),
        )
        .filter(VasySalesItem.invoice_date != None)          # noqa: E711
        .filter(VasySalesItem.invoice_date <= today)
    )
    if window_start:
        q = q.filter(VasySalesItem.invoice_date >= window_start)
    q = q.group_by(func.coalesce(VasySalesItem.product_name, VasySalesItem.item_code, "Unknown"))

    agg = [(str(name), float(qty or 0), float(val or 0), int(n)) for name, qty, val, n in q.all()]
    total_value = sum(v for _, _, v, _ in agg) or 0.0

    rows = []
    for name, qty, val, n in agg:
        pct = round(val / total_value * 100, 1) if total_value else 0.0
        rows.append({
            "product": name,
            "qty": round(qty, 2), "qty_fmt": fmt_qty(qty),
            "lines": n,
            "value": round(val, 2), "value_fmt": fmt_inr(val),
            "pct": pct,
        })
    rows.sort(key=lambda x: x["value"], reverse=True)

    return {
        "rows": rows,
        "total_value": round(total_value, 2),
        "total_value_fmt": fmt_inr(total_value),
        "sku_count": len(rows),
        "days": days,
        "window_label": "All time" if days is None else f"Last {days} days",
    }


# ── P1-8 Demand trend (per SKU, over time) ─────────────────────────────────

def demand_trend(db: Session, today: date, months: int = 12) -> dict:
    """P1-8 — ordered demand per SKU per month (production-planning signal),
    from order parsed_items (what customers asked for, not what was billed).

    Returns month labels, and per-SKU monthly quantity series. Quantity is
    summed across units (SKUs are single-unit in practice); the dominant unit
    is reported per SKU for labelling.
    """
    keys = _last_n_months(today, months)
    scan_start = date(int(keys[0][:4]), int(keys[0][5:]), 1)
    key_set = set(keys)

    rows = (
        db.query(Order.business_date, Order.parsed_items)
        .filter(Order.is_cancelled == False,                 # noqa: E712
                Order.business_date != None,                 # noqa: E711
                Order.business_date >= scan_start.strftime("%Y-%m-%d"),
                Order.business_date <= today.strftime("%Y-%m-%d"))
        .all()
    )

    # sku → {month_key: qty}, sku → unit counts
    series = {}
    unit_votes = {}
    for bdate, parsed in rows:
        try:
            mk = _month_key(date.fromisoformat(bdate))
        except (TypeError, ValueError):
            continue
        if mk not in key_set:
            continue
        for it in safe_list(parsed):
            if not isinstance(it, dict):
                continue
            name = erp_display_name(it.get("product", "") or "Unknown")
            unit = (it.get("unit", "kg") or "kg").lower()
            try:
                qty = float(it.get("quantity") or 0)
            except (TypeError, ValueError):
                qty = 0
            series.setdefault(name, {}).setdefault(mk, 0.0)
            series[name][mk] += qty
            uv = unit_votes.setdefault(name, {})
            uv[unit] = uv.get(unit, 0) + 1

    skus = []
    for name, by_month in series.items():
        months_series = [{"key": k, "label": _month_label(k),
                          "qty": round(by_month.get(k, 0.0), 2)} for k in keys]
        total = sum(by_month.values())
        unit = max(unit_votes.get(name, {"kg": 1}).items(), key=lambda kv: kv[1])[0]
        skus.append({
            "product": name,
            "unit": unit,
            "total": round(total, 2),
            "total_fmt": fmt_qty(total),
            "series": months_series,
        })
    skus.sort(key=lambda s: s["total"], reverse=True)

    return {
        "months": [{"key": k, "label": _month_label(k)} for k in keys],
        "skus": skus,
    }


# ── P1-12 Excel export ─────────────────────────────────────────────────────

def build_xlsx(sheet_name: str, headers: list, rows: list) -> bytes:
    """Serialise headers + rows to .xlsx bytes (openpyxl). Bold header row,
    auto-ish column widths. Numbers stay numbers so the sheet is usable."""
    import io
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Sheet1")[:31]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    for i, h in enumerate(headers, start=1):
        width = max(len(str(h)), *(len(str(r[i - 1])) for r in rows)) if rows else len(str(h))
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(48, width + 2)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_dataset(db: Session, today: date, name: str, days=None,
                   area=None, salesperson=None):
    """Return (filename, sheet_name, headers, rows) for an analytics list, so
    the export route can serialise it. Mirrors what each screen shows.
    Optional area/salesperson narrow the rows to the on-screen selection.
    Unknown name → None.
    """
    tag = today.strftime("%Y%m%d")

    def _match(r):
        return (not area or r.get("area") == area) and \
               (not salesperson or r.get("salesperson") == salesperson)

    if name == "customers":
        data = customer_360(db, today, days=days if days is not None else 30)
        win = data.get("window_label", "")
        headers = ["Customer", "Area", "Salesperson",
                   f"Revenue ({win})".strip(), "Outstanding (INR)",
                   "Last invoice", "Last payment"]
        src = [r for r in data["rows"] if _match(r)]
        src.sort(key=lambda r: (r["outstanding"] if r["outstanding"] != "" else 0),
                 reverse=True)
        rows = [[r["name"], r["area"], r["salesperson"],
                 r["revenue"], r["outstanding"],
                 r["last_order"], r["last_payment_display"]] for r in src]
        total_rev = sum(r["revenue"] for r in src if r["revenue"] != "")
        total_out = sum(r["outstanding"] for r in src if r["outstanding"] != "")
        rows.append(["TOTAL", "", "", round(total_rev, 2), round(total_out, 2), "", ""])
        return (f"customers_{tag}.xlsx", "Customer 360", headers, rows)

    if name == "data-health":
        data = customer_split_report(db, today)
        headers = ["Invoice record", "Invoice cust #", "Revenue (INR)", "Invoices",
                   "Last invoice", "AR record", "AR cust #", "Phone", "Salesperson",
                   "Area", "Outstanding (INR)", "Last payment", "Suggested action"]
        rows = [[p["invoice_side"]["name"], p["invoice_side"]["id"],
                 p["invoice_side"]["revenue"], p["invoice_side"]["invoices"],
                 p["invoice_side"]["last_invoice"],
                 p["ar_side"]["name"], p["ar_side"]["id"], p["ar_side"]["phone"],
                 p["ar_side"]["salesperson"], p["ar_side"]["area"],
                 p["ar_side"]["outstanding"], p["ar_side"]["last_payment"],
                 f'Merge #{p["invoice_side"]["id"]} into #{p["ar_side"]["id"]}']
                for p in data["pairs"]]
        return (f"data_health_{tag}.xlsx", "Split customers", headers, rows)

    if name == "churn":
        data = churn_risk(db, today)
        headers = ["Customer", "Phone", "Area", "Salesperson", "Orders",
                   "Cadence (days)", "Days since last", "Last order", "Overdue x", "Severity"]
        rows = [[r["name"], r["phone"], r["area"], r["salesperson"], r["orders"],
                 r["cadence_days"], r["days_since_last"], r["last_order"],
                 r["ratio"], r["severity"]] for r in data["rows"]]
        return (f"churn_{tag}.xlsx", "Churn risk", headers, rows)

    if name == "revenue":
        data = revenue_trends(db, today)
        headers = ["Customer", "Area", "Salesperson",
                   f"{data['prev_label']} (INR)", f"{data['current_label']} (INR)",
                   "Delta (INR)", "MoM %", "Movement"]
        rows = [[r["name"], r["area"], r["salesperson"], r["prev"], r["curr"],
                 r["delta"], ("" if r["pct"] is None else r["pct"]), r["direction"]]
                for r in data["customers"]]
        return (f"revenue_mom_{tag}.xlsx", "Revenue MoM", headers, rows)

    if name == "products":
        data = product_mix(db, today, days=days if days is not None else 30)
        headers = ["SKU", "Qty", "Lines", "Value (INR)", "% of value"]
        rows = [[r["product"], r["qty"], r["lines"], r["value"], r["pct"]] for r in data["rows"]]
        return (f"product_mix_{tag}.xlsx", "Product mix", headers, rows)

    if name == "team":
        data = team_performance(db, today, days=days if days is not None else 30)
        headers = ["Group", "Name", "Revenue (INR)", "Collected (INR)",
                   "Outstanding (INR)", "Orders", "Active", "Portfolio"]
        rows = ([["Salesperson", r["name"], r["revenue"], r["collected"], r["outstanding"], r["orders"], r["active"], r["portfolio"]]
                 for r in data["by_salesperson"]] +
                [["Area", r["name"], r["revenue"], r["collected"], r["outstanding"], r["orders"], r["active"], r["portfolio"]]
                 for r in data["by_area"]])
        return (f"team_area_{tag}.xlsx", "Team & area", headers, rows)

    if name == "receivables":
        data = receivables(db, today)
        if not data.get("has_data"):
            return (f"receivables_{tag}.xlsx", "Receivables", ["Note"], [["No outstanding snapshot imported yet."]])
        headers = ["Customer", "Area", "Salesperson", "Outstanding (INR)", "Trend",
                   "Days since payment", "Last payment", "Avg receipt"]
        rows = [[r["name"], r["area"], r["salesperson"], r["outstanding"],
                 r["direction"] or "", r["days_since_payment"], r["last_payment"],
                 r["avg_receipt_fmt"]] for r in data["rows"]]
        return (f"receivables_{tag}.xlsx", "Receivables", headers, rows)

    if name == "chase":
        data = chase_list(db, today)
        if not data.get("has_data"):
            return (f"chase_{tag}.xlsx", "Chase", ["Note"], [["No money data imported yet."]])
        headers = ["Rank", "Call today", "Customer", "Phone", "Area", "Salesperson",
                   "Outstanding (INR)", "Risk", "Class", "Days since payment",
                   "Rs at risk", "Reason"]
        rows = [[r["rank"], "Yes" if r["call_today"] else "", r["name"], r["phone"],
                 r["area"], r["salesperson"], r["outstanding"], r["score"],
                 r["classification"], r["days_since_payment"], r["priority"], r["reason"]]
                for r in data["rows"]]
        return (f"chase_{tag}.xlsx", "Chase", headers, rows)

    if name == "credit":
        data = credit_intelligence(db, today)
        if not data.get("has_data"):
            return (f"credit_{tag}.xlsx", "Credit", ["Note"], [["No money data imported yet."]])
        headers = ["Customer", "Area", "Salesperson", "Outstanding (INR)", "Credit limit (INR)",
                   "Breach", "Risk score", "Classification", "At risk",
                   "Days since payment", "Exposure (months)", "Reasons"]
        rows = [[r["name"], r["area"], r["salesperson"], r["outstanding"],
                 (r["credit_limit"] if r["credit_limit"] != "" else ""),
                 "Yes" if r["breach"] else "", r["score"], r["classification"],
                 "Yes" if r["at_risk"] else "", r["days_since_payment"],
                 r["exposure_months"], "; ".join(r["reasons"])] for r in data["rows"]]
        return (f"credit_{tag}.xlsx", "Credit", headers, rows)

    if name == "rfm":
        data = rfm(db, today)
        headers = ["Customer", "Area", "Salesperson", "Recency (days)",
                   "Frequency", "Monetary (INR)", "R", "F", "M", "Segment"]
        rows = [[r["name"], r["area"], r["salesperson"], r["recency_days"],
                 r["frequency"], r["monetary"], r["R"], r["F"], r["M"], r["segment"]]
                for r in data["rows"]]
        return (f"rfm_{tag}.xlsx", "RFM", headers, rows)

    return None


# ── P1-13 RFM segmentation ─────────────────────────────────────────────────

def _quintile_scores(pairs, higher_better=True):
    """Map each (id, value) to a 1–5 score by rank-based quintile.
    higher_better=False inverts (used for recency, where fewer days = better).
    Robust to ties and tiny samples."""
    vals = sorted(v for _, v in pairs)
    n = len(vals) or 1
    out = {}
    for cid, v in pairs:
        rank = bisect.bisect_right(vals, v)        # count of values ≤ v (1..n)
        score = min(5, max(1, int((rank / n - 1e-9) * 5) + 1))
        out[cid] = (6 - score) if not higher_better else score
    return out


def _rfm_segment(r: int, f: int, m: int) -> str:
    """Name a segment from R/F/M scores (1–5). Standard RFM matrix, simplified
    and explainable."""
    fm = round((f + m) / 2)
    if r >= 4 and fm >= 4:
        return "Champions"
    if r >= 3 and fm >= 4:
        return "Loyal"
    if r >= 4 and fm >= 2:
        return "Potential loyalist"
    if r >= 4 and fm < 2:
        return "New"
    if r <= 2 and fm >= 4:
        return "Can't lose"
    if r <= 2 and fm == 3:
        return "At risk"
    if r <= 2 and fm <= 2:
        return "Hibernating"
    return "Needs attention"


# display order + accent for segments
RFM_SEGMENTS = [
    ("Champions", "good"), ("Loyal", "good"), ("Potential loyalist", "blue"),
    ("New", "blue"), ("Needs attention", "amber"), ("At risk", "amber"),
    ("Can't lose", "red"), ("Hibernating", "red"),
]


def rfm(db: Session, today: date) -> dict:
    """P1-13 — Recency/Frequency/Monetary scoring & segmentation.

    Recency = days since last Vasy invoice, Frequency = lifetime invoice count,
    Monetary = lifetime Vasy revenue. Each scored 1–5 by quintile across the
    active base; a named segment is derived from the trio. Uses Vasy invoices
    (billing = actual business), not OrdeRR orders. Only customers with at
    least one invoice are scored.
    """
    dates_by_customer = _vasy_invoice_dates_by_customer(db, today)
    revenue = _vasy_revenue_by_customer(db)              # all-time, by customer_id
    customers = {c.id: c for c in db.query(Customer).all()}
    sp_name = {s.id: s.name for s in db.query(Salesperson).all()}

    base = []  # (customer_id, recency_days, frequency, monetary)
    for cid, dates in dates_by_customer.items():
        if cid not in customers or not dates:
            continue
        recency = (today - dates[-1]).days
        base.append((cid, recency, len(dates), revenue.get(cid, 0.0)))

    if not base:
        return {"rows": [], "segments": [], "areas": [], "salespeople": [], "total": 0}

    r_scores = _quintile_scores([(p, rec) for p, rec, _, _ in base], higher_better=False)
    f_scores = _quintile_scores([(p, fr) for p, _, fr, _ in base], higher_better=True)
    m_scores = _quintile_scores([(p, mo) for p, _, _, mo in base], higher_better=True)

    rows = []
    areas, salespeople = set(), set()
    seg_counts = {}
    for cid, recency, frequency, monetary in base:
        cust = customers[cid]
        R, F, M = r_scores[cid], f_scores[cid], m_scores[cid]
        seg = _rfm_segment(R, F, M)
        seg_counts[seg] = seg_counts.get(seg, 0) + 1
        sp = (sp_name.get(cust.salesperson_id) if cust.salesperson_id else "") or ""
        area = cust.area or ""
        if area:
            areas.add(area)
        if sp:
            salespeople.add(sp)
        rows.append({
            "customer_id": cust.id,
            "name": cust.restaurant_name or (cust.phone_number or f"#{cid}"),
            "area": area, "salesperson": sp,
            "recency_days": recency, "frequency": frequency,
            "monetary": round(monetary, 2), "monetary_fmt": fmt_inr(monetary),
            "R": R, "F": F, "M": M, "rfm": f"{R}{F}{M}", "segment": seg,
        })
    rows.sort(key=lambda x: (x["R"] + x["F"] + x["M"], x["monetary"]), reverse=True)

    segments = [{"name": name, "accent": accent, "count": seg_counts.get(name, 0)}
                for name, accent in RFM_SEGMENTS if seg_counts.get(name, 0)]

    return {
        "rows": rows,
        "segments": segments,
        "areas": sorted(areas),
        "salespeople": sorted(salespeople),
        "total": len(rows),
    }


# ── P1-11 Salesperson & area performance (sales) ───────────────────────────

def team_performance(db: Session, today: date, days=30) -> dict:
    """P1-11 — sales rolled up by salesperson and by area.

    Per group: revenue (₹, window, from invoices), order count (window),
    active customers (ordered in window) and portfolio size (assigned
    customers). `days` bounds revenue/orders; None = all time.
    Customers with no salesperson/area roll into an "Unassigned" bucket.
    """
    window_start = None if days is None else today - timedelta(days=days - 1)
    ws = window_start.strftime("%Y-%m-%d") if window_start else None
    today_str = today.strftime("%Y-%m-%d")

    # revenue + invoice count by customer within window (Vasy = revenue truth)
    revenue = _vasy_revenue_by_customer(db, window_start, today)
    vc_q = db.query(VasyInvoice.customer_id, func.count(VasyInvoice.id)).filter(
        VasyInvoice.customer_id != None, VasyInvoice.invoice_date <= today)  # noqa: E711
    if window_start:
        vc_q = vc_q.filter(VasyInvoice.invoice_date >= window_start)
    inv_win = {cid: n for cid, n in vc_q.group_by(VasyInvoice.customer_id).all()}

    sp_name = {s.id: s.name for s in db.query(Salesperson).all()}

    # collections (receipts in window) by customer_id
    col_q = db.query(CustomerReceipt.customer_id, func.coalesce(func.sum(CustomerReceipt.amount), 0)) \
        .filter(CustomerReceipt.customer_id != None, CustomerReceipt.receipt_date <= today)  # noqa: E711
    if window_start:
        col_q = col_q.filter(CustomerReceipt.receipt_date >= window_start)
    collected = {cid: float(t) for cid, t in col_q.group_by(CustomerReceipt.customer_id).all()}

    # current outstanding (latest snapshot) by customer_id — summed per customer
    latest_snap, _ = _latest_two_snapshot_dates(db)
    outstanding = _closing_by_customer(db, latest_snap)

    def _bucket():
        return {"revenue": 0.0, "orders": 0, "active": 0, "portfolio": 0,
                "collected": 0.0, "outstanding": 0.0}
    by_sp, by_area = {}, {}

    for c in db.query(Customer).all():
        rev = revenue.get(c.id, 0.0)
        no = inv_win.get(c.id, 0)
        col = collected.get(c.id, 0.0)
        out = outstanding.get(c.id, 0.0)
        sp = (sp_name.get(c.salesperson_id) if c.salesperson_id else None) or "Unassigned"
        area = c.area or "Unassigned"
        for key, table in ((sp, by_sp), (area, by_area)):
            b = table.setdefault(key, _bucket())
            b["revenue"] += rev
            b["orders"] += no
            b["portfolio"] += 1
            b["collected"] += col
            b["outstanding"] += out
            if no > 0:
                b["active"] += 1

    def _rows(table):
        rows = [{
            "name": k,
            "revenue": round(v["revenue"], 2), "revenue_fmt": fmt_inr(v["revenue"]),
            "orders": v["orders"], "active": v["active"], "portfolio": v["portfolio"],
            "collected": round(v["collected"], 2), "collected_fmt": fmt_inr(v["collected"]),
            "outstanding": round(v["outstanding"], 2), "outstanding_fmt": fmt_inr(v["outstanding"]),
        } for k, v in table.items()]
        rows.sort(key=lambda r: r["revenue"], reverse=True)
        return rows

    return {
        "by_salesperson": _rows(by_sp),
        "by_area": _rows(by_area),
        "days": days,
        "window_label": "All time" if days is None else f"Last {days} days",
        "has_money": bool(collected) or bool(outstanding),
    }


# ── Volume (KG) sold by route/area, day-level windows ───────────────────────

VOLUME_WINDOWS = ("today", "yesterday", "7", "30", "all")


def _volume_window(today: date, window: str):
    """Resolve a volume window key → (start, end, label, normalized_window).
    Shared by volume_report and volume_breakdown so the day-level windows
    (today / yesterday / 7d / 30d / all) always mean the same span."""
    if window == "yesterday":
        y = today - timedelta(days=1)
        return y, y, f"Yesterday · {y.strftime('%d %b')}", "yesterday"
    if window == "7":
        return today - timedelta(days=6), today, "Last 7 days", "7"
    if window == "30":
        return today - timedelta(days=29), today, "Last 30 days", "30"
    if window == "all":
        return None, today, "All time", "all"
    return today, today, f"Today · {today.strftime('%d %b')}", "today"


def volume_report(db: Session, today: date, window: str = "today") -> dict:
    """KG sold by route (customer area) and by SKU, for a day-level window
    (today / yesterday / 7d / 30d / all), from the Vasy Sales Item Register.

    Everything is billed by weight (verified: every SKU shows fractional
    quantities), so a cross-SKU KG total is meaningful. ₹0 lines are EXCLUDED —
    they are internal transfers (Dead Bird, Plant Wastage, workers' food), not
    sales; their kg is reported separately as `internal_kg`. Sales Returns
    (negative-value lines) subtract from both kg and value.
    """
    from sqlalchemy import case
    from orderr_core.models.vasy_sales_item import VasySalesItem

    start, end, label, window = _volume_window(today, window)

    signed_qty = case((VasySalesItem.sale_type == "Sales Return", -VasySalesItem.qty),
                      else_=VasySalesItem.qty)

    def base(q):
        q = q.filter(VasySalesItem.invoice_date != None,          # noqa: E711
                     VasySalesItem.invoice_date <= end)
        if start:
            q = q.filter(VasySalesItem.invoice_date >= start)
        return q

    sold = VasySalesItem.net_amount != 0    # ₹0 = internal transfer, not a sale

    # by route/area (via the customer master; unmatched/blank → Unassigned)
    area_q = base(
        db.query(func.coalesce(Customer.area, "Unassigned"),
                 func.coalesce(func.sum(signed_qty), 0),
                 func.coalesce(func.sum(VasySalesItem.net_amount), 0),
                 func.count(func.distinct(VasySalesItem.voucher_no)))
        .outerjoin(Customer, Customer.id == VasySalesItem.customer_id)
        .filter(sold)
    ).group_by(func.coalesce(Customer.area, "Unassigned")).all()

    # by SKU
    prod_q = base(
        db.query(func.coalesce(VasySalesItem.product_name, VasySalesItem.item_code, "Unknown"),
                 func.coalesce(func.sum(signed_qty), 0),
                 func.coalesce(func.sum(VasySalesItem.net_amount), 0))
        .filter(sold)
    ).group_by(func.coalesce(VasySalesItem.product_name, VasySalesItem.item_code, "Unknown")).all()

    # internal (₹0) kg in the same window — context, not sales
    internal_kg = float(base(
        db.query(func.coalesce(func.sum(signed_qty), 0))
        .filter(VasySalesItem.net_amount == 0)
    ).scalar() or 0)

    areas = [{"name": a, "kg": round(float(kg), 1), "kg_fmt": fmt_kg(float(kg)),
              "value": round(float(v), 2), "value_fmt": fmt_inr(float(v)),
              "invoices": int(n)} for a, kg, v, n in area_q]
    areas.sort(key=lambda r: r["kg"], reverse=True)

    products = [{"name": p, "kg": round(float(kg), 1), "kg_fmt": fmt_kg(float(kg)),
                 "value": round(float(v), 2), "value_fmt": fmt_inr(float(v))}
                for p, kg, v in prod_q]
    products.sort(key=lambda r: r["kg"], reverse=True)

    total_kg = sum(r["kg"] for r in areas)
    total_value = sum(r["value"] for r in areas)
    return {
        "window": window, "window_label": label,
        "areas": areas, "products": products,
        "total_kg": round(total_kg, 1), "total_kg_fmt": fmt_kg(total_kg),
        "total_value": round(total_value, 2), "total_value_fmt": fmt_inr(total_value),
        "invoices": sum(r["invoices"] for r in areas),
        "internal_kg": round(internal_kg, 1), "internal_kg_fmt": fmt_kg(internal_kg),
        "has_data": bool(areas),
    }


def volume_breakdown(db: Session, today: date, window: str, dim: str, key: str) -> dict:
    """Hotels (customers) behind ONE product SKU or ONE route/area, for the
    same day-level window as `volume_report`. Drill-down for the Team & area
    volume tables: dim='product' → hotels that bought that SKU; dim='area' →
    hotels in that route. Grouped by the billed party name (Vasy register
    truth), ₹0 internal transfers excluded, Sales Returns netted."""
    from sqlalchemy import case
    from orderr_core.models.vasy_sales_item import VasySalesItem

    start, end, label, window = _volume_window(today, window)

    signed_qty = case((VasySalesItem.sale_type == "Sales Return", -VasySalesItem.qty),
                      else_=VasySalesItem.qty)
    sold = VasySalesItem.net_amount != 0
    hotel = func.coalesce(VasySalesItem.party_name, Customer.restaurant_name, "Unknown")

    q = (db.query(hotel,
                  func.coalesce(func.sum(signed_qty), 0),
                  func.coalesce(func.sum(VasySalesItem.net_amount), 0),
                  func.count(func.distinct(VasySalesItem.voucher_no)))
         .outerjoin(Customer, Customer.id == VasySalesItem.customer_id)
         .filter(VasySalesItem.invoice_date != None,          # noqa: E711
                 VasySalesItem.invoice_date <= end, sold))
    if start:
        q = q.filter(VasySalesItem.invoice_date >= start)

    if dim == "product":
        prod = func.coalesce(VasySalesItem.product_name, VasySalesItem.item_code, "Unknown")
        q = q.filter(prod == key)
    else:  # area / route
        q = q.filter(func.coalesce(Customer.area, "Unassigned") == key)

    rows = [{"name": h, "kg": round(float(kg), 1), "kg_fmt": fmt_kg(float(kg)),
             "value": round(float(v), 2), "value_fmt": fmt_inr(float(v)),
             "invoices": int(n)}
            for h, kg, v, n in q.group_by(hotel).all()]
    rows.sort(key=lambda r: r["kg"], reverse=True)

    total_kg = sum(r["kg"] for r in rows)
    total_value = sum(r["value"] for r in rows)
    return {
        "dim": dim, "key": key,
        "window": window, "window_label": label,
        "rows": rows,
        "total_kg_fmt": fmt_kg(total_kg),
        "total_value_fmt": fmt_inr(total_value),
        "hotel_count": len(rows),
        "has_data": bool(rows),
    }


# ── P3-8 Demand forecast (per SKU, for production planning) ────────────────

def demand_forecast(db: Session, today: date, lookback: int = 28) -> dict:
    """P3-8 — per-SKU demand forecast for production planning.

    From ordered quantities (parsed order items). Baseline avg daily demand
    over the last `lookback` days vs the last 7 days; next-day forecast =
    recent daily rate, next-week = ×7. Trend = recent vs baseline. Simple,
    explainable moving-average — good enough to plan daily production.
    """
    start = today - timedelta(days=lookback - 1)
    start7 = today - timedelta(days=6)
    ss, ss7, ts = start.strftime("%Y-%m-%d"), start7.strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")

    rows = db.query(Order.business_date, Order.parsed_items).filter(
        Order.is_cancelled == False,                        # noqa: E712
        Order.business_date != None,                        # noqa: E711
        Order.business_date >= ss, Order.business_date <= ts).all()

    agg = {}  # erp_name → {full, last7, unit_votes}
    for bdate, parsed in rows:
        in7 = bdate >= ss7
        for it in safe_list(parsed):
            if not isinstance(it, dict):
                continue
            name = erp_display_name(it.get("product", "") or "Unknown")
            unit = (it.get("unit", "kg") or "kg").lower()
            try:
                qty = float(it.get("quantity") or 0)
            except (TypeError, ValueError):
                qty = 0
            a = agg.setdefault(name, {"full": 0.0, "last7": 0.0, "units": {}})
            a["full"] += qty
            if in7:
                a["last7"] += qty
            a["units"][unit] = a["units"].get(unit, 0) + 1

    out = []
    for name, a in agg.items():
        avg_daily = a["full"] / lookback
        recent_daily = a["last7"] / 7
        trend = None
        if avg_daily > 0:
            trend = round((recent_daily - avg_daily) / avg_daily * 100, 1)
        unit = max(a["units"].items(), key=lambda kv: kv[1])[0] if a["units"] else "kg"
        out.append({
            "product": name, "unit": unit,
            "avg_daily": round(avg_daily, 1), "avg_daily_fmt": fmt_qty(round(avg_daily, 1)),
            "recent_daily": round(recent_daily, 1), "recent_daily_fmt": fmt_qty(round(recent_daily, 1)),
            "forecast_next_day": round(recent_daily, 1), "forecast_next_day_fmt": fmt_qty(round(recent_daily, 1)),
            "forecast_next_week": round(recent_daily * 7, 1), "forecast_next_week_fmt": fmt_qty(round(recent_daily * 7, 1)),
            "trend_pct": trend,
        })
    out.sort(key=lambda r: r["forecast_next_week"], reverse=True)
    return {"lookback": lookback, "rows": out, "has_data": bool(out)}


# ── P1-9 Fill rate (ordered vs delivered) ──────────────────────────────────

def fill_rate(db: Session, today: date, days=90) -> dict:
    """P1-9 — delivered vs ordered quantity, from OrderItemActual.

    Fill % = Σ actual / Σ ordered, per SKU and overall, over rows where an
    actual was captured AND the actual unit matches the ordered unit (mixed
    units aren't summable). `days` bounds by the order's business_date;
    None = all time. Rows without an actual are counted as `pending`
    (awaiting weigh-in), not as a shortfall.
    """
    window_start = None if days is None else today - timedelta(days=days - 1)

    q = (
        db.query(OrderItemActual.product, OrderItemActual.ordered_quantity,
                 OrderItemActual.ordered_unit, OrderItemActual.actual_quantity,
                 OrderItemActual.actual_unit)
        .join(Order, OrderItemActual.order_id == Order.id)
        .filter(Order.is_cancelled == False,          # noqa: E712
                Order.business_date <= today.strftime("%Y-%m-%d"))
    )
    if window_start:
        q = q.filter(Order.business_date >= window_start.strftime("%Y-%m-%d"))

    agg = {}  # erp_name → {ordered, actual, unit, lines, pending}
    pending_total = 0
    for product, oq, ou, aq, au in q.all():
        name = erp_display_name(product or "Unknown")
        unit = (ou or "kg").lower()
        row = agg.setdefault(name, {"ordered": 0.0, "actual": 0.0, "unit": unit,
                                    "lines": 0, "pending": 0})
        row["lines"] += 1
        if aq is None:
            row["pending"] += 1
            pending_total += 1
            continue
        if (au or unit).lower() != unit:
            # unit mismatch — can't sum; skip from fill math (rare)
            continue
        row["ordered"] += float(oq or 0)
        row["actual"] += float(aq or 0)

    rows = []
    tot_ordered = tot_actual = 0.0
    for name, r in agg.items():
        fill = round(r["actual"] / r["ordered"] * 100, 1) if r["ordered"] else None
        tot_ordered += r["ordered"]
        tot_actual += r["actual"]
        rows.append({
            "product": name, "unit": r["unit"],
            "ordered": round(r["ordered"], 3), "ordered_fmt": fmt_qty(r["ordered"]),
            "actual": round(r["actual"], 3), "actual_fmt": fmt_qty(r["actual"]),
            "fill_pct": fill, "lines": r["lines"], "pending": r["pending"],
        })
    rows.sort(key=lambda x: x["ordered"], reverse=True)

    overall = round(tot_actual / tot_ordered * 100, 1) if tot_ordered else None
    return {
        "rows": rows,
        "overall_fill_pct": overall,
        "ordered_fmt": fmt_qty(tot_ordered),
        "actual_fmt": fmt_qty(tot_actual),
        "pending": pending_total,
        "has_data": bool(rows),
        "days": days,
        "window_label": "All time" if days is None else f"Last {days} days",
    }


# ── P1-10 Parse-quality monitor (unclear rate over time) ───────────────────

def parse_quality(db: Session, today: date, days: int = 30) -> dict:
    """P1-10 — daily unclear-order rate (data-hygiene signal).

    % unclear = orders flagged is_unclear / total orders, per business_date,
    over the last `days` days. Cancelled orders excluded.
    """
    start = today - timedelta(days=days - 1)
    rows = (
        db.query(Order.business_date, Order.is_unclear)
        .filter(Order.is_cancelled == False,        # noqa: E712
                Order.business_date != None,        # noqa: E711
                Order.business_date >= start.strftime("%Y-%m-%d"),
                Order.business_date <= today.strftime("%Y-%m-%d"))
        .all()
    )
    by_day = {}  # 'YYYY-MM-DD' → [total, unclear]
    for bdate, unclear in rows:
        d = by_day.setdefault(bdate, [0, 0])
        d[0] += 1
        if unclear:
            d[1] += 1

    series = []
    d = start
    tot = unc = 0
    while d <= today:
        ds = d.strftime("%Y-%m-%d")
        t, u = by_day.get(ds, [0, 0])
        tot += t
        unc += u
        series.append({
            "date": ds, "label": d.strftime("%d %b"),
            "total": t, "unclear": u,
            "pct": round(u / t * 100, 1) if t else None,
        })
        d += timedelta(days=1)

    return {
        "series": series,
        "total_orders": tot,
        "total_unclear": unc,
        "overall_pct": round(unc / tot * 100, 1) if tot else None,
        "days": days,
    }
