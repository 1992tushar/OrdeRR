import io
from decimal import Decimal, InvalidOperation

from sqlalchemy import func
from sqlalchemy.orm import Session

from orderr_core.models.customer import Customer


def normalize_phone(phone: str) -> str:
    """
    Normalize phone number to E.164 format without leading +.
    Handles numbers with/without country code.

    Examples:
        9876543210      → 919876543210
        919876543210    → 919876543210  (already correct, not double-prefixed)
        +919876543210   → 919876543210
    """

    # Strip whitespace, dashes, plus sign
    phone = (
        phone
        .replace("+", "")
        .replace(" ", "")
        .replace("-", "")
        .strip()
    )

    # Only add 91 prefix if the number is exactly 10 digits (raw Indian mobile)
    # This avoids the bug where 911234567890 would incorrectly get prefixed again
  


    if len(phone) == 10:
        phone = f"91{phone}"   # always prefix 10-digit numbers, no startswith check
    return phone


    return phone


def validate_phone(phone: str) -> str | None:
    """
    Validate a phone number BEFORE normalizing/storing it.
    Accepts either:
        - a bare 10-digit Indian mobile number (e.g. 9876543210), or
        - an already-prefixed 12-digit number starting with 91 (e.g. 919876543210)
    Also accepts a leading +, spaces, or dashes, which are stripped first.

    Returns an error message string if invalid, or None if valid.
    """
    if not phone or not phone.strip():
        return "Phone number cannot be empty."

    cleaned = (
        phone
        .replace("+", "")
        .replace(" ", "")
        .replace("-", "")
        .strip()
    )

    if not cleaned.isdigit():
        return f"Phone number '{phone}' contains invalid characters. Use digits only (e.g. 9876543210)."

    if len(cleaned) == 10:
        if cleaned[0] not in "6789":
            return f"'{phone}' doesn't look like a valid Indian mobile number (should start with 6-9)."
        return None

    if len(cleaned) == 12:
        if not cleaned.startswith("91"):
            return f"'{phone}' is 12 digits but doesn't start with 91 (India country code)."
        if cleaned[2] not in "6789":
            return f"'{phone}' doesn't look like a valid Indian mobile number after the 91 prefix."
        return None

    return (
        f"'{phone}' should be a 10-digit mobile number (e.g. 9876543210), "
        f"got {len(cleaned)} digits."
    )

def get_customer_by_phone(
    db: Session,
    phone: str
):
    normalized_phone = normalize_phone(phone)

    return db.query(Customer).filter(
        Customer.phone_number == normalized_phone
    ).first()


def create_new_customer(
    db: Session,
    phone: str
):
    normalized_phone = normalize_phone(phone)

    customer = Customer(
        phone_number=normalized_phone,
        onboarding_status="awaiting_name"
    )

    db.add(customer)
    db.commit()
    db.refresh(customer)

    return customer


def create_customer_manually(
    db: Session,
    phone: str,
    restaurant_name: str,
    area: str = None,
    salesperson_id: int = None,
) -> Customer:
    """
    Create a customer record directly (no onboarding flow).
    Used by dashboard Add Customer form and manager WhatsApp command.
    Raises ValueError if phone already exists.
    """

    error = validate_phone(phone)
    if error:
        raise ValueError(error)
        
    normalized = normalize_phone(phone)

    existing = db.query(Customer).filter(Customer.phone_number == normalized).first()
    if existing:
        raise ValueError(f"Customer with phone {normalized} already exists.")

    customer = Customer(
        phone_number=normalized,
        restaurant_name=restaurant_name.strip(),
        area=area.strip() if area else None,
        salesperson_id=salesperson_id,
        is_daily_order_customer=True,
        onboarding_status="active",   # skip onboarding since manager added them
        is_active=True,
    )
    db.add(customer)
    db.commit()
    db.refresh(customer)
    return customer


# ── Bulk import from the "Customer Outstanding" spreadsheet ──────────────────

def _to_amount(value) -> Decimal:
    """Coerce a spreadsheet cell into a Decimal amount; blanks/junk → 0.

    Handles the Vasy quirk where credit/negative balances export as STRINGS with
    a non-breaking space after the sign and comma grouping, e.g.
    '-\\xa08,29,681.00'. All whitespace (incl. \\xa0) and commas are stripped
    before parsing so those don't silently become 0.
    """
    if value is None:
        return Decimal("0")
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return Decimal("0")
    import re
    s = re.sub(r"[\s ]", "", str(value)).replace(",", "")
    if not s:
        return Decimal("0")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _find_header_and_columns(ws):
    """
    Locate the header row and map the columns we care about by their labels,
    so the import is resilient to column re-ordering in the ERP export.

    Returns (header_row_index, {"name": idx, "phone": idx|None, "closing": idx|None}).
    Raises ValueError if a Party Name column can't be found.
    """
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
        labels = [str(c).strip().lower() if c is not None else "" for c in row]
        name_col = phone_col = closing_col = None
        for i, lbl in enumerate(labels):
            if lbl in ("party name", "name", "customer name", "party"):
                name_col = i
            elif "contact" in lbl or "phone" in lbl or "mobile" in lbl:
                phone_col = i
            elif lbl in ("closing", "closing balance", "outstanding", "balance"):
                closing_col = i
        if name_col is not None:
            return r_idx, {"name": name_col, "phone": phone_col, "closing": closing_col}
    raise ValueError(
        "Could not find a 'Party Name' column. Expected a header row with "
        "columns like 'Party Name', 'Contact No.' and 'Closing'."
    )


def import_customers_from_xlsx(db: Session, file_bytes: bytes) -> dict:
    """
    Import / update customers from a "Customer Outstanding" .xlsx export.

    Upsert rules:
      • Row with a valid phone → matched by normalized phone number.
      • Row without a phone (or an invalid one) → matched by restaurant name
        (case-insensitive). Stored with phone_number = NULL → flagged RED.
      • Existing match → outstanding is refreshed; a missing phone is filled in.
      • No match → a new customer is created (onboarding_status='active',
        is_daily_order_customer=False so bulk-imported receivables customers are
        NOT auto-chased by the daily-order reminder jobs).

    Returns a summary dict: created / updated / skipped / no_phone counts +
    per-row issues.
    """
    import openpyxl  # local import — only needed for imports, keeps startup lean

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"Could not read the Excel file: {e}")

    ws = wb.active
    header_idx, colmap = _find_header_and_columns(ws)
    name_col = colmap["name"]
    phone_col = colmap["phone"]
    closing_col = colmap["closing"]

    created = updated = skipped = no_phone = 0
    issues: list[str] = []
    seen_keys: set = set()   # guard against duplicate rows within one file

    # Manual aliases outrank phone/name matching — an alias is an explicit
    # human decision about which customer a party name IS. Needed because two
    # real outlets can share one contact number in Vasy (SANTOSH MAGGIE:
    # KHANDALA + Lonvla outlets, same owner phone) — phone-first matching
    # would rename/merge them into one customer. Local imports: vasy_import
    # imports this module at load time, so importing it at module level here
    # would be circular.
    from orderr_core.models.customer_alias import CustomerAlias
    from orderr_core.services.vasy_import import normalize_name
    alias_map = {a.alias_key: a.customer_id for a in db.query(CustomerAlias).all()}

    rows = ws.iter_rows(min_row=header_idx + 2, values_only=True)
    for row in rows:
        if not row:
            continue
        name = row[name_col] if name_col < len(row) else None
        name = (str(name).strip() if name is not None else "")
        if not name:
            continue  # skip blank / separator rows silently

        raw_phone = None
        if phone_col is not None and phone_col < len(row):
            raw_phone = row[phone_col]
        raw_phone = (str(raw_phone).strip() if raw_phone not in (None, "") else "")

        outstanding = _to_amount(row[closing_col]) if (closing_col is not None and closing_col < len(row)) else Decimal("0")

        # Resolve phone → normalized or None
        normalized = None
        if raw_phone:
            err = validate_phone(raw_phone)
            if err:
                issues.append(f"'{name}': {err} — imported without a phone number.")
            else:
                normalized = normalize_phone(raw_phone)

        # Duplicate-within-file guard
        key = normalized or f"name:{name.lower()}"
        if key in seen_keys:
            issues.append(f"'{name}': duplicate row in file — skipped.")
            skipped += 1
            continue
        seen_keys.add(key)

        # ── Find an existing customer to update ────────────────────────────
        # alias first (explicit human decision), then phone, then name
        existing = None
        alias_cid = alias_map.get(normalize_name(name))
        if alias_cid:
            existing = db.query(Customer).filter(Customer.id == alias_cid).first()
        if existing is None and normalized:
            existing = db.query(Customer).filter(
                Customer.phone_number == normalized
            ).first()
        if existing is None:
            # match by name (covers phone-less rows and pre-existing name-only records)
            existing = db.query(Customer).filter(
                func.lower(Customer.restaurant_name) == name.lower()
            ).first()

        if existing:
            existing.outstanding = outstanding
            if normalized and not existing.phone_number:
                # backfill only if no OTHER customer owns this number — two
                # outlets can share a contact in Vasy, and stealing the phone
                # here would make phone-matching nondeterministic everywhere
                phone_taken = db.query(Customer.id).filter(
                    Customer.phone_number == normalized,
                    Customer.id != existing.id,
                ).first() is not None
                if not phone_taken:
                    existing.phone_number = normalized
            # Refresh the name from Vasy (source of truth). Previously only set
            # when empty, which meant corrected Vasy names never overwrote an
            # existing (possibly glued Name+Company) customer name. Vasy is
            # authoritative for the party name, so keep OrdeRR in sync — EXCEPT
            # when we matched via an alias: the alias name is the known-wrong
            # variant, and renaming the canonical customer to it would undo the
            # data-health merge that created the alias.
            if name and not alias_cid:
                existing.restaurant_name = name
            updated += 1
        else:
            db.add(Customer(
                restaurant_name=name,
                phone_number=normalized,
                outstanding=outstanding,
                onboarding_status="active",
                is_active=True,
                is_daily_order_customer=False,
            ))
            created += 1

        if not normalized:
            no_phone += 1

    db.commit()
    wb.close()

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "no_phone": no_phone,
        "total_processed": created + updated,
        "issues": issues[:50],   # cap so a totally-wrong file doesn't flood the UI
    }
